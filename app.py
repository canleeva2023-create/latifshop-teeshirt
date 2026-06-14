import streamlit as st
import streamlit.components.v1 as components
from PIL import Image
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import base64
import pandas as pd
import json
import os
import zipfile
import urllib.parse
from datetime import datetime, timedelta
from collections import defaultdict
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Latif Shop — Inventaire", page_icon="👕", layout="wide", initial_sidebar_state="collapsed")

SAVE_FILE        = "inventaire_historique.json"
STOCK_FILE       = "stock.json"
VENTES_FILE      = "ventes.json"
PROFILS_FILE     = "profils.json"
ARRIVAGES_FILE   = "arrivages.json"
CATEGORIES_FILE          = "categories.json"
ALIMENTATION_PENDING_FILE = "alimentation_pending.json"
ALIMENTATION_HIST_FILE    = "alimentation_historique.json"
ALL_DATA_FILES   = [SAVE_FILE, STOCK_FILE, VENTES_FILE, PROFILS_FILE,
                    ARRIVAGES_FILE, CATEGORIES_FILE,
                    ALIMENTATION_PENDING_FILE, ALIMENTATION_HIST_FILE]

# ══════════════════════════════════════════════════════════════════════════════
# PALETTE COULEURS
# ══════════════════════════════════════════════════════════════════════════════
COLOR_PALETTE = [
    {"name":"Blanc","hex":"#FFFFFF","text":"#111"},{"name":"Crème","hex":"#F5F0E8","text":"#111"},
    {"name":"Beige","hex":"#D4B896","text":"#111"},{"name":"Gris clair","hex":"#D0D0D0","text":"#111"},
    {"name":"Gris","hex":"#888888","text":"#FFF"},{"name":"Gris foncé","hex":"#444444","text":"#FFF"},
    {"name":"Anthracite","hex":"#2B2B2B","text":"#FFF"},{"name":"Noir","hex":"#111111","text":"#FFF"},
    {"name":"Camel","hex":"#C19A6B","text":"#FFF"},{"name":"Marron","hex":"#7B4F2E","text":"#FFF"},
    {"name":"Chocolat","hex":"#3E1C00","text":"#FFF"},{"name":"Corail","hex":"#FF6B5B","text":"#FFF"},
    {"name":"Saumon","hex":"#FA8072","text":"#FFF"},{"name":"Orange","hex":"#FF7F2A","text":"#FFF"},
    {"name":"Moutarde","hex":"#D4A017","text":"#FFF"},{"name":"Jaune","hex":"#FFD700","text":"#111"},
    {"name":"Rouge","hex":"#CC0000","text":"#FFF"},{"name":"Bordeaux","hex":"#6E0E0A","text":"#FFF"},
    {"name":"Rose pâle","hex":"#FFB6C1","text":"#111"},{"name":"Rose","hex":"#FF69B4","text":"#FFF"},
    {"name":"Fuchsia","hex":"#C2006F","text":"#FFF"},{"name":"Menthe","hex":"#98D4B0","text":"#111"},
    {"name":"Vert clair","hex":"#5CB85C","text":"#FFF"},{"name":"Vert","hex":"#228B22","text":"#FFF"},
    {"name":"Vert forêt","hex":"#1A4A1A","text":"#FFF"},{"name":"Kaki","hex":"#7B7440","text":"#FFF"},
    {"name":"Olive","hex":"#556B2F","text":"#FFF"},{"name":"Bleu ciel","hex":"#87CEEB","text":"#111"},
    {"name":"Turquoise","hex":"#00CED1","text":"#FFF"},{"name":"Bleu","hex":"#1565C0","text":"#FFF"},
    {"name":"Bleu marine","hex":"#002366","text":"#FFF"},{"name":"Navy","hex":"#001A33","text":"#FFF"},
    {"name":"Jean clair","hex":"#A8C4D4","text":"#111"},{"name":"Jean","hex":"#4A7FA5","text":"#FFF"},
    {"name":"Jean foncé","hex":"#1E3A5F","text":"#FFF"},{"name":"Lavande","hex":"#E6E6FA","text":"#111"},
    {"name":"Mauve","hex":"#C8A2C8","text":"#111"},{"name":"Violet","hex":"#6A0DAD","text":"#FFF"},
    {"name":"Prune","hex":"#4B0040","text":"#FFF"},{"name":"Or","hex":"#CFB53B","text":"#FFF"},
    {"name":"Argent","hex":"#C0C0C0","text":"#111"},{"name":"Bronze","hex":"#8C6239","text":"#FFF"},
]
HEX_MAP = {c["name"]: c["hex"] for c in COLOR_PALETTE}

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for key, val in [
        ("user_name", None), ("user_role", None),
        ("pin_error", False), ("selected_colors", []),
        ("login_space", None),
        # Alimentation flow
        ("ali_step", 0), ("ali_cat", None), ("ali_art_key", None),
        ("ali_qtys", {}), ("ali_livreur", ""), ("ali_pin_error", False),
        # Edition catégories
        ("edit_cat_idx", None),
        # Edition vendeurs
        ("edit_vend_idx", None),
        # Alimentation admin: modification avant approbation
        ("ali_edit_idx", None),
        # Calendrier admin
        ("v_admin_date", None),
        ("ali_admin_date", None),
        ("v_admin_month", None),
        ("ali_admin_month", None),
]:
    if key not in st.session_state:
        st.session_state[key] = val

# ══════════════════════════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* ═══════════════════════════════════════════════
   LATIF SHOP — UI 4K  ·  v4.0  MODE CLAIR
   Police : Inter + Space Grotesk
   Thème  : Blanc & Bleu Marine & Or
═══════════════════════════════════════════════ */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Space+Grotesk:wght@400;600;700&display=swap');

:root{
  --bg:          #F4F6FA;
  --surface:     #FFFFFF;
  --surface2:    #F8F9FC;
  --surface3:    #EEF1F7;
  --border:      #E0E5EF;
  --border2:     #C8D0DF;
  --navy:        #1B2B4B;
  --navy2:       #2C4270;
  --gold:        #C09020;
  --gold2:       #E0B030;
  --gold-soft:   rgba(192,144,32,.10);
  --text:        #1A1F2E;
  --text2:       #4A5568;
  --text3:       #8A9AB5;
  --accent:      #2563EB;
  --green:       #059669;
  --red:         #DC2626;
  --radius:      12px;
  --radius-sm:   8px;
  --shadow:      0 2px 12px rgba(27,43,75,.08);
  --shadow-md:   0 4px 24px rgba(27,43,75,.12);
  --shadow-lg:   0 8px 40px rgba(27,43,75,.18);
}

/* ─── Reset global ─── */
*,*::before,*::after{box-sizing:border-box;}
html,body,[class*="css"]{
  font-family:'Inter',sans-serif;
  background:var(--bg)!important;
  color:var(--text)!important;
}

/* ─── Page principale ─── */
.stApp{background:var(--bg)!important;}
.main .block-container{
  padding:2rem 2.5rem 4rem!important;
  max-width:1800px!important;
}

/* Masquer sidebar toggle & header Streamlit */
#MainMenu,header[data-testid="stHeader"],
[data-testid="collapsedControl"],
footer{display:none!important;}

/* ─── BANDEAU TITRE ─── */
.app-header{
  display:flex;align-items:center;gap:1.4rem;
  background:linear-gradient(135deg,var(--navy) 0%,var(--navy2) 100%);
  border:none;
  border-radius:var(--radius);
  padding:1.4rem 2rem;
  margin-bottom:1.8rem;
  box-shadow:var(--shadow-md);
  position:relative;overflow:hidden;
}
.app-header::before{
  content:'';position:absolute;top:-80px;right:-80px;
  width:260px;height:260px;border-radius:50%;
  background:radial-gradient(circle,rgba(224,176,48,.22) 0%,transparent 70%);
  pointer-events:none;
}
.app-header::after{
  content:'';position:absolute;bottom:-40px;left:180px;
  width:140px;height:140px;border-radius:50%;
  background:radial-gradient(circle,rgba(255,255,255,.06) 0%,transparent 70%);
  pointer-events:none;
}
.app-logo{font-size:2.6rem;line-height:1;filter:drop-shadow(0 2px 4px rgba(0,0,0,.3));}
.app-title{
  font-family:'Space Grotesk',sans-serif;
  font-size:clamp(1.3rem,3vw,2.1rem);
  font-weight:700;color:#FFFFFF;
  letter-spacing:.03em;line-height:1.1;
}
.app-sub{
  font-size:.75rem;color:rgba(255,255,255,.6);
  letter-spacing:.18em;text-transform:uppercase;margin-top:.2rem;
}
.app-badge{
  margin-left:auto;
  background:rgba(255,255,255,.15);
  border:1px solid rgba(255,255,255,.3);
  color:#FFF;font-size:.68rem;font-weight:600;
  letter-spacing:.12em;text-transform:uppercase;
  padding:.4rem 1rem;border-radius:100px;white-space:nowrap;
}

/* ─── ONGLETS ─── */
.stTabs [data-baseweb="tab-list"]{
  background:var(--surface)!important;
  border:1px solid var(--border)!important;
  border-radius:var(--radius)!important;
  padding:.35rem .4rem!important;
  gap:.3rem!important;
  box-shadow:var(--shadow)!important;
  margin-bottom:1.6rem!important;
}
.stTabs [role="tab"]{
  background:transparent!important;
  color:var(--text2)!important;
  border-radius:var(--radius-sm)!important;
  border:none!important;
  font-size:.82rem!important;font-weight:500!important;
  letter-spacing:.04em!important;
  padding:.65rem 1.3rem!important;
  transition:all .2s!important;
}
.stTabs [role="tab"]:hover{
  background:var(--surface3)!important;
  color:var(--navy)!important;
}
.stTabs [aria-selected="true"]{
  background:linear-gradient(135deg,var(--navy) 0%,var(--navy2) 100%)!important;
  color:#FFF!important;
  font-weight:700!important;
  box-shadow:0 3px 12px rgba(27,43,75,.25)!important;
}
.stTabs [data-baseweb="tab-highlight"]{display:none!important;}
.stTabs [data-baseweb="tab-border"]{display:none!important;}

/* ─── CARTES / SECTIONS ─── */
.card{
  background:var(--surface);
  border:1px solid var(--border);
  border-radius:var(--radius);
  padding:1.6rem;
  margin-bottom:1.2rem;
  box-shadow:var(--shadow);
}
.card-title{
  font-family:'Space Grotesk',sans-serif;
  font-size:1rem;font-weight:700;color:var(--navy);
  letter-spacing:.06em;text-transform:uppercase;
  display:flex;align-items:center;gap:.6rem;
  margin-bottom:1.1rem;padding-bottom:.8rem;
  border-bottom:2px solid var(--border);
}

/* ─── SECTION LABEL ─── */
.section-label{
  font-size:.68rem;letter-spacing:.2em;text-transform:uppercase;
  color:var(--text3);margin-bottom:8px;display:block;font-weight:600;
}

/* ─── INPUTS ─── */
input,select,textarea{font-size:15px!important;}
.stTextInput>label,.stNumberInput>label,.stSelectbox>label,
.stMultiSelect>label,.stRadio>label,.stDateInput>label,
.stTextArea>label,.stFileUploader>label{
  font-size:.72rem!important;letter-spacing:.14em!important;
  text-transform:uppercase!important;color:var(--text2)!important;
  font-weight:600!important;
}
.stTextInput input,.stNumberInput input,.stTextArea textarea{
  background:#FFF!important;
  border:1.5px solid var(--border2)!important;
  border-radius:var(--radius-sm)!important;
  color:var(--text)!important;
  font-size:15px!important;
  transition:border-color .2s,box-shadow .2s!important;
}
.stTextInput input:focus,.stNumberInput input:focus,.stTextArea textarea:focus{
  border-color:var(--accent)!important;
  box-shadow:0 0 0 3px rgba(37,99,235,.12)!important;
}
[data-baseweb="select"]>div{
  background:#FFF!important;
  border:1.5px solid var(--border2)!important;
  border-radius:var(--radius-sm)!important;
  color:var(--text)!important;
}
[data-baseweb="select"] [role="option"]:hover{background:var(--surface3)!important;}
[data-baseweb="tag"]{
  background:rgba(37,99,235,.1)!important;
  border:1px solid var(--accent)!important;
  color:var(--accent)!important;
  border-radius:6px!important;
}
[data-baseweb="select"] svg{color:var(--text3)!important;}

/* Radio */
.stRadio [data-testid="stMarkdownContainer"] p{
  color:var(--text2)!important;font-size:.85rem!important;
}

/* ─── BOUTONS principaux ─── */
.stButton>button{
  background:linear-gradient(135deg,var(--navy) 0%,var(--navy2) 100%)!important;
  color:#FFF!important;border:none!important;
  border-radius:var(--radius-sm)!important;
  padding:.75rem 1.6rem!important;
  font-size:.8rem!important;letter-spacing:.1em!important;
  text-transform:uppercase!important;font-weight:700!important;
  width:100%!important;min-height:48px!important;
  box-shadow:0 3px 12px rgba(27,43,75,.22)!important;
  transition:all .2s!important;
}
.stButton>button:hover{
  background:linear-gradient(135deg,var(--navy2) 0%,#3B5998 100%)!important;
  box-shadow:0 5px 20px rgba(27,43,75,.35)!important;
  transform:translateY(-1px);
}
.stButton>button:active{transform:translateY(0)!important;}

.stDownloadButton>button{
  background:#FFF!important;color:var(--navy)!important;
  border:1.5px solid var(--navy)!important;
  border-radius:var(--radius-sm)!important;
  padding:.75rem 1.6rem!important;font-size:.8rem!important;
  letter-spacing:.1em!important;text-transform:uppercase!important;
  font-weight:600!important;width:100%!important;min-height:48px!important;
  transition:all .2s!important;box-shadow:none!important;
}
.stDownloadButton>button:hover{
  background:var(--navy)!important;color:#FFF!important;
}

/* ─── FILE UPLOADER ─── */
.stFileUploader{
  background:#FFF!important;
  border:2px dashed var(--border2)!important;
  border-radius:var(--radius)!important;
  padding:1rem!important;
  transition:border-color .2s!important;
}
.stFileUploader:hover{border-color:var(--accent)!important;}
.stFileUploader [data-testid="stFileUploaderDropzone"]{background:transparent!important;}
.stFileUploader [data-testid="stFileUploaderDropzone"] p,
.stFileUploader [data-testid="stFileUploaderDropzone"] small,
.stFileUploader [data-testid="stFileUploaderDropzone"] span{
  color:var(--text3)!important;
}

/* ─── SÉPARATEUR ─── */
hr{border:none;border-top:1.5px solid var(--border);margin:1.4rem 0;}

/* ─── GALERIE PHOTOS ─── */
.photo-gallery{display:flex;flex-wrap:wrap;gap:10px;margin-top:12px;}
.photo-thumb{width:80px;}
.photo-thumb img{
  width:80px;height:100px;object-fit:cover;
  border-radius:var(--radius-sm);
  border:2px solid var(--border);
  transition:border-color .2s,box-shadow .2s;
}
.photo-thumb img:hover{border-color:var(--accent);box-shadow:0 0 0 3px rgba(37,99,235,.15);}
.photo-thumb .photo-name{
  font-size:.52rem;color:var(--text3);text-align:center;
  overflow:hidden;white-space:nowrap;text-overflow:ellipsis;margin-top:4px;
}
.photo-count-badge{
  display:inline-flex;align-items:center;
  background:rgba(37,99,235,.1);border:1px solid var(--accent);
  color:var(--accent);font-size:.68rem;padding:3px 12px;
  letter-spacing:.1em;border-radius:100px;margin-left:8px;font-weight:600;
}

/* ─── COULEURS ─── */
.selected-colors-bar{
  display:flex;flex-wrap:wrap;gap:8px;margin:8px 0;
  min-height:44px;padding:10px 12px;
  background:var(--surface2);border:1.5px solid var(--border);
  border-radius:var(--radius-sm);
}
.color-tag{
  display:inline-flex;align-items:center;gap:6px;
  padding:5px 12px;font-size:.74rem;font-weight:500;
  letter-spacing:.04em;
  background:#FFF;border:1px solid var(--border2);
  border-radius:100px;color:var(--text);
  box-shadow:0 1px 4px rgba(0,0,0,.06);
}
.color-dot{
  width:12px;height:12px;border-radius:50%;
  display:inline-block;border:1.5px solid rgba(0,0,0,.12);
  flex-shrink:0;
}

/* ─── TABLEAU ─── */
.table-wrapper{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;}
table{
  border-collapse:collapse;min-width:460px;
  font-size:.84rem;background:#FFF;
  border-radius:var(--radius);overflow:hidden;
  box-shadow:var(--shadow-md);margin-bottom:1.5rem;width:100%;
  border:1px solid var(--border);
}
th{
  background:linear-gradient(135deg,var(--navy) 0%,var(--navy2) 100%);
  color:#FFF;padding:14px 16px;text-align:center;
  font-weight:600;letter-spacing:.1em;text-transform:uppercase;
  font-size:.72rem;white-space:nowrap;
}
td{
  padding:12px 14px;border:1px solid var(--border);
  text-align:center;color:var(--text);min-width:56px;height:54px;
}
tr:nth-child(even) td{background:var(--surface2);}
tr:hover td{background:#EEF4FF!important;}

/* ─── CARTES HISTORIQUE ─── */
.history-card{
  background:#FFF;border:1px solid var(--border);
  border-left:4px solid var(--navy);
  padding:16px 18px;margin-bottom:12px;
  border-radius:var(--radius);box-shadow:var(--shadow);
  transition:box-shadow .2s,border-left-color .2s;
}
.history-card:hover{
  box-shadow:var(--shadow-md);
  border-left-color:var(--accent);
}
.history-date{font-size:.68rem;color:var(--text3);letter-spacing:.1em;}
.history-title{font-weight:700;color:var(--navy);font-size:.95rem;margin:4px 0;}
.history-meta{color:var(--text2);font-size:.8rem;}

/* ─── TAGS ─── */
.tag{
  display:inline-block;background:var(--surface3);color:var(--text2);
  font-size:.68rem;padding:4px 10px;margin:2px;
  letter-spacing:.05em;border-radius:var(--radius-sm);
  border:1px solid var(--border);
}
.tag-dark{
  background:var(--navy)!important;color:#FFF!important;
  border-color:var(--navy)!important;
}

/* ─── AVATAR PROFIL ─── */
.profil-avatar{
  width:64px;height:64px;border-radius:50%;
  background:linear-gradient(135deg,var(--navy) 0%,var(--navy2) 100%);
  color:#FFF;font-size:1.3rem;font-weight:700;
  display:flex;align-items:center;justify-content:center;
  margin:0 auto 8px auto;
  box-shadow:0 4px 16px rgba(27,43,75,.25);
}

/* ─── MÉTRIQUES STREAMLIT ─── */
[data-testid="metric-container"]{
  background:#FFF!important;
  border:1px solid var(--border)!important;
  border-top:3px solid var(--navy)!important;
  border-radius:var(--radius)!important;
  padding:1.1rem 1.4rem!important;
  box-shadow:var(--shadow)!important;
}
[data-testid="metric-container"] [data-testid="stMetricLabel"]{
  color:var(--text3)!important;font-size:.72rem!important;
  letter-spacing:.12em!important;text-transform:uppercase!important;font-weight:600!important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"]{
  color:var(--navy)!important;font-size:2rem!important;font-weight:700!important;
}
[data-testid="metric-container"] [data-testid="stMetricDelta"]{
  color:var(--green)!important;font-size:.8rem!important;
}

/* ─── DATAFRAME ─── */
[data-testid="stDataFrame"]{
  border:1px solid var(--border)!important;
  border-radius:var(--radius)!important;
  overflow:hidden!important;
  box-shadow:var(--shadow)!important;
}

/* ─── ALERTES ─── */
[data-testid="stAlert"]{
  border-radius:var(--radius)!important;
  border-left:3px solid var(--accent)!important;
  background:rgba(37,99,235,.06)!important;
  color:var(--text)!important;
}
[data-testid="stAlert"] p{color:var(--text)!important;}

/* ─── CAPTION / INFO ─── */
.stCaptionContainer,[data-testid="stCaptionContainer"]{
  color:var(--text3)!important;font-size:.75rem!important;
}

/* ─── EXPANDER ─── */
.streamlit-expanderHeader{
  background:var(--surface2)!important;
  border:1px solid var(--border)!important;
  border-radius:var(--radius-sm)!important;
  color:var(--text)!important;font-weight:500!important;
}
.streamlit-expanderContent{
  background:#FFF!important;
  border:1px solid var(--border)!important;
  border-top:none!important;
  border-radius:0 0 var(--radius-sm) var(--radius-sm)!important;
}

/* ─── LOGIN SCREEN ─── */
.login-card{
  background:#FFF;border:1px solid var(--border);
  border-radius:20px;padding:2.5rem;width:100%;max-width:640px;
  box-shadow:var(--shadow-lg);text-align:center;
}
.login-logo{font-size:4rem;margin-bottom:.5rem;}
.login-title{
  font-family:'Space Grotesk',sans-serif;font-size:2rem;font-weight:700;
  background:linear-gradient(135deg,var(--navy) 0%,var(--navy2) 100%);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  margin-bottom:.3rem;
}
.login-sub{
  color:var(--text3);font-size:.85rem;letter-spacing:.1em;
  text-transform:uppercase;margin-bottom:2rem;
}

/* ─── SCROLLBAR ─── */
::-webkit-scrollbar{width:6px;height:6px;}
::-webkit-scrollbar-track{background:var(--bg);}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:6px;}
::-webkit-scrollbar-thumb:hover{background:var(--navy);}

/* ─── BOUTONS +/− STOCK ─── */
.stock-pm button{
  background:var(--surface3)!important;
  color:var(--navy)!important;
  border:1.5px solid var(--border2)!important;
  border-radius:6px!important;
  padding:.3rem .5rem!important;
  font-size:1rem!important;font-weight:700!important;
  min-height:36px!important;
  letter-spacing:0!important;text-transform:none!important;
  box-shadow:none!important;width:100%!important;
}
.stock-pm button:hover{
  background:rgba(37,99,235,.1)!important;
  border-color:var(--accent)!important;
  color:var(--accent)!important;
  box-shadow:none!important;transform:none!important;
}

/* ─── RESPONSIVE MOBILE ─── */
@media(max-width:768px){
  .main .block-container{padding:.9rem .9rem 3rem!important;}
  .app-title{font-size:1.2rem!important;}
  .app-badge{display:none;}
  [data-testid="column"]{width:100%!important;flex:1 1 100%!important;}
  .stTabs [role="tab"]{font-size:.72rem!important;padding:.5rem .8rem!important;}
  .login-card{padding:1.5rem!important;}
  .login-title{font-size:1.5rem!important;}
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE SHEETS — SYNCHRONISATION CLOUD
# ══════════════════════════════════════════════════════════════════════════════
try:
    import gspread
    from google.oauth2.service_account import Credentials as _GCredentials
    _GSPREAD_AVAILABLE = True
except ImportError:
    _GSPREAD_AVAILABLE = False

_GS_SCOPES = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive",
]

@st.cache_resource(show_spinner=False)
def _gs_client():
    if not _GSPREAD_AVAILABLE:
        return None
    try:
        creds = _GCredentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=_GS_SCOPES
        )
        return gspread.authorize(creds)
    except Exception:
        return None

def _gs_spreadsheet():
    client = _gs_client()
    if client is None:
        return None
    try:
        return client.open_by_key(st.secrets["sheets"]["spreadsheet_id"])
    except Exception:
        return None

def _gs_load(sheet_name, default):
    ss = _gs_spreadsheet()
    if ss is None:
        return None
    try:
        ws = ss.worksheet(sheet_name)
        # Le stock utilise 1 ligne par modèle (col A = clé, col B = JSON)
        if sheet_name == "stock":
            rows = ws.get_all_values()
            if not rows:
                return default
            result = {}
            for r in rows:
                if len(r) >= 2 and r[0] and r[1]:
                    try:
                        result[r[0]] = json.loads(r[1])
                    except Exception:
                        pass
            return result if result else default
        # Les autres feuilles utilisent A1 = JSON complet
        val = ws.acell("A1").value
        return json.loads(val) if val else default
    except Exception:
        return None

def _gs_save(sheet_name, data):
    ss = _gs_spreadsheet()
    if ss is None:
        return
    try:
        try:
            ws = ss.worksheet(sheet_name)
        except Exception:
            ws = ss.add_worksheet(title=sheet_name, rows=2000, cols=2)
        # Le stock : 1 ligne par modèle (pas de limite de taille)
        if sheet_name == "stock" and isinstance(data, dict):
            rows = []
            for mk, v in data.items():
                entry = {kk: vv for kk, vv in v.items() if kk != "b64_thumb"} \
                        if isinstance(v, dict) else v
                rows.append([mk, json.dumps(entry, ensure_ascii=False)])
            ws.clear()
            if rows:
                ws.update("A1", rows)
            return
        # Autres feuilles : JSON complet en A1 (sans b64_thumb)
        if sheet_name == "inventaire_historique":
            if isinstance(data, list):
                data_gs = [{kk: vv for kk, vv in item.items() if kk != "b64_thumb"}
                           if isinstance(item, dict) else item for item in data]
            else:
                data_gs = data
        else:
            data_gs = data
        ws.update("A1", [[json.dumps(data_gs, ensure_ascii=False)]])
    except Exception:
        pass

def _gs_save_photos(photos_dict):
    """Sauvegarde {model_key: b64_thumb} dans une feuille dédiée (1 ligne par article)."""
    ss = _gs_spreadsheet()
    if ss is None or not photos_dict:
        return
    try:
        try:
            ws = ss.worksheet("_photos")
        except Exception:
            ws = ss.add_worksheet(title="_photos", rows=200, cols=2)
        rows = [[mk, thumb] for mk, thumb in photos_dict.items() if thumb]
        if rows:
            ws.clear()
            ws.update("A1", rows)
    except Exception:
        pass

def _gs_load_photos():
    """Charge {model_key: b64_thumb} depuis la feuille dédiée."""
    ss = _gs_spreadsheet()
    if ss is None:
        return {}
    try:
        ws = ss.worksheet("_photos")
        rows = ws.get_all_values()
        return {r[0]: r[1] for r in rows if len(r) >= 2 and r[0] and r[1]}
    except Exception:
        return {}

# ══════════════════════════════════════════════════════════════════════════════
# FONCTIONS I/O
# ══════════════════════════════════════════════════════════════════════════════
def _load(path, default):
    sheet_name = os.path.splitext(os.path.basename(path))[0]
    gs_data = _gs_load(sheet_name, default)
    if gs_data is not None:
        # Fusionner les photos depuis la feuille GSheets _photos (cloud-persistent)
        if sheet_name == "stock":
            gs_photos = _gs_load_photos()
            if gs_photos:
                for mk, entry in gs_data.items():
                    if isinstance(entry, dict) and mk in gs_photos:
                        entry["b64_thumb"] = gs_photos[mk]
            elif os.path.exists(path):
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        local_data = json.load(f)
                    for mk, entry in gs_data.items():
                        if isinstance(entry, dict) and mk in local_data:
                            thumb = local_data[mk].get("b64_thumb", "")
                            if thumb:
                                entry["b64_thumb"] = thumb
                except Exception:
                    pass
        return gs_data
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default

def _save(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    sheet_name = os.path.splitext(os.path.basename(path))[0]
    _gs_save(sheet_name, data)

def load_history():   return _load(SAVE_FILE, [])
def load_stock():     return _load(STOCK_FILE, {})

def save_stock(d):
    _save(STOCK_FILE, d)
    # Sauvegarder les photos dans une feuille GSheets dédiée
    photos = {mk: v["b64_thumb"] for mk, v in d.items()
              if isinstance(v, dict) and v.get("b64_thumb")}
    _gs_save_photos(photos)
def load_ventes():    return _load(VENTES_FILE, [])
def load_arrivages():    return _load(ARRIVAGES_FILE, [])
def load_categories():      return _load(CATEGORIES_FILE, [])
def save_categories(d):     _save(CATEGORIES_FILE, d)
def load_ali_pending():     return _load(ALIMENTATION_PENDING_FILE, [])
def save_ali_pending(d):    _save(ALIMENTATION_PENDING_FILE, d)
def load_ali_hist():        return _load(ALIMENTATION_HIST_FILE, [])
def save_ali_hist(d):       _save(ALIMENTATION_HIST_FILE, d)
def load_profils():
    d = _load(PROFILS_FILE, {"admin_pin": "1234", "vendeurs": []})
    if not os.path.exists(PROFILS_FILE): _save(PROFILS_FILE, d)
    return d

def save_entry(e):
    h = load_history(); h.insert(0, e); _save(SAVE_FILE, h)
def save_ventes(v):   _save(VENTES_FILE, v)
def save_profils(d):  _save(PROFILS_FILE, d)
def save_arrivage(a):
    arr = load_arrivages(); arr.insert(0, a); _save(ARRIVAGES_FILE, arr)

# ── Backup ZIP (mode hors-ligne) ──────────────────────────────────────────────
def export_backup_zip() -> bytes:
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in ALL_DATA_FILES:
            if os.path.exists(f):
                zf.write(f, os.path.basename(f))
    return buf.getvalue()

def import_backup_zip(zip_bytes: bytes):
    buf = BytesIO(zip_bytes)
    with zipfile.ZipFile(buf, "r") as zf:
        for name in zf.namelist():
            if name in [os.path.basename(f) for f in ALL_DATA_FILES]:
                zf.extract(name, ".")

# ── Rapport journalier ────────────────────────────────────────────────────────
def build_rapport_journalier() -> str:
    today = datetime.now().strftime("%d/%m/%Y")
    ventes = [v for v in load_ventes() if v.get("date","").startswith(today)]
    stock_data = load_stock()
    orders = get_order_list(stock_data)

    total_pieces = sum(v.get("quantite",1) for v in ventes)
    nb_ventes = len(ventes)

    # Agrégation par modèle
    agg_mod = defaultdict(int)
    for v in ventes: agg_mod[v.get("model_name","")]+=v.get("quantite",1)
    top_mod = sorted(agg_mod.items(), key=lambda x:-x[1])

    # Par vendeur
    agg_vend = defaultdict(int)
    for v in ventes: agg_vend[v.get("vendeur","")]+=v.get("quantite",1)

    lines = [
        f"📊 *RAPPORT JOURNALIER — {today}*",
        f"🕐 Généré à {datetime.now().strftime('%H:%M')}\n",
        f"━━━━━━━━━━━━━━━━━━━━━",
        f"🧾 *VENTES DU JOUR*",
        f"• {nb_ventes} transaction(s) — {total_pieces} pièce(s) vendues",
    ]
    if agg_vend:
        lines.append("\n👥 *Par vendeur :*")
        for vend,qty in sorted(agg_vend.items(),key=lambda x:-x[1]):
            lines.append(f"  • {vend} : {qty} pièce(s)")
    if top_mod:
        lines.append("\n🏆 *Top articles :*")
        for mod,qty in top_mod[:3]:
            lines.append(f"  • {mod} : {qty} pièce(s)")
    if orders:
        ruptures = [o for o in orders if o["En stock"]==0]
        faibles  = [o for o in orders if o["En stock"]>0]
        lines.append(f"\n━━━━━━━━━━━━━━━━━━━━━")
        lines.append(f"⚠️ *ALERTES STOCK*")
        if ruptures:
            lines.append(f"🔴 {len(ruptures)} rupture(s) :")
            for o in ruptures[:5]: lines.append(f"  • {o['Modèle']} {o['Couleur']}/{o['Taille']}")
        if faibles:
            lines.append(f"🟡 {len(faibles)} stock(s) faible(s)")
    else:
        lines.append(f"\n✅ *Stock : aucune alerte*")

    lines.append(f"\n_Rapport généré automatiquement_")
    return "\n".join(lines)

# ── Articles à rotation lente ─────────────────────────────────────────────────
def get_rotation_lente(seuil_jours: int = 14) -> list:
    """Retourne les articles sans vente depuis `seuil_jours` jours."""
    ventes = load_ventes()
    stock_data = load_stock()
    today = datetime.now()
    # Dernière vente par (model_key, couleur, taille)
    last_sale = {}
    for v in ventes:
        key = (v.get("model_key",""), v.get("couleur",""), v.get("taille",""))
        try:
            d = datetime.strptime(v.get("date","").split(" ")[0], "%d/%m/%Y")
            if key not in last_sale or d > last_sale[key]: last_sale[key] = d
        except: pass
    results = []
    for mk, data in stock_data.items():
        for color in data.get("colors",[]):
            for size in data.get("sizes",[]):
                qty = data.get("stock",{}).get(color,{}).get(size,0)
                if qty == 0: continue  # déjà en rupture, géré ailleurs
                key = (mk, color, size)
                if key in last_sale:
                    jours = (today - last_sale[key]).days
                    if jours >= seuil_jours:
                        results.append({"model_key":mk,"Modèle":data.get("model_name",mk),
                                        "Couleur":color,"Taille":size,"En stock":qty,
                                        "Dernière vente":last_sale[key].strftime("%d/%m/%Y"),
                                        "Jours sans vente":jours})
                else:
                    results.append({"model_key":mk,"Modèle":data.get("model_name",mk),
                                    "Couleur":color,"Taille":size,"En stock":qty,
                                    "Dernière vente":"Jamais vendu","Jours sans vente":999})
    return sorted(results, key=lambda x: -x["Jours sans vente"])

def delete_history(idx):
    h = load_history()
    if 0 <= idx < len(h): h.pop(idx); _save(SAVE_FILE, h)

def model_key(name): return (name or "sans_nom").strip().lower().replace(" ", "_")

def save_vente_and_update_stock(vente):
    v = load_ventes(); v.insert(0, vente); _save(VENTES_FILE, v)
    sd = load_stock()
    mk = vente["model_key"]
    if mk in sd:
        c, s, q = vente["couleur"], vente["taille"], vente["quantite"]
        cur = sd[mk]["stock"].get(c, {}).get(s, 0)
        sd[mk]["stock"].setdefault(c, {})[s] = max(0, cur - q)
        sd[mk]["last_update"] = vente["date"]
        save_stock(sd)

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS VISUELS
# ══════════════════════════════════════════════════════════════════════════════
def img_to_b64(fobj, w=70, h=90):
    img = Image.open(fobj); img.thumbnail((w*2, h*2))
    buf = BytesIO(); img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()

def make_thumb(fobj):
    """Miniature très compressée (80×100 JPEG) pour stockage GSheets."""
    img = Image.open(fobj).convert("RGB")
    img.thumbnail((80, 100), Image.LANCZOS)
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=50, optimize=True)
    return base64.b64encode(buf.getvalue()).decode()

def wa_button(url, label="Envoyer sur WhatsApp", height=48):
    svg = '<svg width="18" height="18" viewBox="0 0 32 32" fill="white"><path d="M16 2C8.268 2 2 8.268 2 16c0 2.492.664 4.833 1.822 6.854L2 30l7.338-1.797A13.93 13.93 0 0016 30c7.732 0 14-6.268 14-14S23.732 2 16 2zm0 25.5a11.45 11.45 0 01-5.824-1.587l-.418-.248-4.354 1.067 1.097-4.232-.272-.435A11.46 11.46 0 014.5 16C4.5 9.649 9.649 4.5 16 4.5S27.5 9.649 27.5 16 22.351 27.5 16 27.5zm6.29-8.388c-.344-.172-2.036-1.004-2.352-1.118-.316-.115-.546-.172-.776.172-.23.344-.89 1.118-1.09 1.348-.2.23-.4.258-.744.086-.344-.172-1.452-.535-2.766-1.707-1.022-.912-1.712-2.038-1.912-2.382-.2-.344-.021-.53.15-.701.155-.154.344-.402.516-.603.172-.2.23-.344.344-.574.115-.23.058-.43-.029-.603-.086-.172-.776-1.872-1.064-2.563-.28-.672-.564-.58-.776-.59l-.66-.012c-.23 0-.603.086-.918.43-.316.344-1.205 1.176-1.205 2.868s1.233 3.327 1.405 3.557c.172.23 2.426 3.705 5.878 5.196.822.355 1.463.567 1.963.726.824.263 1.574.226 2.167.137.66-.099 2.036-.832 2.323-1.635.287-.803.287-1.492.2-1.635-.086-.143-.316-.23-.66-.402z"/></svg>'
    components.html(f"""<style>.wab{{display:flex;align-items:center;justify-content:center;gap:8px;
    width:100%;background:#25D366;color:#FFF;padding:10px 0;font-size:.78rem;letter-spacing:.1em;
    text-transform:uppercase;font-weight:600;font-family:Inter,sans-serif;border:none;cursor:pointer;
    text-decoration:none;box-sizing:border-box;}}.wab:hover{{background:#1ebe5d}}</style>
    <a class="wab" href="{url}" onclick="window.open('{url}','_blank');return false;">{svg} {label}</a>""",
    height=height)

# ══════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION EXCEL / PDF
# ══════════════════════════════════════════════════════════════════════════════
def _border():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def build_excel(photos, colors, sizes):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    B=Border(left=Side(style="thin",color="FFD0D0D0"),right=Side(style="thin",color="FFD0D0D0"),
             top=Side(style="thin",color="FFD0D0D0"),bottom=Side(style="thin",color="FFD0D0D0"))
    items = photos if photos else [{"name":"Inventaire"}]
    for idx,p in enumerate(items,1):
        tab = (p.get("name",f"Modèle {idx}") or f"Modèle {idx}")[:28]
        ws = wb.create_sheet(title=tab)
        ws.cell(1,1,f"INVENTAIRE — {tab.upper()}").font=Font(name="Calibri",bold=True,size=13,color="FF111111")
        ws.cell(1,1).alignment=Alignment(horizontal="left",vertical="center")
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2+len(sizes))
        ws.row_dimensions[1].height=32
        for ci,h in enumerate(["Photo","Couleur"]+sizes,1):
            c=ws.cell(2,ci,h); c.font=Font(name="Calibri",bold=True,color="FFFFFFFF",size=10)
            c.fill=PatternFill("solid",fgColor="FF111111"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=B
        ws.row_dimensions[2].height=28
        ws.column_dimensions["A"].width=18; ws.column_dimensions["B"].width=22
        for i in range(len(sizes)): ws.column_dimensions[get_column_letter(3+i)].width=12
        for ri,col in enumerate(colors,3):
            rf=PatternFill("solid",fgColor="FFFFFFFF") if ri%2==0 else PatternFill("solid",fgColor="FFF9F9F9")
            c=ws.cell(ri,1,"[Photo]"); c.font=Font(name="Calibri",italic=True,color="FF999999",size=9)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.fill=rf; c.border=B
            c=ws.cell(ri,2,col); c.font=Font(name="Calibri",bold=True,size=10)
            c.alignment=Alignment(horizontal="left",vertical="center"); c.fill=rf; c.border=B
            for ci in range(3,3+len(sizes)):
                c=ws.cell(ri,ci,""); c.fill=rf; c.border=B; c.alignment=Alignment(horizontal="center",vertical="center")
            ws.row_dimensions[ri].height=40
    buf=BytesIO(); wb.save(buf); return buf.getvalue()

def build_order_excel(orders):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Liste de commande"
    B=Border(left=Side(style="thin",color="FFD0D0D0"),right=Side(style="thin",color="FFD0D0D0"),
             top=Side(style="thin",color="FFD0D0D0"),bottom=Side(style="thin",color="FFD0D0D0"))
    ws.cell(1,1,f"LISTE DE COMMANDE — {datetime.now().strftime('%d/%m/%Y')}").font=Font(name="Calibri",bold=True,size=13,color="FF111111")
    ws.merge_cells("A1:F1"); ws.row_dimensions[1].height=30
    for ci,h in enumerate(["Modèle","Couleur","Taille","En stock","Seuil min","À commander"],1):
        c=ws.cell(2,ci,h); c.font=Font(name="Calibri",bold=True,color="FFFFFFFF",size=10)
        c.fill=PatternFill("solid",fgColor="FF111111"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=B
    ws.row_dimensions[2].height=26
    for ri,row in enumerate(orders,3):
        fill=PatternFill("solid",fgColor="FFFFE0E0") if row["En stock"]==0 else PatternFill("solid",fgColor="FFFFF3CD")
        for ci,v in enumerate([row["Modèle"],row["Couleur"],row["Taille"],row["En stock"],row["Seuil"],row["À commander"]],1):
            c=ws.cell(ri,ci,v); c.fill=fill; c.border=B
            c.alignment=Alignment(horizontal="left" if ci<=2 else "center",vertical="center")
            if ci==6: c.font=Font(name="Calibri",bold=True,size=10,color="FFCC0000")
        ws.row_dimensions[ri].height=22
    for i,w in enumerate([28,20,10,12,12,14],1): ws.column_dimensions[get_column_letter(i)].width=w
    buf=BytesIO(); wb.save(buf); return buf.getvalue()

def build_ventes_excel(ventes):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Ventes"
    B=Border(left=Side(style="thin",color="FFD0D0D0"),right=Side(style="thin",color="FFD0D0D0"),
             top=Side(style="thin",color="FFD0D0D0"),bottom=Side(style="thin",color="FFD0D0D0"))
    ws.cell(1,1,"HISTORIQUE DES VENTES").font=Font(name="Calibri",bold=True,size=13,color="FF111111")
    ws.merge_cells("A1:H1"); ws.row_dimensions[1].height=30
    for ci,h in enumerate(["Date","Heure","Vendeur","Modèle","Couleur","Taille","Qté","Note"],1):
        c=ws.cell(2,ci,h); c.font=Font(name="Calibri",bold=True,color="FFFFFFFF",size=10)
        c.fill=PatternFill("solid",fgColor="FF1A7A3C"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=B
    ws.row_dimensions[2].height=26
    for ri,v in enumerate(ventes,3):
        dp=v.get("date","").split(" "); date_p=dp[0] if dp else ""; time_p=dp[1] if len(dp)>1 else ""
        fill=PatternFill("solid",fgColor="FFFFFFFF") if ri%2==0 else PatternFill("solid",fgColor="FFF0FFF4")
        for ci,val in enumerate([date_p,time_p,v.get("vendeur",""),v.get("model_name",""),
                                  v.get("couleur",""),v.get("taille",""),v.get("quantite",1),v.get("note","")],1):
            c=ws.cell(ri,ci,val); c.fill=fill; c.border=B
            c.alignment=Alignment(horizontal="left" if ci in[3,4,8] else "center",vertical="center")
        ws.row_dimensions[ri].height=20
    for i,w in enumerate([13,10,16,26,18,10,8,22],1): ws.column_dimensions[get_column_letter(i)].width=w
    buf=BytesIO(); wb.save(buf); return buf.getvalue()

def build_pdf(photos_data, colors, sizes, model_name):
    buf=BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=1.5*cm,bottomMargin=1.5*cm)
    BLACK=rl_colors.HexColor("#111111"); GREY=rl_colors.HexColor("#F5F5F5"); LGREY=rl_colors.HexColor("#E0E0E0")
    s_title=ParagraphStyle("t",fontName="Helvetica-Bold",fontSize=16,textColor=BLACK,spaceAfter=4,letterSpacing=2)
    s_sub=ParagraphStyle("s",fontName="Helvetica",fontSize=8,textColor=rl_colors.HexColor("#888"),spaceAfter=10)
    s_pname=ParagraphStyle("pn",fontName="Helvetica-Bold",fontSize=10,textColor=BLACK,spaceBefore=14,spaceAfter=6)
    s_hdr=ParagraphStyle("hdr",fontName="Helvetica-Bold",fontSize=8,textColor=rl_colors.white,alignment=TA_CENTER)
    s_cell=ParagraphStyle("cell",fontName="Helvetica",fontSize=8,textColor=BLACK)
    story=[]
    titre=(model_name.upper() if model_name else "INVENTAIRE")
    story.append(Paragraph(f"INVENTAIRE — {titre}",s_title))
    story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}  •  {len(colors)} couleur(s)  •  Tailles: {', '.join(sizes)}",s_sub))
    story.append(HRFlowable(width="100%",thickness=1,color=BLACK,spaceAfter=10))
    items=photos_data if photos_data and any(p for p in photos_data) else [None]
    for photo in items:
        if photo and photo.get("name"): story.append(Paragraph(photo["name"],s_pname))
        hrow=[Paragraph("PHOTO",s_hdr),Paragraph("COULEUR",s_hdr)]+[Paragraph(s,s_hdr) for s in sizes]
        drows=[]
        for i,color in enumerate(colors):
            if i==0 and photo and photo.get("img_bytes"):
                try:
                    rl_img=RLImage(BytesIO(photo["img_bytes"]),width=2*cm,height=2.5*cm); photo_cell=rl_img
                except: photo_cell=Paragraph("[Photo]",s_cell)
            elif i==0: photo_cell=Paragraph("[Photo]",s_cell)
            else: photo_cell=""
            hex_v=HEX_MAP.get(color,"#888")
            cpara=Paragraph(f'<font color="{hex_v}">■</font> {color}',
                            ParagraphStyle("c",fontName="Helvetica",fontSize=8,textColor=BLACK))
            drows.append([photo_cell,cpara]+[""]*len(sizes))
        pw=A4[0]-3*cm; phw=2.4*cm; cw=4.0*cm; sw=(pw-phw-cw)/max(len(sizes),1)
        tbl=Table([hrow]+drows,colWidths=[phw,cw]+[sw]*len(sizes),repeatRows=1)
        ts=TableStyle([("BACKGROUND",(0,0),(-1,0),BLACK),("TEXTCOLOR",(0,0),(-1,0),rl_colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8),
            ("ALIGN",(0,0),(-1,0),"CENTER"),("VALIGN",(0,0),(-1,0),"MIDDLE"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.white,GREY]),("GRID",(0,0),(-1,-1),0.4,LGREY),
            ("LINEBELOW",(0,0),(-1,0),0.8,BLACK),("VALIGN",(0,1),(-1,-1),"MIDDLE"),
            ("ALIGN",(2,1),(-1,-1),"CENTER"),("LEFTPADDING",(1,1),(1,-1),6),
            ("TOPPADDING",(0,1),(-1,-1),6),("BOTTOMPADDING",(0,1),(-1,-1),6),
            ("SPAN",(0,1),(0,len(colors))),("ALIGN",(0,1),(0,-1),"CENTER"),("VALIGN",(0,1),(0,-1),"MIDDLE")])
        tbl.setStyle(ts); story.append(tbl); story.append(Spacer(1,0.5*cm))
    doc.build(story); return buf.getvalue()

def _df_to_excel(df: pd.DataFrame, sheet: str) -> bytes:
    """Convertit un DataFrame en Excel simple."""
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=sheet
    B=Border(left=Side(style="thin",color="FFD0D0D0"),right=Side(style="thin",color="FFD0D0D0"),
             top=Side(style="thin",color="FFD0D0D0"),bottom=Side(style="thin",color="FFD0D0D0"))
    for ci,h in enumerate(df.columns,1):
        c=ws.cell(1,ci,str(h)); c.font=Font(name="Calibri",bold=True,color="FFFFFFFF",size=10)
        c.fill=PatternFill("solid",fgColor="FF111111"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=B
    for ri,row in df.iterrows():
        for ci,v in enumerate(row,1):
            c=ws.cell(ri+2,ci,v); c.border=B
            c.fill=PatternFill("solid",fgColor="FFFFFFFF" if (ri+2)%2==0 else "FFF9F9F9")
    buf=BytesIO(); wb.save(buf); return buf.getvalue()

def build_rank_excel(df_rank, periode):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Classement"
    B=Border(left=Side(style="thin",color="FFD0D0D0"),right=Side(style="thin",color="FFD0D0D0"),
             top=Side(style="thin",color="FFD0D0D0"),bottom=Side(style="thin",color="FFD0D0D0"))
    GOLD="FFCFB53B"; SILVER="FFC0C0C0"; BRONZE="FFCD7F32"
    ws.cell(1,1,f"CLASSEMENT DES VENTES — {periode.upper()} — {datetime.now().strftime('%d/%m/%Y')}").font=Font(name="Calibri",bold=True,size=13,color="FF111111")
    ws.merge_cells("A1:E1"); ws.row_dimensions[1].height=30
    for ci,h in enumerate(["Rang","Modèle","Couleur","Taille","Pièces vendues"],1):
        c=ws.cell(2,ci,h); c.font=Font(name="Calibri",bold=True,color="FFFFFFFF",size=10)
        c.fill=PatternFill("solid",fgColor="FF111111"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=B
    ws.row_dimensions[2].height=26
    medal_fills={"1":PatternFill("solid",fgColor="FFFFF9E6"),"2":PatternFill("solid",fgColor="FFF8F8F8"),"3":PatternFill("solid",fgColor="FFFFF8F0")}
    medals={"1":"🥇","2":"🥈","3":"🥉"}
    for ri,row in df_rank.iterrows():
        r=ri+3; rang=str(int(row["Rang"]))
        fill=medal_fills.get(rang,PatternFill("solid",fgColor="FFFFFFFF"))
        for ci,v in enumerate([medals.get(rang,f"#{rang}"),row["Modèle"],row["Couleur"],row["Taille"],row["Pièces vendues"]],1):
            c=ws.cell(r,ci,v); c.fill=fill; c.border=B
            c.alignment=Alignment(horizontal="center" if ci in[1,4,5] else "left",vertical="center")
            if ci==5: c.font=Font(name="Calibri",bold=True,size=11)
        ws.row_dimensions[r].height=22
    for i,w in enumerate([8,28,18,10,16],1): ws.column_dimensions[get_column_letter(i)].width=w
    buf=BytesIO(); wb.save(buf); return buf.getvalue()

def get_order_list(stock_data):
    orders=[]
    for key,data in stock_data.items():
        seuil=data.get("seuil_min",0); name=data.get("model_name",key)
        for color in data.get("colors",[]):
            for size in data.get("sizes",[]):
                qty=data.get("stock",{}).get(color,{}).get(size,0)
                if seuil-qty>0:
                    orders.append({"Modèle":name,"Couleur":color,"Taille":size,
                                   "En stock":qty,"Seuil":seuil,"À commander":seuil-qty})
    return orders

# ══════════════════════════════════════════════════════════════════════════════
# WIDGET PALETTE COULEURS
# ══════════════════════════════════════════════════════════════════════════════
def color_palette_picker(key_suffix=""):
    st.markdown('<span class="section-label">Couleurs disponibles</span>', unsafe_allow_html=True)

    palette_names = [c["name"] for c in COLOR_PALETTE]
    ss_key = f"selected_colors_{key_suffix}"
    if ss_key not in st.session_state:
        st.session_state[ss_key] = []

    # ── Menu déroulant multi-sélection ──
    chosen = st.multiselect(
        "",
        options=palette_names,
        default=[c for c in st.session_state[ss_key] if c in palette_names],
        format_func=lambda name: name,
        placeholder="Choisissez une ou plusieurs couleurs…",
        label_visibility="collapsed",
        key=f"multiselect_colors_{key_suffix}",
    )

    # ── Couleur personnalisée ──
    st.markdown('<span class="section-label" style="margin-top:.8rem">Ajouter une couleur personnalisée</span>', unsafe_allow_html=True)
    c1, c2 = st.columns([3, 1])
    with c1:
        cc = st.text_input("", placeholder="Ex: Vert sauge, Taupe…",
                           key=f"custom_color_input_{key_suffix}", label_visibility="collapsed")
    with c2:
        if st.button("Ajouter", key=f"add_custom_{key_suffix}"):
            n = cc.strip()
            if n and n not in st.session_state[ss_key]:
                st.session_state[ss_key].append(n)
                st.rerun()
            elif not n:
                st.warning("Saisissez un nom.")

    # Fusionne la sélection du multiselect + les couleurs custom
    custom_colors = [c for c in st.session_state[ss_key] if c not in palette_names]
    all_selected = chosen + custom_colors
    st.session_state[ss_key] = all_selected

    # ── Aperçu des couleurs choisies ──
    if all_selected:
        st.markdown('<span class="section-label" style="margin-top:.8rem">Aperçu</span>', unsafe_allow_html=True)
        tags = '<div class="selected-colors-bar">'
        for nm in all_selected:
            hx = HEX_MAP.get(nm, "#888888")
            tags += (f'<span class="color-tag">'
                     f'<span class="color-dot" style="background:{hx}"></span>'
                     f'{nm}</span>')
        tags += "</div>"
        st.markdown(tags, unsafe_allow_html=True)

        if custom_colors:
            st.caption(f"Couleurs personnalisées : {', '.join(custom_colors)}")
            if st.button("✕  Effacer les couleurs personnalisées",
                         key=f"clear_custom_{key_suffix}"):
                st.session_state[ss_key] = [c for c in st.session_state[ss_key]
                                             if c in palette_names]
                st.rerun()

    return all_selected

# ══════════════════════════════════════════════════════════════════════════════
# ÉCRAN DE CONNEXION
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.user_name is None:
    profils_data = load_profils()
    vendeurs = profils_data.get("vendeurs", [])

    # ── Logo centré ──────────────────────────────────────────────────────────
    st.markdown("""
    <div style="text-align:center;padding:2rem 0 2rem">
      <div style="font-size:3.5rem">👕</div>
      <div style="font-family:'Space Grotesk',sans-serif;font-size:1.8rem;font-weight:700;
                  color:#1B2B4B;letter-spacing:.04em">LATIF SHOP</div>
      <div style="font-size:.72rem;color:#8A9AB5;letter-spacing:.18em;
                  text-transform:uppercase;margin-top:.3rem">Gestion Inventaire &amp; Ventes</div>
    </div>
    """, unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════
    # VUE 1 — Choix de l'espace (deux grandes cartes)
    # ════════════════════════════════════════════════════════
    if st.session_state.login_space is None:
        _, c1, g1, c2, g2, c3, _ = st.columns([0.5, 3, 0.2, 3, 0.2, 3, 0.5])
        with c1:
            st.markdown("""
            <div style="background:#FFFFFF;border:3px solid #1B2B4B;border-radius:20px;
                        padding:2.5rem 1.5rem;text-align:center;
                        box-shadow:0 4px 24px rgba(27,43,75,.12)">
              <div style="font-size:3rem;margin-bottom:.8rem">🧑‍💼</div>
              <div style="font-family:'Space Grotesk',sans-serif;font-size:1.2rem;
                          font-weight:700;color:#1B2B4B">Espace Vendeurs</div>
              <div style="font-size:.75rem;color:#8A9AB5;margin-top:.4rem">Saisir les ventes</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Entrer →", key="btn_space_vendeur", use_container_width=True):
                st.session_state.login_space = "vendeur"
                st.rerun()

        with c2:
            st.markdown("""
            <div style="background:#FFFFFF;border:3px solid #059669;border-radius:20px;
                        padding:2.5rem 1.5rem;text-align:center;
                        box-shadow:0 4px 24px rgba(5,150,105,.12)">
              <div style="font-size:3rem;margin-bottom:.8rem">📦</div>
              <div style="font-family:'Space Grotesk',sans-serif;font-size:1.2rem;
                          font-weight:700;color:#1B2B4B">Alimentation Stock</div>
              <div style="font-size:.75rem;color:#8A9AB5;margin-top:.4rem">Ajouter des articles</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Entrer →", key="btn_space_ali", use_container_width=True):
                st.session_state.login_space = "alimentation"
                st.rerun()

        with c3:
            st.markdown("""
            <div style="background:#FFFFFF;border:3px solid #C09020;border-radius:20px;
                        padding:2.5rem 1.5rem;text-align:center;
                        box-shadow:0 4px 24px rgba(192,144,32,.15)">
              <div style="font-size:3rem;margin-bottom:.8rem">👑</div>
              <div style="font-family:'Space Grotesk',sans-serif;font-size:1.2rem;
                          font-weight:700;color:#1B2B4B">Espace Admin</div>
              <div style="font-size:.75rem;color:#8A9AB5;margin-top:.4rem">Gestion &amp; stock</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Entrer →", key="btn_space_admin", use_container_width=True):
                st.session_state.login_space = "admin"
                st.rerun()

    # ════════════════════════════════════════════════════════
    # VUE 2A — Espace Vendeurs : liste des noms
    # ════════════════════════════════════════════════════════
    elif st.session_state.login_space == "vendeur":
        st.markdown("""
        <div style="text-align:center;margin-bottom:1.4rem">
          <span style="background:#1B2B4B;color:#FFF;font-size:.72rem;font-weight:600;
                       letter-spacing:.15em;text-transform:uppercase;
                       padding:.4rem 1.2rem;border-radius:100px">🧑‍💼 Espace Vendeurs</span>
        </div>
        """, unsafe_allow_html=True)

        if vendeurs:
            nb = min(len(vendeurs), 4)
            cols_v = st.columns(nb)
            for i, nom in enumerate(vendeurs):
                with cols_v[i % nb]:
                    ini = "".join(p[0].upper() for p in nom.split()[:2])
                    st.markdown(f'<div class="profil-avatar">{ini}</div>',
                                unsafe_allow_html=True)
                    if st.button(nom, key=f"login_{nom}", use_container_width=True):
                        st.session_state.user_name = nom
                        st.session_state.user_role = "vendeur"
                        st.session_state.login_space = None
                        st.rerun()
        else:
            st.info("Aucun vendeur enregistré. Connectez-vous en admin pour en ajouter.")

        st.markdown("<div style='height:.8rem'></div>", unsafe_allow_html=True)
        if st.button("← Retour", key="back_vendeur"):
            st.session_state.login_space = None
            st.rerun()

    # ════════════════════════════════════════════════════════
    # VUE 2B — Espace Admin : PIN
    # ════════════════════════════════════════════════════════
    elif st.session_state.login_space == "admin":
        st.markdown("""
        <div style="text-align:center;margin-bottom:1.4rem">
          <span style="background:#C09020;color:#FFF;font-size:.72rem;font-weight:600;
                       letter-spacing:.15em;text-transform:uppercase;
                       padding:.4rem 1.2rem;border-radius:100px">👑 Espace Admin</span>
        </div>
        """, unsafe_allow_html=True)

        _, pin_col, _ = st.columns([1, 2, 1])
        with pin_col:
            pin = st.text_input("Code PIN", type="password", placeholder="• • • •",
                                label_visibility="visible", key="pin_input")
            if st.button("🔐  Connexion Admin", use_container_width=True, key="btn_admin"):
                if pin == profils_data.get("admin_pin", "1234"):
                    st.session_state.user_name = "Admin"
                    st.session_state.user_role = "admin"
                    st.session_state.pin_error = False
                    st.session_state.login_space = None
                    st.rerun()
                else:
                    st.session_state.pin_error = True
                    st.rerun()
            if st.session_state.pin_error:
                st.error("❌  Code PIN incorrect.")

        st.markdown("<div style='height:.8rem'></div>", unsafe_allow_html=True)
        if st.button("← Retour", key="back_admin"):
            st.session_state.pin_error = False
            st.session_state.login_space = None
            st.rerun()

    # ════════════════════════════════════════════════════════
    # VUE 2C — Espace Alimentation : PIN livreur
    # ════════════════════════════════════════════════════════
    elif st.session_state.login_space == "alimentation":
        profils_ali = load_profils()
        ali_pin_stored = profils_ali.get("alimentation_pin", "")

        st.markdown("""
        <div style="text-align:center;margin-bottom:1.4rem">
          <span style="background:#059669;color:#FFF;font-size:.72rem;font-weight:600;
                       letter-spacing:.15em;text-transform:uppercase;
                       padding:.4rem 1.2rem;border-radius:100px">📦 Alimentation Stock</span>
        </div>
        """, unsafe_allow_html=True)

        if not ali_pin_stored:
            st.warning("⚠️ Aucun code PIN alimentation configuré. Contactez l'administrateur.")
        else:
            _, ac, _ = st.columns([1, 2, 1])
            with ac:
                ali_nom = st.text_input("Votre nom (livreur)", placeholder="Ex: Ahmed",
                                        key="ali_nom_input")
                ali_pin_input = st.text_input("Code PIN alimentation", type="password",
                                              placeholder="• • • •", key="ali_pin_input")
                if st.button("📦 Accéder à l'alimentation", use_container_width=True,
                             key="btn_ali_login"):
                    if not ali_nom.strip():
                        st.warning("Saisissez votre nom.")
                    elif ali_pin_input == ali_pin_stored:
                        st.session_state.user_name = ali_nom.strip()
                        st.session_state.user_role = "livreur"
                        st.session_state.ali_livreur = ali_nom.strip()
                        st.session_state.ali_pin_error = False
                        st.session_state.login_space = None
                        st.rerun()
                    else:
                        st.session_state.ali_pin_error = True
                        st.rerun()
                if st.session_state.ali_pin_error:
                    st.error("❌ Code PIN incorrect.")

        st.markdown("<div style='height:.8rem'></div>", unsafe_allow_html=True)
        if st.button("← Retour", key="back_ali"):
            st.session_state.ali_pin_error = False
            st.session_state.login_space = None
            st.rerun()

    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# FLUX ALIMENTATION LIVREUR
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.user_role == "livreur":
    livreur_nom = st.session_state.user_name
    now_str_ali = datetime.now().strftime("%A %d %b %Y  —  %H:%M")
    lh1, lh2 = st.columns([5, 1])
    with lh1:
        st.markdown(f"""
        <div class="app-header">
          <div class="app-logo">📦</div>
          <div><div class="app-title">ALIMENTATION STOCK</div>
          <div class="app-sub">Latif Shop</div></div>
          <div class="app-badge">🚚 {livreur_nom}</div>
          <div style="margin-left:1rem;color:rgba(255,255,255,.6);font-size:.72rem">{now_str_ali}</div>
        </div>""", unsafe_allow_html=True)
    with lh2:
        st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
        if st.button("⏻ Déconnexion", use_container_width=True, key="ali_logout"):
            for k in ["user_name","user_role","ali_step","ali_cat","ali_art_key","ali_qtys","ali_livreur"]:
                st.session_state[k] = None if k in ["user_name","user_role","ali_cat","ali_art_key"] else (0 if k=="ali_step" else ({} if k=="ali_qtys" else ""))
            st.rerun()

    stock_ali  = load_stock()
    history_ali = load_history()
    cats_ali   = load_categories()
    cur_ali    = st.session_state.ali_step

    # ── ÉTAPE 0 : Choisir la catégorie ──────────────────────────────────────
    if cur_ali == 0:
        st.markdown('<span class="section-label">Choisissez une catégorie</span>', unsafe_allow_html=True)
        if not cats_ali:
            st.info("Aucune catégorie disponible.")
        else:
            clicked_ali_cat = None
            nb_c = min(len(cats_ali), 4)
            rows_ali = [cats_ali[i:i+nb_c] for i in range(0, len(cats_ali), nb_c)]
            for row in rows_ali:
                rcols = st.columns(nb_c)
                for ci, cat in enumerate(row):
                    cat_n = cat.get("name","") if isinstance(cat,dict) else cat
                    cat_i = cat.get("icon","📦") if isinstance(cat,dict) else "📦"
                    cat_c = cat.get("color","#1B2B4B") if isinstance(cat,dict) else "#1B2B4B"
                    nb_art_ali = sum(1 for e in history_ali if e.get("categorie","") == cat_n)
                    with rcols[ci]:
                        st.markdown(
                            f'<div style="background:#FFF;border:2px solid {cat_c}44;'
                            f'border-radius:14px;padding:1.2rem;text-align:center;margin-bottom:.4rem">'
                            f'<div style="font-size:2.2rem">{cat_i}</div>'
                            f'<div style="font-weight:700;color:#1B2B4B;font-size:.9rem;margin:.4rem 0">{cat_n}</div>'
                            f'<div style="font-size:.7rem;color:#8A9AB5">{nb_art_ali} article(s)</div></div>',
                            unsafe_allow_html=True)
                        if st.button("Sélectionner", key=f"ali_cat_{ci}", use_container_width=True):
                            clicked_ali_cat = cat_n
            if clicked_ali_cat:
                st.session_state.ali_cat = clicked_ali_cat
                st.session_state.ali_step = 1
                st.rerun()

    # ── ÉTAPE 1 : Choisir l'article ─────────────────────────────────────────
    elif cur_ali == 1:
        st.markdown(f'<div style="font-size:.78rem;color:#8A9AB5;margin-bottom:.8rem">📂 {st.session_state.ali_cat}</div>', unsafe_allow_html=True)
        st.markdown('<span class="section-label">Choisissez un article</span>', unsafe_allow_html=True)
        arts_ali = []
        seen_ali = set()
        for e in history_ali:
            mk = model_key(e["model_name"])
            if mk in seen_ali or mk not in stock_ali: continue
            if e.get("categorie","") != st.session_state.ali_cat: continue
            seen_ali.add(mk); arts_ali.append((mk, e))
        if not arts_ali:
            st.info("Aucun article dans cette catégorie.")
        else:
            clicked_ali_art = None
            for row in [arts_ali[i:i+3] for i in range(0, len(arts_ali), 3)]:
                r0, r1, r2 = st.columns(3)
                for ci2, col_ref in enumerate([r0, r1, r2]):
                    if ci2 >= len(row): break
                    mk_a, entry_a = row[ci2]
                    b64_a = entry_a.get("b64_thumb") or stock_ali.get(mk_a,{}).get("b64_thumb","")
                    photo_h = (f'<img src="data:image/png;base64,{b64_a}" style="width:100%;height:100px;object-fit:cover;border-radius:8px;margin-bottom:.4rem">'
                               if b64_a else '<div style="height:100px;background:#EEF1F7;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:2rem;margin-bottom:.4rem">👕</div>')
                    with col_ref:
                        st.markdown(
                            f'<div style="background:#FFF;border:2px solid #E0E5EF;border-radius:12px;padding:.7rem">'
                            f'{photo_h}<div style="font-weight:700;color:#1B2B4B;font-size:.82rem">{entry_a["model_name"]}</div></div>',
                            unsafe_allow_html=True)
                        if st.button("Sélectionner", key=f"ali_art_{mk_a}", use_container_width=True):
                            clicked_ali_art = mk_a
            if clicked_ali_art:
                st.session_state.ali_art_key = clicked_ali_art
                st.session_state.ali_qtys = {}
                st.session_state.ali_step = 2
                st.rerun()
        if st.button("← Retour catégories", key="ali_back_cat"):
            st.session_state.ali_step = 0; st.rerun()

    # ── ÉTAPE 2 : Tailles +/− ───────────────────────────────────────────────
    elif cur_ali == 2:
        mk_ali   = st.session_state.ali_art_key
        sd_ali   = stock_ali.get(mk_ali, {})
        art_name_ali = sd_ali.get("model_name", mk_ali)
        colors_ali = sd_ali.get("colors", [])
        sizes_ali  = sd_ali.get("sizes", [])
        b64_ali    = sd_ali.get("b64_thumb","")

        st.markdown(f'<div style="font-size:.78rem;color:#8A9AB5;margin-bottom:.8rem">📂 {st.session_state.ali_cat} › <strong style="color:#1B2B4B">{art_name_ali}</strong></div>', unsafe_allow_html=True)

        ah1, ah2 = st.columns([1, 3])
        with ah1:
            if b64_ali:
                st.markdown(f'<img src="data:image/png;base64,{b64_ali}" style="width:100%;border-radius:12px">', unsafe_allow_html=True)
            else:
                st.markdown('<div style="font-size:4rem;text-align:center">👕</div>', unsafe_allow_html=True)
        with ah2:
            st.markdown(f'<div style="font-size:1.1rem;font-weight:700;color:#1B2B4B;margin-bottom:.8rem">{art_name_ali}</div>', unsafe_allow_html=True)
            st.markdown('<span class="section-label">Quantités à ajouter par taille</span>', unsafe_allow_html=True)

            if "ali_qtys" not in st.session_state or not isinstance(st.session_state.ali_qtys, dict):
                st.session_state.ali_qtys = {}

            for color in colors_ali:
                hx_ali = HEX_MAP.get(color, "#888")
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:.5rem;margin:.6rem 0 .3rem">'
                    f'<span style="width:14px;height:14px;border-radius:50%;background:{hx_ali};'
                    f'border:1px solid rgba(0,0,0,.15);display:inline-block"></span>'
                    f'<strong style="font-size:.85rem;color:#1B2B4B">{color}</strong></div>',
                    unsafe_allow_html=True)
                sz_cols = st.columns(min(len(sizes_ali), 5))
                for si, sz in enumerate(sizes_ali):
                    key_q = f"{color}__{sz}"
                    cur_stock = sd_ali.get("stock",{}).get(color,{}).get(sz, 0)
                    with sz_cols[si % 5]:
                        st.markdown(f'<div style="text-align:center;font-size:.7rem;color:#8A9AB5;margin-bottom:.2rem">{sz}<br><span style="color:#059669">{cur_stock} en stock</span></div>', unsafe_allow_html=True)
                        val_q = st.number_input("", min_value=0, max_value=500, value=st.session_state.ali_qtys.get(key_q, 0), step=1, key=f"ali_q_{mk_ali}_{color}_{sz}", label_visibility="collapsed")
                        st.session_state.ali_qtys[key_q] = int(val_q)

        total_added = sum(v for v in st.session_state.ali_qtys.values() if v > 0)
        st.markdown(f'<div style="background:#F0FDF4;border:1px solid #059669;border-radius:10px;padding:.8rem 1rem;margin:1rem 0;font-size:.85rem;color:#065F46"><strong>Total à ajouter : {total_added} pièce(s)</strong></div>', unsafe_allow_html=True)

        ok1, ok2 = st.columns(2)
        with ok1:
            if st.button("← Retour articles", key="ali_back_art", use_container_width=True):
                st.session_state.ali_step = 1; st.rerun()
        with ok2:
            if st.button("✅ Envoyer la demande", key="ali_send", use_container_width=True):
                if total_added == 0:
                    st.warning("Ajoutez au moins une pièce.")
                else:
                    pending = load_ali_pending()
                    pending.append({
                        "date":      datetime.now().strftime("%d/%m/%Y %H:%M"),
                        "livreur":   livreur_nom,
                        "model_key": mk_ali,
                        "model_name": art_name_ali,
                        "categorie": st.session_state.ali_cat,
                        "qtys":      {k: v for k, v in st.session_state.ali_qtys.items() if v > 0},
                        "statut":    "en_attente",
                    })
                    save_ali_pending(pending)
                    st.success("✅ Demande envoyée ! En attente d'approbation de l'administrateur.")
                    st.session_state.ali_step = 0
                    st.session_state.ali_qtys = {}
                    st.session_state.ali_art_key = None
                    st.rerun()

    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# BARRE UTILISATEUR
# ══════════════════════════════════════════════════════════════════════════════
# ── Bandeau en-tête principal ─────────────────────────────────────────────
badge_icon = "👑" if st.session_state.user_role == "admin" else "🧑‍💼"
badge_label = "Administrateur" if st.session_state.user_role == "admin" else "Vendeur"
user_name = st.session_state.user_name
now_str = datetime.now().strftime("%A %d %b %Y  —  %H:%M")

header_left, header_right = st.columns([5, 1])
with header_left:
    st.markdown(f"""
    <div class="app-header">
      <div class="app-logo">👕</div>
      <div>
        <div class="app-title">LATIF SHOP</div>
        <div class="app-sub">Gestion Inventaire &amp; Ventes</div>
      </div>
      <div class="app-badge">{badge_icon} {badge_label} — {user_name}</div>
      <div style="margin-left:1rem;color:var(--text3);font-size:.72rem;letter-spacing:.08em;
                  white-space:nowrap">{now_str}</div>
    </div>""", unsafe_allow_html=True)
with header_right:
    st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
    if st.button("⏻  Déconnexion", use_container_width=True):
        for k in ["user_name", "user_role", "pin_error"]:
            st.session_state[k] = None if k != "pin_error" else False
        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# ONGLETS SELON RÔLE
# ══════════════════════════════════════════════════════════════════════════════
is_admin = st.session_state.user_role == "admin"

if is_admin:
    tab_dashboard, tab_new, tab_hist, tab_stock, tab_commandes, tab_ventes, tab_ali_admin, tab_admin = st.tabs([
        "📊  Dashboard", "✦  Nouvelle Grille", "📋  Historique",
        "📦  Stock", "🛒  Commandes", "💰  Ventes", "📥  Alimentation", "⚙  Profils & Config"])
else:
    tab_ventes = st.tabs(["💰  Saisie des Ventes"])[0]
    tab_dashboard = tab_new = tab_hist = tab_stock = tab_commandes = tab_ali_admin = tab_admin = None

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET DASHBOARD (admin seulement)
# ══════════════════════════════════════════════════════════════════════════════
if tab_dashboard:
    with tab_dashboard:
        ventes_all  = load_ventes()
        stock_dash  = load_stock()
        today_str   = datetime.now().strftime("%d/%m/%Y")
        week_start  = (datetime.now() - timedelta(days=datetime.now().weekday())).strftime("%d/%m/%Y")

        ventes_today = [v for v in ventes_all if v.get("date","").startswith(today_str)]
        ventes_week  = [v for v in ventes_all
                        if v.get("date","") >= week_start.replace("/","") or
                        v.get("date","")[:10].replace("/","") >= week_start.replace("/","")]

        # Recalcul correct semaine
        def in_this_week(date_str):
            try:
                d = datetime.strptime(date_str[:10], "%d/%m/%Y")
                return (datetime.now() - d).days <= 7
            except: return False
        ventes_week = [v for v in ventes_all if in_this_week(v.get("date",""))]

        pieces_today = sum(v.get("quantite",1) for v in ventes_today)
        pieces_week  = sum(v.get("quantite",1) for v in ventes_week)
        nb_ali_pending = len(load_ali_pending())

        # Articles en rupture (calculé avant les métriques)
        ruptures = [(k, d) for k, d in stock_dash.items()
                    if sum(q for c in d.get("stock",{}).values() for q in c.values()) == 0]

        # ── Métriques grandes cartes ──────────────────────────────────────────
        m1, m2, m3, m4 = st.columns(4)
        for col_m, icon_m, label_m, val_m, sub_m, color_m in [
            (m1, "💰", "Ventes aujourd'hui",   str(len(ventes_today)),  f"{pieces_today} pièces",    "#2563EB"),
            (m2, "📅", "Ventes cette semaine",  str(len(ventes_week)),   f"{pieces_week} pièces",     "#7C3AED"),
            (m3, "🔴", "Ruptures stock",        str(len(ruptures)),      "articles épuisés",          "#DC2626"),
            (m4, "📥", "Attente approbation",   str(nb_ali_pending),     "alimentations",             "#D97706"),
        ]:
            with col_m:
                st.markdown(
                    f'<div style="background:#FFF;border:2px solid {color_m}33;border-radius:16px;'
                    f'padding:1.4rem 1.2rem;text-align:center;box-shadow:0 2px 8px {color_m}18">'
                    f'<div style="font-size:2.2rem;margin-bottom:.4rem">{icon_m}</div>'
                    f'<div style="font-size:3rem;font-weight:800;color:{color_m};line-height:1">{val_m}</div>'
                    f'<div style="font-size:1rem;font-weight:700;color:#1B2B4B;margin:.5rem 0 .2rem">{label_m}</div>'
                    f'<div style="font-size:.85rem;color:#8A9AB5">{sub_m}</div>'
                    f'</div>',
                    unsafe_allow_html=True)

        st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

        # ── Top 3 articles ───────────────────────────────────────────────────
        st.markdown(
            '<div style="font-size:1.5rem;font-weight:800;color:#1B2B4B;margin-bottom:1rem">🏆 Top 3 articles les plus vendus</div>',
            unsafe_allow_html=True)
        from collections import Counter
        # Compter par model_key (clé réelle) pour retrouver la photo correctement
        top_keys_counter = Counter()
        key_to_name = {}
        for v in ventes_all:
            mk_v = v.get("model_key") or model_key(v.get("model_name","?"))
            top_keys_counter[mk_v] += v.get("quantite", 1)
            if mk_v not in key_to_name:
                key_to_name[mk_v] = v.get("model_name","?")
        top3 = top_keys_counter.most_common(3)
        if top3:
            t1, t2, t3 = st.columns(3)
            medailles = ["🥇", "🥈", "🥉"]
            for idx, (col_t, (mk_t, qty_t)) in enumerate(zip([t1,t2,t3], top3)):
                with col_t:
                    nom_t = key_to_name.get(mk_t, mk_t)
                    b64_t = stock_dash.get(mk_t,{}).get("b64_thumb","")
                    photo_t = (
                        f'<img src="data:image/png;base64,{b64_t}" style="width:100%;height:160px;object-fit:cover;border-radius:12px;margin-bottom:.7rem">'
                        if b64_t else
                        f'<div style="height:160px;background:#EEF1F7;border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:3.5rem;margin-bottom:.7rem">👕</div>'
                    )
                    st.markdown(
                        f'<div style="background:#FFF;border:2px solid #E0E5EF;border-radius:16px;padding:1.2rem;text-align:center;box-shadow:0 2px 10px #0001">'
                        f'{photo_t}'
                        f'<div style="font-size:2rem;margin-bottom:.4rem">{medailles[idx]}</div>'
                        f'<div style="font-weight:800;color:#1B2B4B;font-size:1.05rem;margin:.3rem 0">{nom_t}</div>'
                        f'<div style="color:#059669;font-weight:800;font-size:1.3rem">{qty_t}</div>'
                        f'<div style="color:#8A9AB5;font-size:.85rem">pièces vendues</div>'
                        f'</div>',
                        unsafe_allow_html=True)
        else:
            st.info("Aucune vente enregistrée.")

        st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

        # ── Alertes stock faible ─────────────────────────────────────────────
        st.markdown(
            '<div style="font-size:1.5rem;font-weight:800;color:#1B2B4B;margin-bottom:1rem">🔔 Alertes stock</div>',
            unsafe_allow_html=True)
        alertes = []
        for mk, d in stock_dash.items():
            nom_a = d.get("model_name", mk)
            seuil_a = d.get("seuil_min", 2)
            total_a = sum(q for c in d.get("stock",{}).values() for q in c.values())
            if total_a == 0:
                alertes.append(("🔴 RUPTURE", nom_a, total_a, seuil_a))
            elif total_a <= seuil_a:
                alertes.append(("🟡 STOCK FAIBLE", nom_a, total_a, seuil_a))

        if not alertes:
            st.success("✅ Tous les articles ont un stock suffisant.")
        else:
            for badge_a, nom_a, total_a, seuil_a in alertes:
                couleur_a = "#DC2626" if "RUPTURE" in badge_a else "#D97706"
                st.markdown(
                    f'<div style="background:#FFF;border-left:6px solid {couleur_a};'
                    f'border:1px solid {couleur_a}44;border-radius:12px;padding:1rem 1.4rem;'
                    f'margin-bottom:.7rem;display:flex;align-items:center;justify-content:space-between">'
                    f'<div style="display:flex;align-items:center;gap:.8rem">'
                    f'<span style="background:{couleur_a};color:#FFF;font-size:.8rem;'
                    f'font-weight:800;padding:4px 12px;border-radius:100px">{badge_a}</span>'
                    f'<span style="font-weight:700;color:#1B2B4B;font-size:1rem">{nom_a}</span>'
                    f'</div>'
                    f'<div style="font-size:1rem;font-weight:700;color:{couleur_a}">{total_a} pcs</div>'
                    f'</div>',
                    unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)

        # ── Rapport WhatsApp ─────────────────────────────────────────────────
        st.markdown("### 📱 Rapport de fin de journée — WhatsApp")
        WA_NUMBER = "212661376059"
        lignes_rapport = [f"📊 *Rapport Latif Shop — {today_str}*", ""]
        lignes_rapport.append(f"💰 Ventes du jour : {len(ventes_today)} transaction(s) — {pieces_today} pièce(s)")
        lignes_rapport.append(f"📅 Ventes de la semaine : {len(ventes_week)} — {pieces_week} pièce(s)")
        lignes_rapport.append("")
        if top3:
            lignes_rapport.append("🏆 Top articles :")
            for i,(n,q) in enumerate(top3):
                lignes_rapport.append(f"  {['🥇','🥈','🥉'][i]} {n} : {q} pcs")
            lignes_rapport.append("")
        if ventes_today:
            lignes_rapport.append("📋 Détail ventes :")
            agg_v = {}
            for v in ventes_today:
                agg_v[v.get("vendeur","?")] = agg_v.get(v.get("vendeur","?"),0) + v.get("quantite",1)
            for vendeur_r, qty_r in sorted(agg_v.items(), key=lambda x:-x[1]):
                lignes_rapport.append(f"  • {vendeur_r} : {qty_r} pcs")
            lignes_rapport.append("")
        if alertes:
            lignes_rapport.append(f"⚠️ {len(alertes)} alerte(s) stock à vérifier")
        lignes_rapport.append("")
        lignes_rapport.append("_Latif Shop — Gestion automatique_")

        message_wa = "\n".join(lignes_rapport)
        wa_url = f"https://wa.me/{WA_NUMBER}?text={urllib.parse.quote(message_wa)}"
        st.markdown(
            f'<a href="{wa_url}" target="_blank">'
            f'<button style="background:#25D366;color:#FFF;border:none;border-radius:10px;'
            f'padding:.8rem 2rem;font-size:.9rem;font-weight:700;cursor:pointer;width:100%">'
            f'📱 Envoyer le rapport sur WhatsApp</button></a>',
            unsafe_allow_html=True)
        with st.expander("👁 Aperçu du message"):
            st.text(message_wa)

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 1 : NOUVELLE GRILLE (admin seulement)
# ══════════════════════════════════════════════════════════════════════════════
if tab_new:
    with tab_new:
        cats = load_categories()

        # ══ PAS DE CATÉGORIE : message et formulaire de création rapide ══
        if not cats:
            st.markdown("""
            <div style="background:#EEF4FF;border:2px solid #2563EB;
            border-radius:12px;padding:1.8rem;text-align:center;margin-bottom:1.5rem">
              <div style="font-size:2.5rem;margin-bottom:.6rem">🗂️</div>
              <div style="color:#1B2B4B;font-size:1.1rem;font-weight:700;margin-bottom:.5rem">
                Aucune catégorie créée
              </div>
              <div style="color:#4A5568;font-size:.88rem">
                Allez dans <strong>⚙ Profils &amp; Config</strong>
                et créez vos catégories (ex: T-shirts Homme, Polos…)
                avant d'ajouter des articles.
              </div>
            </div>""", unsafe_allow_html=True)

        # ══ ÉTAPE 1 : Choisir la catégorie ═══════════════════════════════
        elif not st.session_state.get("new_grid_cat"):
            # ── Produits déjà enregistrés ───────────────────────────────
            stock_ng = load_stock()
            if stock_ng:
                st.markdown(
                    f'<div style="font-size:1.1rem;font-weight:800;color:#1B2B4B;'
                    f'margin-bottom:.8rem">📋 {len(stock_ng)} produit(s) déjà enregistré(s)</div>',
                    unsafe_allow_html=True)
                ng_search = st.text_input("🔍 Rechercher un produit", placeholder="Nom du modèle…", key="ng_search")
                ng_items = [(mk, v) for mk, v in stock_ng.items()
                            if ng_search.lower() in v.get("model_name","").lower()]
                for row_i in range(0, len(ng_items), 4):
                    row_ng = ng_items[row_i:row_i+4]
                    cols_ng = st.columns(4)
                    for ci_ng, (mk_ng, v_ng) in enumerate(row_ng):
                        total_ng = sum(q for cd in v_ng.get("stock",{}).values() for q in cd.values())
                        b64_ng = v_ng.get("b64_thumb","")
                        sbc_ng = "#059669" if total_ng > 5 else "#D97706" if total_ng > 0 else "#DC2626"
                        with cols_ng[ci_ng]:
                            photo_ng = (
                                f'<img src="data:image/jpeg;base64,{b64_ng}" '
                                f'style="width:100%;height:80px;object-fit:cover;'
                                f'border-radius:8px 8px 0 0;display:block">'
                                if b64_ng else
                                f'<div style="width:100%;height:80px;background:#EEF1F7;'
                                f'border-radius:8px 8px 0 0;display:flex;align-items:center;'
                                f'justify-content:center;font-size:1.8rem">👕</div>'
                            )
                            st.markdown(
                                f'<div style="border:1.5px solid #E0E5EF;border-radius:10px;'
                                f'overflow:hidden;margin-bottom:.5rem">'
                                f'{photo_ng}'
                                f'<div style="padding:.4rem .5rem">'
                                f'<div style="font-weight:700;font-size:.75rem;color:#1B2B4B;'
                                f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis">'
                                f'{v_ng.get("model_name",mk_ng)}</div>'
                                f'<div style="font-size:.65rem;color:{sbc_ng};font-weight:600">'
                                f'{total_ng} pcs</div></div></div>',
                                unsafe_allow_html=True)
                st.markdown("<hr>", unsafe_allow_html=True)

            st.markdown('<span class="section-label">Étape 1 — Choisir la catégorie pour le nouveau produit</span>',
                        unsafe_allow_html=True)
            nb_cols = min(len(cats), 5)
            cat_cols = st.columns(nb_cols)
            for i, cat in enumerate(cats):
                with cat_cols[i % nb_cols]:
                    icon  = cat.get("icon", "📦") if isinstance(cat, dict) else "📦"
                    name  = cat.get("name", cat)  if isinstance(cat, dict) else cat
                    color = cat.get("color", "#2563EB") if isinstance(cat, dict) else "#2563EB"
                    st.markdown(
                        f'<div style="background:#F8F9FC;border:2px solid #E0E5EF;'
                        f'border-radius:12px;padding:1rem .6rem;text-align:center;'
                        f'margin-bottom:.4rem">'
                        f'<div style="font-size:2rem">{icon}</div>'
                        f'<div style="font-size:.82rem;font-weight:600;color:#1B2B4B;'
                        f'margin-top:.4rem">{name}</div></div>',
                        unsafe_allow_html=True)
                    if st.button(f"Choisir", key=f"cat_sel_{i}", use_container_width=True):
                        st.session_state["new_grid_cat"] = name
                        st.rerun()

        # ══ ÉTAPE 2 : Formulaire article ═════════════════════════════════
        else:
            selected_cat = st.session_state["new_grid_cat"]

            # ── Clé de réinitialisation du formulaire ─────────────────
            if "grille_articles" not in st.session_state:
                st.session_state.grille_articles = []
            if "form_reset_key" not in st.session_state:
                st.session_state.form_reset_key = 0
            fk = st.session_state.form_reset_key

            # Bandeau catégorie sélectionnée
            cat_obj = next((c for c in cats if
                (c.get("name") if isinstance(c,dict) else c) == selected_cat), None)
            cat_icon  = cat_obj.get("icon","📦")  if isinstance(cat_obj,dict) else "📦"
            cat_color = cat_obj.get("color","#2563EB") if isinstance(cat_obj,dict) else "#2563EB"
            st.markdown(
                f'<div style="background:{cat_color}15;border:1.5px solid {cat_color};'
                f'border-radius:10px;padding:.7rem 1.2rem;display:inline-flex;'
                f'align-items:center;gap:.6rem;margin-bottom:1.2rem">'
                f'<span style="font-size:1.4rem">{cat_icon}</span>'
                f'<span style="color:{cat_color};font-weight:700;font-size:.95rem">'
                f'{selected_cat}</span>'
                f'<span style="color:#8A9AB5;font-size:.75rem;margin-left:.3rem">'
                f'— catégorie sélectionnée</span></div>',
                unsafe_allow_html=True)

            rc1, _ = st.columns([1, 4])
            with rc1:
                if st.button("↩  Changer de catégorie", key="btn_change_cat"):
                    st.session_state["new_grid_cat"] = None
                    st.rerun()

            # ── Produits existants de cette catégorie ─────────────────
            stock_ng2 = load_stock()
            arts_cat_ng = [(mk, v) for mk, v in stock_ng2.items()
                           if v.get("categorie","").strip().lower() == selected_cat.strip().lower()]
            if arts_cat_ng:
                st.markdown(
                    f'<div style="font-size:1rem;font-weight:700;color:#1B2B4B;margin:.6rem 0 .4rem">'
                    f'📦 {len(arts_cat_ng)} produit(s) déjà dans « {selected_cat} »</div>',
                    unsafe_allow_html=True)
                for row_i2 in range(0, len(arts_cat_ng), 5):
                    row_ng2 = arts_cat_ng[row_i2:row_i2+5]
                    cols_ng2 = st.columns(5)
                    for ci2, (mk2, v2) in enumerate(row_ng2):
                        total2 = sum(q for cd in v2.get("stock",{}).values() for q in cd.values())
                        b64_2 = v2.get("b64_thumb","")
                        sbc2 = "#059669" if total2 > 5 else "#D97706" if total2 > 0 else "#DC2626"
                        with cols_ng2[ci2]:
                            photo2 = (
                                f'<img src="data:image/jpeg;base64,{b64_2}" '
                                f'style="width:100%;height:70px;object-fit:cover;'
                                f'border-radius:6px 6px 0 0;display:block">'
                                if b64_2 else
                                f'<div style="width:100%;height:70px;background:#EEF1F7;'
                                f'border-radius:6px 6px 0 0;display:flex;align-items:center;'
                                f'justify-content:center;font-size:1.4rem">👕</div>'
                            )
                            st.markdown(
                                f'<div style="border:1.5px solid #E0E5EF;border-radius:8px;'
                                f'overflow:hidden;margin-bottom:.4rem">'
                                f'{photo2}'
                                f'<div style="padding:.3rem .4rem">'
                                f'<div style="font-weight:700;font-size:.7rem;color:#1B2B4B;'
                                f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis">'
                                f'{v2.get("model_name",mk2)}</div>'
                                f'<div style="font-size:.62rem;color:{sbc2};font-weight:600">'
                                f'{total2} pcs</div></div></div>',
                                unsafe_allow_html=True)

            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown('<span class="section-label">Étape 2 — Photos & informations du modèle</span>',
                        unsafe_allow_html=True)
            cl, cr = st.columns([1,2], gap="large")
            with cl:
                st.markdown('<span class="section-label">Photos des modèles</span>', unsafe_allow_html=True)
                photo_mode = st.radio("",["📁 Depuis la galerie","📷 Prendre une photo"],
                                      horizontal=True, label_visibility="collapsed",
                                      key=f"photo_mode_new_{fk}")
                if photo_mode == "📷 Prendre une photo":
                    cam_img = st.camera_input("", label_visibility="collapsed", key=f"cam_new_{fk}")
                    uploaded_files = [cam_img] if cam_img else []
                else:
                    uploaded_files = st.file_uploader("",type=["jpg","jpeg","png"],
                                                      accept_multiple_files=True,
                                                      label_visibility="collapsed",
                                                      key=f"uploader_new_{fk}")
                if uploaded_files:
                    n=len(uploaded_files)
                    st.markdown(f'<span class="photo-count-badge">{n} photo{"s" if n>1 else ""}</span>',unsafe_allow_html=True)
                    th='<div class="photo-gallery">'
                    for f in uploaded_files:
                        b64=img_to_b64(f); sn=f.name if len(f.name)<=12 else f.name[:10]+"…"
                        th+=f'<div class="photo-thumb"><img src="data:image/png;base64,{b64}"/><div class="photo-name">{sn}</div></div>'
                    st.markdown(th+"</div>",unsafe_allow_html=True)

            with cr:
                st.markdown('<span class="section-label">Nom / Référence</span>', unsafe_allow_html=True)
                model_name = st.text_input("", placeholder="Ex: Veste oversize REF-001",
                                           label_visibility="collapsed",
                                           key=f"model_name_input_{fk}")

                st.markdown('<span class="section-label" style="margin-top:1rem">Tailles disponibles & quantités</span>',
                            unsafe_allow_html=True)
                st.caption("Cochez les tailles puis saisissez la quantité en stock pour chacune.")

                ALL_SIZES = ["XS","S","M","L","XL","XXL","3XL"]
                # Sélection des tailles actives
                selected_sizes = st.multiselect("", options=ALL_SIZES,
                                                default=["S","M","L","XL","XXL"],
                                                label_visibility="collapsed",
                                                key=f"sizes_multisel_{fk}")

                # Cases de quantité pour chaque taille sélectionnée
                size_qtys = {}
                if selected_sizes:
                    st.markdown('<span class="section-label" style="margin-top:.6rem">Quantités par taille</span>',
                                unsafe_allow_html=True)
                    n_sz = len(selected_sizes)
                    sz_cols = st.columns(n_sz)
                    for i, sz in enumerate(selected_sizes):
                        with sz_cols[i]:
                            st.markdown(
                                f'<div style="text-align:center;font-size:.72rem;font-weight:700;'
                                f'color:#1B2B4B;letter-spacing:.1em;margin-bottom:4px">{sz}</div>',
                                unsafe_allow_html=True)
                            size_qtys[sz] = st.number_input(
                                "", min_value=0, max_value=9999, value=0, step=1,
                                key=f"qty_{sz}_{fk}", label_visibility="collapsed")

            st.markdown("<hr>", unsafe_allow_html=True)
            selected_colors = color_palette_picker(key_suffix=str(fk))
            st.markdown("<hr>", unsafe_allow_html=True)

            # ── Bouton Ajouter à la grille ────────────────────────
            add_col, gen_col, clear_col = st.columns([2, 2, 1])
            with add_col:
                if st.button("➕  Ajouter cet article à la grille", key="btn_add_article",
                             use_container_width=True):
                    errs = []
                    if not selected_colors: errs.append("Sélectionnez au moins une couleur.")
                    if not selected_sizes:  errs.append("Sélectionnez au moins une taille.")
                    if errs:
                        for e in errs: st.warning(e)
                    else:
                        photos_data_tmp = []
                        first_thumb = None
                        if uploaded_files:
                            for f in uploaded_files:
                                try:
                                    raw = f.getvalue() if hasattr(f, "getvalue") else f.read()
                                    if not raw:
                                        continue
                                    b64_display = img_to_b64(BytesIO(raw))
                                    fname = getattr(f, "name", "photo")
                                    photos_data_tmp.append({
                                        "name": fname.rsplit(".",1)[0] if "." in fname else fname,
                                        "b64": b64_display
                                    })
                                    if first_thumb is None:
                                        first_thumb = make_thumb(BytesIO(raw))
                                except Exception:
                                    pass
                        article = {
                            "model_name": model_name or "Sans nom",
                            "categorie":  selected_cat,
                            "colors":     list(selected_colors),
                            "sizes":      list(selected_sizes),
                            "size_qtys":  dict(size_qtys),
                            "photos":     photos_data_tmp,
                            "thumb":      first_thumb,
                            "date":       datetime.now().strftime("%d/%m/%Y %H:%M"),
                        }
                        st.session_state.grille_articles.append(article)
                        # ── Réinitialiser tout le formulaire ──────────────
                        st.session_state.form_reset_key += 1
                        new_fk = st.session_state.form_reset_key
                        # Vider aussi la liste des couleurs de l'ancien fk
                        old_colors_key = f"selected_colors_{fk}"
                        if old_colors_key in st.session_state:
                            del st.session_state[old_colors_key]
                        st.success(f"✓ **{article['model_name']}** ajouté — formulaire réinitialisé !")
                        st.rerun()

            with clear_col:
                if st.session_state.grille_articles:
                    if st.button("🗑", key="btn_clear_grille", use_container_width=True,
                                 help="Vider la grille"):
                        st.session_state.grille_articles = []
                        st.rerun()

            # ── Prévisualisation des articles en attente ──────────
            if st.session_state.grille_articles:
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown(
                    f'<div style="background:#EEF4FF;border:1.5px solid #2563EB;'
                    f'border-radius:10px;padding:.8rem 1.2rem;margin-bottom:1rem;'
                    f'display:flex;align-items:center;gap:.8rem">'
                    f'<span style="font-size:1.4rem">📋</span>'
                    f'<span style="color:#1B2B4B;font-weight:700">'
                    f'{len(st.session_state.grille_articles)} article(s) dans la grille</span>'
                    f'</div>', unsafe_allow_html=True)

                for idx, art in enumerate(st.session_state.grille_articles):
                    color_dots = "".join(
                        f'<span style="width:10px;height:10px;border-radius:50%;'
                        f'background:{HEX_MAP.get(c,"#888")};border:1px solid #ccc;'
                        f'display:inline-block;margin-right:2px"></span>'
                        for c in art["colors"][:6])
                    size_tags = "".join(
                        f'<span style="background:#1B2B4B;color:#FFF;font-size:.62rem;'
                        f'padding:2px 7px;border-radius:4px;margin:1px;display:inline-block">'
                        f'{s} <strong style="color:#F0C060">{art["size_qtys"].get(s,0)}</strong>'
                        f'</span>' for s in art["sizes"])
                    thumb_html = ""
                    if art["photos"]:
                        thumb_html = (f'<img src="data:image/png;base64,{art["photos"][0]["b64"]}" '
                                      f'style="width:44px;height:54px;object-fit:cover;'
                                      f'border-radius:6px;border:1px solid #E0E5EF;'
                                      f'margin-right:.6rem;flex-shrink:0">')
                    rc1, rc2 = st.columns([8, 1])
                    with rc1:
                        st.markdown(
                            f'<div style="background:#FFF;border:1px solid #E0E5EF;'
                            f'border-left:3px solid #2563EB;border-radius:8px;'
                            f'padding:.7rem 1rem;display:flex;align-items:center;gap:.6rem;'
                            f'margin-bottom:.4rem">'
                            f'{thumb_html}'
                            f'<div style="flex:1">'
                            f'<div style="font-weight:700;color:#1B2B4B;font-size:.9rem">'
                            f'#{idx+1} — {art["model_name"]}</div>'
                            f'<div style="margin-top:3px">{color_dots}</div>'
                            f'<div style="margin-top:4px">{size_tags}</div>'
                            f'</div></div>', unsafe_allow_html=True)
                    with rc2:
                        if st.button("✕", key=f"del_art_{idx}", use_container_width=True,
                                     help="Retirer cet article"):
                            st.session_state.grille_articles.pop(idx)
                            st.rerun()

                st.markdown("<hr>", unsafe_allow_html=True)

                # ── Bouton Générer la grille complète ─────────────
                with gen_col:
                    if st.button("✦  Enregistrer toute la grille", key="btn_save_grille",
                                 use_container_width=True):
                        stock_data_new = load_stock()
                        saved_names = []
                        for art in st.session_state.grille_articles:
                            mk = model_key(art["model_name"])
                            # Miniature compressée (JPEG 80x100) — petite pour GSheets
                            b64_thumb_art = art.get("thumb") or (
                                art["photos"][0]["b64"] if art.get("photos") else None)
                            # Sauvegarde entrée historique
                            save_entry({
                                "date":       art["date"],
                                "model_name": art["model_name"],
                                "categorie":  art["categorie"],
                                "photos":     [p["name"] for p in art["photos"]],
                                "b64_thumb":  b64_thumb_art,
                                "colors":     art["colors"],
                                "sizes":      art["sizes"],
                            })
                            # Initialise le stock avec les quantités saisies
                            init_stock = {}
                            for c in art["colors"]:
                                init_stock[c] = {s: art["size_qtys"].get(s, 0)
                                                 for s in art["sizes"]}
                            stock_data_new[mk] = {
                                "model_name":  art["model_name"],
                                "categorie":   art["categorie"],
                                "colors":      art["colors"],
                                "sizes":       art["sizes"],
                                "seuil_min":   2,
                                "stock":       init_stock,
                                "last_update": art["date"],
                                "b64_thumb":   b64_thumb_art,
                            }
                            saved_names.append(art["model_name"])
                        save_stock(stock_data_new)

                        # Afficher la grille récap
                        for art in st.session_state.grille_articles:
                            photos_d = art["photos"]
                            sh = "".join(f"<th>{s}</th>" for s in art["sizes"])
                            qty_header = "".join(
                                f'<th style="color:#F0C060">{s}<br>'
                                f'<span style="font-weight:400;font-size:.75em">'
                                f'{art["size_qtys"].get(s,0)} pcs</span></th>'
                                for s in art["sizes"])
                            st.markdown(
                                f'<h3 style="color:#1B2B4B;margin-top:1rem">'
                                f'{art["model_name"]}'
                                f'<span style="font-size:.7rem;color:#8A9AB5;'
                                f'font-weight:400;margin-left:.8rem">'
                                f'📂 {art["categorie"]}</span></h3>',
                                unsafe_allow_html=True)
                            if photos_d:
                                pc = (f'<img src="data:image/png;base64,{photos_d[0]["b64"]}"'
                                      f' width="60" style="border-radius:6px">')
                            else:
                                pc = "<span style='color:#999;font-size:.75rem'>[Photo]</span>"
                            rows = ""
                            for i, color in enumerate(art["colors"]):
                                hx = HEX_MAP.get(color, "#888")
                                cc = (f'<span style="display:inline-flex;align-items:center;gap:6px">'
                                      f'<span style="width:12px;height:12px;border-radius:50%;'
                                      f'background:{hx};border:1px solid #ccc;display:inline-block">'
                                      f'</span>{color}</span>')
                                cells = "".join(
                                    f'<td style="color:#1B2B4B;font-weight:600">'
                                    f'{art["size_qtys"].get(s,0)}</td>'
                                    for s in art["sizes"])
                                if i == 0:
                                    rs = (f'rowspan="{len(art["colors"])}"'
                                          if len(art["colors"]) > 1 else "")
                                    rows += (f'<tr><td {rs} style="vertical-align:middle;'
                                             f'min-width:90px">{pc}</td>'
                                             f'<td style="text-align:left">{cc}</td>'
                                             f'{cells}</tr>')
                                else:
                                    rows += (f'<tr><td style="text-align:left">{cc}</td>'
                                             f'{cells}</tr>')
                            st.markdown(
                                f'<div class="table-wrapper">'
                                f'<table><thead><tr>'
                                f'<th>Photo</th><th>Couleur</th>{qty_header}'
                                f'</tr></thead><tbody>{rows}</tbody></table></div>',
                                unsafe_allow_html=True)

                        st.success(f"✅  {len(saved_names)} article(s) enregistré(s) : "
                                   f"{', '.join(saved_names)}")
                        st.session_state.grille_articles = []
                        st.session_state["new_grid_cat"] = None
                        st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 2 : HISTORIQUE (admin seulement)
# ══════════════════════════════════════════════════════════════════════════════
if tab_hist:
    with tab_hist:
        history=load_history()
        if not history:
            st.info("Aucune grille sauvegardée.")
        else:
            # ── Filtres ────────────────────────────────────────────────────
            cats_h = load_categories()
            cat_names_h = ["Toutes"] + [
                (c.get("name",c) if isinstance(c,dict) else c) for c in cats_h
            ] + ["Non classé"]
            f1, f2, f3 = st.columns([2,2,1])
            with f1:
                fil_cat = st.selectbox("📂 Catégorie", cat_names_h, key="hist_cat_fil")
            with f2:
                srch = st.text_input("🔍 Rechercher…", placeholder="Nom du modèle", key="hist_srch")
            with f3:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                df_e = pd.DataFrame([{"Date":e["date"],"Catégorie":e.get("categorie","Non classé"),
                    "Modèle":e["model_name"],"Couleurs":", ".join(e["colors"]),
                    "Tailles":", ".join(e["sizes"])} for e in history])
                st.download_button("📥 CSV", df_e.to_csv(index=False).encode("utf-8-sig"),
                    f"historique_{datetime.now().strftime('%Y%m%d')}.csv","text/csv",
                    use_container_width=True)
            st.markdown("<hr>", unsafe_allow_html=True)

            # Applique filtres
            filt = history
            if fil_cat != "Toutes":
                if fil_cat == "Non classé":
                    filt = [e for e in filt if not e.get("categorie")]
                else:
                    filt = [e for e in filt if e.get("categorie","") == fil_cat]
            if srch:
                filt = [e for e in filt if srch.lower() in e["model_name"].lower()]

            st.markdown(f'<span style="color:var(--text3);font-size:.8rem">'
                        f'{len(filt)} article(s) affiché(s)</span>', unsafe_allow_html=True)

            for idx,entry in enumerate(filt):
                cat_label = entry.get("categorie","Non classé")
                cdots="".join(
                    f'<span style="display:inline-flex;align-items:center;gap:3px;margin:2px 4px 2px 0">'
                    f'<span style="width:10px;height:10px;border-radius:50%;background:{HEX_MAP.get(c,"#888")};'
                    f'border:1.5px solid rgba(255,255,255,.15);display:inline-block"></span>'
                    f'<span style="font-size:.68rem;color:var(--text2)">{c}</span></span>'
                    for c in entry["colors"])
                size_tags="".join(f'<span class="tag tag-dark">{s}</span>' for s in entry["sizes"])
                st.markdown(
                    f'<div class="history-card">'
                    f'<div style="display:flex;align-items:center;justify-content:space-between">'
                    f'<div class="history-date">📅 {entry["date"]}</div>'
                    f'<span style="background:var(--gold-soft);border:1px solid var(--gold);'
                    f'color:var(--gold);font-size:.62rem;padding:2px 8px;border-radius:100px;'
                    f'font-weight:600">📂 {cat_label}</span></div>'
                    f'<div class="history-title">{entry["model_name"]}</div>'
                    f'<div style="margin-top:6px">{cdots}</div>'
                    f'<div style="margin-top:4px">{size_tags}</div>'
                    f'</div>', unsafe_allow_html=True)
                if st.button("Supprimer",key=f"dh_{idx}"):
                    ri=history.index(entry) if entry in history else -1
                    delete_history(ri); st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 3 : STOCK & COMMANDES (admin seulement)
# ══════════════════════════════════════════════════════════════════════════════
if tab_stock:
    with tab_stock:
        history_s=load_history(); stock_data=load_stock()
        cats_s = load_categories()
        if not history_s:
            st.info("Créez d'abord une grille dans l'onglet « Nouvelle grille ».")
        else:
            # ── Session state navigation ──────────────────────────────────
            for sk, sv in [("stock_edit_values", {}),
                            ("stock_pending_confirm", False),
                            ("stock_last_saved", ""),
                            ("stk_cat", None),
                            ("stk_art", None)]:
                if sk not in st.session_state:
                    st.session_state[sk] = sv

            # Index mk → entry historique
            seen = {}
            for e in history_s:
                k = model_key(e["model_name"])
                if k not in seen:
                    seen[k] = e

            # ── Fil d'Ariane ─────────────────────────────────────────────
            bc_parts = ["📦 Stock"]
            if st.session_state.stk_cat:
                bc_parts.append(st.session_state.stk_cat)
            if st.session_state.stk_art:
                art_name_bc = stock_data.get(st.session_state.stk_art,{}).get("model_name","")
                bc_parts.append(art_name_bc)
            st.markdown(
                f'<div style="font-size:.72rem;color:#8A9AB5;margin-bottom:.6rem">'
                + " › ".join(
                    f'<span style="color:#1B2B4B;font-weight:600">{p}</span>'
                    if i == len(bc_parts)-1
                    else f'<span>{p}</span>'
                    for i, p in enumerate(bc_parts))
                + '</div>', unsafe_allow_html=True)

            if st.session_state.stk_cat and not st.session_state.stk_art:
                if st.button("← Retour catégories", key="stk_back_cat0"):
                    st.session_state.stk_cat = None
                    st.rerun()
            elif st.session_state.stk_art:
                n1, n2 = st.columns([1,1])
                with n1:
                    if st.button("← Retour articles", key="stk_back_art0"):
                        st.session_state.stk_art = None
                        st.session_state.stock_edit_values = {}
                        st.session_state.stock_pending_confirm = False
                        st.rerun()
                with n2:
                    if st.button("⌂ Retour catégories", key="stk_back_cat1"):
                        st.session_state.stk_cat = None
                        st.session_state.stk_art = None
                        st.session_state.stock_edit_values = {}
                        st.session_state.stock_pending_confirm = False
                        st.rerun()

            # ════════════════════════════════════════════════════════════
            # ÉTAPE A — Choisir une catégorie
            # ════════════════════════════════════════════════════════════
            if st.session_state.stk_cat is None:
                st.markdown(
                    '<span class="section-label">Choisissez une catégorie</span>',
                    unsafe_allow_html=True)

                # Construire les catégories présentes dans le stock
                cat_map = {}   # cat_name → {"nb": int, "mks": set}
                for mk, entry in seen.items():
                    cat_e = entry.get("categorie","") or "Non classé"
                    if cat_e not in cat_map:
                        cat_map[cat_e] = {"nb": 0, "mks": set()}
                    cat_map[cat_e]["nb"]  += 1
                    cat_map[cat_e]["mks"].add(mk)

                # "Toutes" en premier
                all_cats_stk = [("Toutes", {"nb": len(seen), "mks": set(seen.keys())})] \
                               + list(cat_map.items())

                clicked_stk_cat = None
                NB_CAT = 3
                for ri in range(0, len(all_cats_stk), NB_CAT):
                    cc0, cc1, cc2 = st.columns(3)
                    for ci_c, ccol in enumerate([cc0, cc1, cc2]):
                        if ri + ci_c >= len(all_cats_stk):
                            break
                        cat_nm_s, cat_info = all_cats_stk[ri + ci_c]
                        nb_arts = cat_info["nb"]
                        # Icône catégorie
                        cat_obj_s = next(
                            (c for c in cats_s
                             if (c.get("name") if isinstance(c,dict) else c) == cat_nm_s),
                            {})
                        cat_icon_s = (cat_obj_s.get("icon","📦")
                                      if isinstance(cat_obj_s,dict) else "📦")
                        if cat_nm_s == "Toutes":
                            cat_icon_s = "🗂️"

                        # Total stock de la catégorie
                        tot_stk_cat = sum(
                            sum(v for cd in stock_data.get(mk,{}).get("stock",{}).values()
                                for v in cd.values())
                            for mk in cat_info["mks"])
                        badge_c_s = ("#059669" if tot_stk_cat > 10
                                     else "#D97706" if tot_stk_cat > 0 else "#DC2626")

                        with ccol:
                            st.markdown(
                                f'<div style="background:#FFFFFF;border:2px solid #E0E5EF;'
                                f'border-radius:14px;padding:1.2rem .8rem 1rem;'
                                f'text-align:center;margin-bottom:.5rem;'
                                f'box-shadow:0 2px 8px rgba(27,43,75,.06)">'
                                f'<div style="font-size:2.4rem;margin-bottom:.5rem">'
                                f'{cat_icon_s}</div>'
                                f'<div style="font-weight:800;color:#1B2B4B;font-size:.92rem;'
                                f'margin-bottom:.25rem">{cat_nm_s}</div>'
                                f'<div style="font-size:.68rem;color:#8A9AB5;'
                                f'margin-bottom:.3rem">{nb_arts} article(s)</div>'
                                f'<div style="font-size:.72rem;font-weight:700;'
                                f'color:{badge_c_s}">{tot_stk_cat} pcs en stock</div>'
                                f'</div>', unsafe_allow_html=True)
                            if st.button(f"Voir {cat_nm_s}", key=f"stk_cat_{cat_nm_s}",
                                         use_container_width=True):
                                clicked_stk_cat = cat_nm_s

                if clicked_stk_cat:
                    st.session_state.stk_cat = clicked_stk_cat
                    st.session_state.stk_art = None
                    st.rerun()

            # ════════════════════════════════════════════════════════════
            # ÉTAPE B — Grille photos articles de la catégorie
            # ════════════════════════════════════════════════════════════
            elif st.session_state.stk_art is None:
                cat_sel_s = st.session_state.stk_cat
                st.markdown(
                    f'<span class="section-label">Articles — {cat_sel_s}</span>',
                    unsafe_allow_html=True)

                # Filtrer les articles
                arts_stk = []
                for mk, entry in seen.items():
                    cat_e = entry.get("categorie","") or "Non classé"
                    if cat_sel_s != "Toutes" and cat_e != cat_sel_s:
                        continue
                    arts_stk.append((mk, entry))

                if not arts_stk:
                    st.info(f"Aucun article dans « {cat_sel_s} ».")
                else:
                    clicked_stk_art = None
                    NB_ACOLS = 3
                    for ri in range(0, len(arts_stk), NB_ACOLS):
                        a0, a1, a2 = st.columns(3)
                        for ci_a, acol in enumerate([a0, a1, a2]):
                            if ri + ci_a >= len(arts_stk):
                                break
                            mk_a, entry_a = arts_stk[ri + ci_a]
                            sd_a   = stock_data.get(mk_a, {})
                            stk_a  = sd_a.get("stock", {})
                            total_a = sum(v for cd in stk_a.values() for v in cd.values())
                            colors_a = sd_a.get("colors", [])

                            b64_a = entry_a.get("b64_thumb") or sd_a.get("b64_thumb")
                            if b64_a:
                                img_a = (f'<img src="data:image/png;base64,{b64_a}" '
                                         f'style="width:100%;height:130px;'
                                         f'object-fit:cover;display:block">')
                            else:
                                img_a = (f'<div style="width:100%;height:130px;'
                                         f'background:linear-gradient(135deg,#EEF1F7,#E0E5EF);'
                                         f'display:flex;align-items:center;'
                                         f'justify-content:center;font-size:3rem">👕</div>')

                            badge_a = ("#059669" if total_a > 10
                                       else "#D97706" if total_a > 0 else "#DC2626")
                            badge_t = ("En stock" if total_a > 10
                                       else "Faible" if total_a > 0 else "Rupture")

                            dots = "".join(
                                f'<span style="width:11px;height:11px;border-radius:50%;'
                                f'background:{HEX_MAP.get(c,"#888")};'
                                f'border:1.5px solid rgba(0,0,0,.12);'
                                f'display:inline-block;margin-right:3px"></span>'
                                for c in colors_a[:5])

                            with acol:
                                st.markdown(
                                    f'<div style="background:#FFFFFF;'
                                    f'border:2px solid #E0E5EF;border-radius:14px;'
                                    f'overflow:hidden;margin-bottom:.5rem;'
                                    f'box-shadow:0 2px 8px rgba(27,43,75,.06)">'
                                    f'<div style="position:relative;overflow:hidden">'
                                    f'{img_a}'
                                    f'<span style="position:absolute;top:7px;right:7px;'
                                    f'background:{badge_a};color:#FFF;font-size:.55rem;'
                                    f'font-weight:700;padding:2px 8px;border-radius:20px;'
                                    f'text-transform:uppercase">{badge_t}</span></div>'
                                    f'<div style="padding:.65rem .8rem .6rem">'
                                    f'<div style="font-weight:800;color:#1B2B4B;'
                                    f'font-size:.88rem;white-space:nowrap;overflow:hidden;'
                                    f'text-overflow:ellipsis;margin-bottom:.25rem">'
                                    f'{entry_a["model_name"]}</div>'
                                    f'<div style="margin-bottom:.3rem">{dots}</div>'
                                    f'<div style="font-size:.68rem;color:#8A9AB5">'
                                    f'{total_a} pcs · {len(colors_a)} couleur(s)</div>'
                                    f'</div></div>', unsafe_allow_html=True)
                                if st.button("📊 Voir la situation",
                                             key=f"stk_art_{mk_a}",
                                             use_container_width=True):
                                    clicked_stk_art = mk_a

                    if clicked_stk_art:
                        st.session_state.stk_art = clicked_stk_art
                        st.session_state.stock_edit_values = {}
                        st.session_state.stock_pending_confirm = False
                        st.rerun()

            # ════════════════════════════════════════════════════════════
            # ÉTAPE C — Situation + édition stock de l'article
            # ════════════════════════════════════════════════════════════
            else:
                mk  = st.session_state.stk_art
                sel_m = stock_data.get(mk, {}).get("model_name", mk)
                me  = seen.get(mk, {})
                model_names = [sel_m]   # pour compatibilité avec la suite

                # ── Carte visuelle article ────────────────────────────────
                colors = me.get("colors", stock_data.get(mk,{}).get("colors",[]))
                sizes  = me.get("sizes",  stock_data.get(mk,{}).get("sizes", []))
                ex_s   = stock_data.get(mk, {}).get("stock", {})
                cur_s  = stock_data.get(mk, {}).get("seuil_min", 2)
                cat_d  = me.get("categorie","") or "Non classé"
                total_d = sum(v for cd in ex_s.values() for v in cd.values())
                badge_cd = ("#059669" if total_d > 10
                            else "#D97706" if total_d > 0 else "#DC2626")
                badge_td = ("En stock" if total_d > 10
                            else "Stock faible" if total_d > 0 else "Rupture totale")

                # Actions en-tête
                hdr_col2, hdr_col3 = st.columns([1, 1])
                with hdr_col2:
                    if st.button("🔄  Actualiser", key="btn_refresh_stock",
                                 use_container_width=True):
                        st.session_state.stock_edit_values = {}
                        st.session_state.stock_pending_confirm = False
                        st.rerun()
                with hdr_col3:
                    if st.button("🗑️  Supprimer l'article", key="btn_del_stock",
                                 use_container_width=True):
                        if mk in stock_data:
                            del stock_data[mk]; save_stock(stock_data)
                        h_all = load_history()
                        h_all = [e for e in h_all if model_key(e["model_name"]) != mk]
                        _save(SAVE_FILE, h_all)
                        st.success(f"✓ Article « {sel_m} » supprimé.")
                        st.session_state.stk_art = None
                        st.session_state.stock_edit_values = {}
                        st.rerun()

                # Photo + infos article
                b64_det = me.get("b64_thumb") or stock_data.get(mk,{}).get("b64_thumb")
                if b64_det:
                    ph_html = (f'<div style="width:150px;min-height:150px;flex-shrink:0;'
                               f'overflow:hidden;position:relative">'
                               f'<img src="data:image/png;base64,{b64_det}" '
                               f'style="width:150px;min-height:150px;object-fit:cover;'
                               f'display:block"></div>')
                else:
                    ph_html = (f'<div style="width:150px;min-height:150px;flex-shrink:0;'
                               f'background:linear-gradient(135deg,#EEF1F7,#E0E5EF);'
                               f'display:flex;align-items:center;justify-content:center;'
                               f'font-size:4rem">👕</div>')

                color_pills = "".join(
                    f'<div style="display:flex;flex-direction:column;align-items:center;'
                    f'gap:3px;min-width:50px">'
                    f'<div style="width:30px;height:30px;border-radius:50%;'
                    f'background:{HEX_MAP.get(c,"#888")};border:3px solid #FFF;'
                    f'box-shadow:0 2px 5px rgba(0,0,0,.18)"></div>'
                    f'<span style="font-size:.58rem;color:#4A5568;font-weight:600;'
                    f'text-align:center;max-width:50px;overflow:hidden;'
                    f'text-overflow:ellipsis;white-space:nowrap">{c}</span>'
                    f'<span style="font-size:.65rem;font-weight:700;'
                    f'color:{"#059669" if sum(ex_s.get(c,{}).values())>0 else "#DC2626"}">'
                    f'{sum(ex_s.get(c,{}).values())} pcs</span></div>'
                    for c in colors)

                st.markdown(
                    f'<div style="display:flex;gap:1rem;align-items:stretch;'
                    f'background:#FFFFFF;border:2px solid #E0E5EF;border-radius:16px;'
                    f'overflow:hidden;margin:.6rem 0 1.2rem;'
                    f'box-shadow:0 4px 16px rgba(27,43,75,.08)">'
                    f'{ph_html}'
                    f'<div style="flex:1;padding:1rem 1.1rem">'
                    f'<div style="font-size:.6rem;color:var(--gold);font-weight:700;'
                    f'text-transform:uppercase;letter-spacing:.1em;margin-bottom:.25rem">'
                    f'📂 {cat_d}</div>'
                    f'<div style="font-size:1.25rem;font-weight:800;color:#1B2B4B;'
                    f'margin-bottom:.25rem">{sel_m}</div>'
                    f'<div style="font-size:.7rem;color:#8A9AB5;margin-bottom:.6rem">'
                    f'{len(colors)} couleur(s) · {len(sizes)} taille(s) · '
                    f'<strong style="color:{badge_cd}">{total_d} pcs</strong>'
                    f'&nbsp;<span style="background:{badge_cd};color:#FFF;font-size:.55rem;'
                    f'font-weight:700;padding:2px 7px;border-radius:20px;'
                    f'text-transform:uppercase;vertical-align:middle">{badge_td}</span></div>'
                    f'<div style="display:flex;flex-wrap:wrap;gap:.5rem">{color_pills}</div>'
                    f'</div></div>', unsafe_allow_html=True)

                # Grille stock couleurs × tailles (colorée)
                st.markdown('<span class="section-label">Situation du stock</span>',
                            unsafe_allow_html=True)

                th_s = "".join(
                    f'<th style="padding:8px 6px;text-align:center;font-size:.65rem;'
                    f'color:#8A9AB5;text-transform:uppercase;letter-spacing:.08em;'
                    f'border-bottom:2px solid #EEF1F7;border-right:1px solid #EEF1F7;'
                    f'min-width:46px">{sz}</th>'
                    for sz in sizes)
                rows_sit = ""
                for ci_s, c_s in enumerate(colors):
                    hx_s   = HEX_MAP.get(c_s, "#888")
                    row_bg = "#FFFFFF" if ci_s % 2 == 0 else "#F8F9FC"
                    cells_s = ""
                    for sz_s in sizes:
                        q_s = ex_s.get(c_s, {}).get(sz_s, 0)
                        if q_s == 0:
                            cbg_s="#FFE4E4"; cq_s="#DC2626"
                        elif q_s <= cur_s:
                            cbg_s="#FFF8E1"; cq_s="#D97706"
                        else:
                            cbg_s="#F0FFF4"; cq_s="#059669"
                        cells_s += (
                            f'<td style="padding:8px 4px;text-align:center;'
                            f'background:{cbg_s};border-bottom:1px solid #EEF1F7;'
                            f'border-right:1px solid #EEF1F7">'
                            f'<div style="font-size:.9rem;font-weight:800;color:{cq_s}">'
                            f'{q_s}</div></td>')
                    rows_sit += (
                        f'<tr><td style="padding:8px 10px;border-bottom:1px solid #EEF1F7;'
                        f'border-right:2px solid #EEF1F7;background:{row_bg};white-space:nowrap">'
                        f'<span style="display:inline-flex;align-items:center;gap:7px">'
                        f'<span style="width:13px;height:13px;border-radius:50%;'
                        f'background:{hx_s};flex-shrink:0;'
                        f'border:1.5px solid rgba(0,0,0,.12)"></span>'
                        f'<span style="font-size:.8rem;font-weight:600;color:#1B2B4B">'
                        f'{c_s}</span></span></td>{cells_s}</tr>')

                st.markdown(
                    f'<div style="background:#FFFFFF;border:2px solid #E0E5EF;'
                    f'border-radius:14px;overflow:hidden;margin-bottom:1rem;'
                    f'box-shadow:0 2px 10px rgba(27,43,75,.07)">'
                    f'<div style="display:flex;gap:.8rem;padding:.4rem .9rem;'
                    f'background:#F4F6FA;border-bottom:1px solid #EEF1F7;'
                    f'font-size:.6rem;color:#8A9AB5">'
                    f'<span style="display:flex;align-items:center;gap:4px">'
                    f'<span style="width:10px;height:10px;border-radius:2px;'
                    f'background:#F0FFF4;border:1px solid #059669"></span>OK</span>'
                    f'<span style="display:flex;align-items:center;gap:4px">'
                    f'<span style="width:10px;height:10px;border-radius:2px;'
                    f'background:#FFF8E1;border:1px solid #D97706"></span>Faible</span>'
                    f'<span style="display:flex;align-items:center;gap:4px">'
                    f'<span style="width:10px;height:10px;border-radius:2px;'
                    f'background:#FFE4E4;border:1px solid #DC2626"></span>Rupture</span>'
                    f'</div>'
                    f'<div style="overflow-x:auto">'
                    f'<table style="width:100%;border-collapse:collapse">'
                    f'<thead><tr style="background:#F4F6FA">'
                    f'<th style="padding:8px 10px;text-align:left;font-size:.62rem;'
                    f'color:#8A9AB5;text-transform:uppercase;letter-spacing:.1em;'
                    f'border-bottom:2px solid #EEF1F7;border-right:2px solid #EEF1F7">'
                    f'Couleur</th>{th_s}</tr></thead>'
                    f'<tbody>{rows_sit}</tbody>'
                    f'</table></div></div>', unsafe_allow_html=True)

                st.markdown('<span class="section-label">Ajuster les quantités</span>',
                            unsafe_allow_html=True)

                seuil = st.number_input("🔔 Seuil minimum (alerte commande)",
                                        min_value=0, max_value=100, value=cur_s,
                                        step=1, key="seuil_in")

                edit_key = f"edit_{mk}"
                if edit_key not in st.session_state.stock_edit_values:
                    st.session_state.stock_edit_values[edit_key] = {
                        c: {sz: ex_s.get(c, {}).get(sz, 0) for sz in sizes}
                        for c in colors}
                vals = st.session_state.stock_edit_values[edit_key]

                # En-tête grille +/-
                hcols = st.columns([2] + [1]*len(sizes))
                hcols[0].markdown(
                    '<div style="font-size:.68rem;letter-spacing:.15em;'
                    'text-transform:uppercase;color:var(--text3);padding:.4rem 0">'
                    'Couleur</div>', unsafe_allow_html=True)
                for i, sz in enumerate(sizes):
                    hcols[i+1].markdown(
                        f'<div style="text-align:center;font-size:.7rem;'
                        f'letter-spacing:.12em;text-transform:uppercase;'
                        f'color:var(--gold);font-weight:600;padding:.4rem 0">{sz}</div>',
                        unsafe_allow_html=True)

                changed = False
                for c in colors:
                    hx = HEX_MAP.get(c, "#888")
                    rcols = st.columns([2] + [1]*len(sizes))
                    with rcols[0]:
                        st.markdown(
                            f'<div style="display:flex;align-items:center;gap:8px;'
                            f'padding:.5rem 0">'
                            f'<span style="width:14px;height:14px;border-radius:50%;'
                            f'background:{hx};border:1.5px solid rgba(0,0,0,.15);'
                            f'display:inline-block;flex-shrink:0"></span>'
                            f'<span style="font-size:.84rem;color:var(--text)">{c}</span>'
                            f'</div>', unsafe_allow_html=True)
                    for i, sz in enumerate(sizes):
                        with rcols[i+1]:
                            cur_val = vals[c][sz]
                            b_col1, v_col, b_col2 = st.columns([1,1,1])
                            with b_col1:
                                if st.button("−", key=f"minus_{mk}_{c}_{sz}"):
                                    vals[c][sz] = max(0, cur_val - 1)
                                    st.session_state.stock_pending_confirm = False
                                    changed = True
                            with v_col:
                                new_val = st.number_input("", value=vals[c][sz],
                                    min_value=0, max_value=9999, step=1,
                                    key=f"nv_{mk}_{c}_{sz}",
                                    label_visibility="collapsed")
                                if new_val != vals[c][sz]:
                                    vals[c][sz] = new_val
                                    st.session_state.stock_pending_confirm = False
                                    changed = True
                            with b_col2:
                                if st.button("＋", key=f"plus_{mk}_{c}_{sz}"):
                                    vals[c][sz] = cur_val + 1
                                    st.session_state.stock_pending_confirm = False
                                    changed = True

                st.markdown("<hr>", unsafe_allow_html=True)
                act_cols = st.columns([1, 1, 1])
                with act_cols[0]:
                    if st.button("👁  Vérifier", key="btn_verify",
                                 use_container_width=True):
                        st.session_state.stock_pending_confirm = True
                with act_cols[1]:
                    if st.button("💾  Enregistrer la grille", key="btn_save_direct",
                                 use_container_width=True, type="primary"):
                        ns = {c: {sz: int(vals[c][sz]) for sz in sizes} for c in colors}
                        stock_data.setdefault(mk, {}).update({
                            "model_name": sel_m, "colors": colors, "sizes": sizes,
                            "seuil_min": int(seuil), "stock": ns,
                            "last_update": datetime.now().strftime("%d/%m/%Y %H:%M")})
                        save_stock(stock_data)
                        st.session_state.stock_pending_confirm = False
                        st.session_state.stock_edit_values.pop(edit_key, None)
                        st.success(f"✓ Stock enregistré pour **{sel_m}**")
                        st.rerun()

                if st.session_state.stock_pending_confirm:
                    st.markdown(
                        '<div style="background:var(--surface2);border:2px solid var(--gold);'
                        'border-radius:var(--radius);padding:1.2rem;margin:.8rem 0">'
                        '<div style="color:var(--gold);font-weight:700;font-size:.85rem;'
                        'letter-spacing:.12em;text-transform:uppercase;margin-bottom:.8rem">'
                        '✅ Récapitulatif — Vérifiez avant de confirmer</div>',
                        unsafe_allow_html=True)
                    total_pieces = 0; rows_html = ""
                    for c in colors:
                        hx = HEX_MAP.get(c,"#888")
                        cells = "".join(
                            f'<td style="padding:8px 12px;border:1px solid var(--border);'
                            f'text-align:center;background:'
                            f'{"rgba(212,168,67,.08)" if vals[c][s]>0 else "var(--surface2)"}">'
                            f'<strong>{vals[c][s]}</strong></td>' for s in sizes)
                        total_pieces += sum(vals[c][s] for s in sizes)
                        rows_html += (
                            f'<tr><td style="padding:8px 12px;border:1px solid var(--border)">'
                            f'<span style="display:inline-flex;align-items:center;gap:6px">'
                            f'<span style="width:12px;height:12px;border-radius:50%;'
                            f'background:{hx};display:inline-block"></span>{c}</span>'
                            f'</td>{cells}</tr>')
                    sh = "".join(
                        f'<th style="padding:10px 14px;background:var(--surface3);'
                        f'color:var(--gold);border:1px solid var(--border);'
                        f'font-size:.7rem;text-transform:uppercase">{s}</th>' for s in sizes)
                    st.markdown(
                        f'<table style="width:100%;border-collapse:collapse;margin:.6rem 0">'
                        f'<thead><tr><th style="padding:10px 14px;background:var(--surface3);'
                        f'color:var(--gold);border:1px solid var(--border);'
                        f'font-size:.7rem;text-transform:uppercase">Couleur</th>'
                        f'{sh}</tr></thead><tbody>{rows_html}</tbody></table>'
                        f'<div style="text-align:right;font-size:.78rem;margin:.4rem 0">'
                        f'Total : <strong style="color:var(--gold)">'
                        f'{total_pieces} pièces</strong></div></div>',
                        unsafe_allow_html=True)
                    cf1, cf2 = st.columns(2)
                    with cf1:
                        if st.button("✅  Confirmer & Enregistrer",
                                     key="btn_confirm_save",
                                     use_container_width=True):
                            ns = {c: {sz: int(vals[c][sz]) for sz in sizes} for c in colors}
                            stock_data.setdefault(mk, {}).update({
                                "model_name": sel_m, "colors": colors, "sizes": sizes,
                                "seuil_min": int(seuil), "stock": ns,
                                "last_update": datetime.now().strftime("%d/%m/%Y %H:%M")})
                            save_stock(stock_data)
                            st.session_state.stock_pending_confirm = False
                            st.session_state.stock_edit_values.pop(edit_key, None)
                            st.success(f"✓ Stock enregistré pour **{sel_m}**")
                            st.rerun()
                    with cf2:
                        if st.button("✏️  Modifier encore",
                                     key="btn_cancel_confirm",
                                     use_container_width=True):
                            st.session_state.stock_pending_confirm = False
                            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET COMMANDES (admin seulement)
# ══════════════════════════════════════════════════════════════════════════════
if tab_commandes:
    with tab_commandes:
        history_s=load_history(); stock_data=load_stock()
        st.markdown("### 🛒 Liste de commandes")

        # ── Session state navigation commandes ──────────────────────
        for _k,_v in [("cmd_cat", None), ("cmd_art", None)]:
            if _k not in st.session_state:
                st.session_state[_k] = _v

        orders=get_order_list(stock_data)
        all_orders_tp = sum(o["À commander"] for o in orders)

        # Index mk → entry historique
        mk_idx_cmd = {model_key(e["model_name"]): e for e in history_s}

        # ── Métriques globales toujours visibles ─────────────────────
        if orders:
            m1,m2,m3=st.columns(3)
            m1.metric("Références", len(orders))
            m2.metric("Pièces totales", all_orders_tp)
            m3.metric("Modèles", len(set(o["Modèle"] for o in orders)))
            st.markdown(
                '<div style="font-size:.7rem;color:#8A9AB5;margin:.3rem 0 .8rem">'
                '🟢 OK &nbsp;|&nbsp; 🟡 Stock faible &nbsp;|&nbsp; 🔴 Rupture</div>',
                unsafe_allow_html=True)
        else:
            st.success("✅ Stock suffisant — aucune commande nécessaire !")

        # ════════════════════════════════════════════════════════
        # ÉTAPE A — Choisir une catégorie
        # ════════════════════════════════════════════════════════
        if st.session_state.cmd_cat is None:
            if orders:
                # Construire la liste des catégories qui ont des articles à commander
                cats_cmd = {}   # cat_name → liste de mk
                for o in orders:
                    mk_o = model_key(o["Modèle"])
                    e_o  = mk_idx_cmd.get(mk_o, {})
                    cat_o = e_o.get("categorie","") or "Non classé"
                    cats_cmd.setdefault(cat_o, set()).add(mk_o)

                st.markdown(
                    '<span class="section-label">Choisissez une catégorie</span>',
                    unsafe_allow_html=True)

                clicked_cmd_cat = None
                NB_CAT_COLS = 3
                cats_list = list(cats_cmd.items())
                for ri in range(0, len(cats_list), NB_CAT_COLS):
                    ca0, ca1, ca2 = st.columns(3)
                    for ci_cat, cc in enumerate([ca0, ca1, ca2]):
                        if ri + ci_cat >= len(cats_list):
                            break
                        cat_nm, mks_set = cats_list[ri + ci_cat]
                        nb_models = len(mks_set)
                        nb_refs   = sum(1 for o in orders
                                        if (mk_idx_cmd.get(model_key(o["Modèle"]),{})
                                            .get("categorie","Non classé") or "Non classé")
                                           == cat_nm)
                        # Icône de la catégorie
                        cat_obj = next((c for c in cats_s
                                        if (c.get("name") if isinstance(c,dict) else c)
                                           == cat_nm), {})
                        cat_icon = cat_obj.get("icon","📦") if isinstance(cat_obj,dict) else "📦"

                        with cc:
                            st.markdown(
                                f'<div style="background:#FFFFFF;border:2px solid #E0E5EF;'
                                f'border-radius:14px;padding:1.1rem .8rem;text-align:center;'
                                f'margin-bottom:.5rem;'
                                f'box-shadow:0 2px 8px rgba(27,43,75,.06)">'
                                f'<div style="font-size:2.2rem;margin-bottom:.4rem">'
                                f'{cat_icon}</div>'
                                f'<div style="font-weight:800;color:#1B2B4B;font-size:.9rem;'
                                f'margin-bottom:.3rem">{cat_nm}</div>'
                                f'<div style="font-size:.68rem;color:#8A9AB5">'
                                f'{nb_models} modèle(s)</div>'
                                f'<div style="font-size:.68rem;color:#DC2626;font-weight:700;'
                                f'margin-top:.2rem">{nb_refs} réf. à commander</div>'
                                f'</div>', unsafe_allow_html=True)
                            if st.button(f"Voir {cat_nm}", key=f"cmd_cat_{cat_nm}",
                                         use_container_width=True):
                                clicked_cmd_cat = cat_nm

                if clicked_cmd_cat:
                    st.session_state.cmd_cat = clicked_cmd_cat
                    st.session_state.cmd_art = None
                    st.rerun()

        # ════════════════════════════════════════════════════════
        # ÉTAPE B — Choisir un article dans la catégorie
        # ════════════════════════════════════════════════════════
        elif st.session_state.cmd_art is None:
            cat_sel_cmd = st.session_state.cmd_cat
            # Bouton retour
            if st.button("← Retour catégories", key="cmd_back_cat"):
                st.session_state.cmd_cat = None
                st.rerun()

            st.markdown(
                f'<div style="font-size:.78rem;color:#8A9AB5;margin:.4rem 0 .7rem">'
                f'📂 {cat_sel_cmd}</div>', unsafe_allow_html=True)
            st.markdown('<span class="section-label">Choisissez un article</span>',
                        unsafe_allow_html=True)

            # Articles de cette catégorie qui ont des commandes
            arts_cmd = []
            seen_cmd = set()
            for o in orders:
                mk_o = model_key(o["Modèle"])
                if mk_o in seen_cmd:
                    continue
                e_o   = mk_idx_cmd.get(mk_o, {})
                cat_o = e_o.get("categorie","") or "Non classé"
                if cat_o != cat_sel_cmd:
                    continue
                seen_cmd.add(mk_o)
                arts_cmd.append((mk_o, o["Modèle"]))

            clicked_cmd_art = None
            NB_ART_COLS = 3
            for ri in range(0, len(arts_cmd), NB_ART_COLS):
                a0, a1, a2 = st.columns(3)
                for ci_art, ac in enumerate([a0, a1, a2]):
                    if ri + ci_art >= len(arts_cmd):
                        break
                    mk_a, name_a = arts_cmd[ri + ci_art]
                    entry_a = mk_idx_cmd.get(mk_a, {})
                    sd_a    = stock_data.get(mk_a, {})
                    seuil_a = sd_a.get("seuil_min", 0)
                    stk_a   = sd_a.get("stock", {})
                    total_a = sum(v for cd in stk_a.values() for v in cd.values())
                    cmd_a   = sum(
                        max(0, seuil_a - stk_a.get(c,{}).get(s,0))
                        for c in sd_a.get("colors",[])
                        for s in sd_a.get("sizes",[]))
                    b64_a   = entry_a.get("b64_thumb") or sd_a.get("b64_thumb")

                    if b64_a:
                        img_a = (f'<img src="data:image/png;base64,{b64_a}" '
                                 f'style="width:100%;height:120px;object-fit:cover;'
                                 f'display:block">')
                    else:
                        img_a = (f'<div style="width:100%;height:120px;'
                                 f'background:linear-gradient(135deg,#EEF1F7,#E0E5EF);'
                                 f'display:flex;align-items:center;justify-content:center;'
                                 f'font-size:3rem">👕</div>')

                    badge_col_a = "#DC2626" if total_a == 0 else "#D97706"
                    badge_txt_a = "Rupture" if total_a == 0 else "Stock faible"

                    with ac:
                        st.markdown(
                            f'<div style="background:#FFFFFF;border:2px solid #E0E5EF;'
                            f'border-radius:14px;overflow:hidden;margin-bottom:.5rem;'
                            f'box-shadow:0 2px 8px rgba(27,43,75,.06)">'
                            f'<div style="position:relative">{img_a}'
                            f'<span style="position:absolute;top:7px;right:7px;'
                            f'background:{badge_col_a};color:#FFF;font-size:.55rem;'
                            f'font-weight:700;padding:2px 7px;border-radius:20px;'
                            f'text-transform:uppercase">{badge_txt_a}</span></div>'
                            f'<div style="padding:.6rem .75rem">'
                            f'<div style="font-weight:800;color:#1B2B4B;font-size:.86rem;'
                            f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis">'
                            f'{name_a}</div>'
                            f'<div style="font-size:.68rem;color:#8A9AB5;margin:.2rem 0">'
                            f'{total_a} pcs en stock</div>'
                            f'<div style="font-size:.72rem;color:#DC2626;font-weight:700">'
                            f'+{cmd_a} pcs à commander</div>'
                            f'</div></div>', unsafe_allow_html=True)
                        if st.button("📊 Voir la situation", key=f"cmd_art_{mk_a}",
                                     use_container_width=True):
                            clicked_cmd_art = mk_a

            if clicked_cmd_art:
                st.session_state.cmd_art = clicked_cmd_art
                st.rerun()

        # ════════════════════════════════════════════════════════
        # ÉTAPE C — Situation complète de l'article
        # ════════════════════════════════════════════════════════
        else:
            mk_det  = st.session_state.cmd_art
            sd_det  = stock_data.get(mk_det, {})
            entry_det = mk_idx_cmd.get(mk_det, {})
            colors_det = sd_det.get("colors", [])
            sizes_det  = sd_det.get("sizes",  [])
            stk_det    = sd_det.get("stock",  {})
            seuil_det  = sd_det.get("seuil_min", 0)
            name_det   = sd_det.get("model_name", mk_det)

            # Boutons navigation
            nav1, nav2 = st.columns([1, 1])
            with nav1:
                if st.button("← Retour articles", key="cmd_back_art"):
                    st.session_state.cmd_art = None
                    st.rerun()
            with nav2:
                if st.button("⌂ Retour catégories", key="cmd_back_cat2"):
                    st.session_state.cmd_cat = None
                    st.session_state.cmd_art = None
                    st.rerun()

            # Carte article
            b64_det2 = entry_det.get("b64_thumb") or sd_det.get("b64_thumb")
            if b64_det2:
                ph_det = (f'<img src="data:image/png;base64,{b64_det2}" '
                          f'style="width:72px;height:88px;object-fit:cover;'
                          f'border-radius:8px;border:2px solid #E0E5EF;flex-shrink:0">')
            else:
                ph_det = (f'<div style="width:72px;height:88px;background:#EEF1F7;'
                          f'border-radius:8px;border:2px solid #E0E5EF;display:flex;'
                          f'align-items:center;justify-content:center;'
                          f'font-size:2rem;flex-shrink:0">👕</div>')

            total_det = sum(v for cd in stk_det.values() for v in cd.values())
            cmd_det   = sum(max(0, seuil_det - stk_det.get(c,{}).get(s,0))
                            for c in colors_det for s in sizes_det)
            cat_det   = entry_det.get("categorie","") or "Non classé"

            color_dots = "".join(
                f'<span style="width:13px;height:13px;border-radius:50%;'
                f'background:{HEX_MAP.get(c,"#888")};border:2px solid #FFF;'
                f'box-shadow:0 1px 3px rgba(0,0,0,.2);display:inline-block;'
                f'margin-right:3px"></span>'
                for c in colors_det)

            st.markdown(
                f'<div style="display:flex;gap:.9rem;align-items:center;'
                f'background:#FFFFFF;border:2px solid #E0E5EF;border-radius:14px;'
                f'padding:.8rem 1rem;margin-bottom:1rem;'
                f'box-shadow:0 3px 12px rgba(27,43,75,.07)">'
                f'{ph_det}'
                f'<div style="flex:1;min-width:0">'
                f'<div style="font-size:.6rem;color:var(--gold);font-weight:700;'
                f'text-transform:uppercase;letter-spacing:.1em">📂 {cat_det}</div>'
                f'<div style="font-size:1.1rem;font-weight:800;color:#1B2B4B;'
                f'margin:.15rem 0">{name_det}</div>'
                f'<div style="margin-bottom:.3rem">{color_dots}</div>'
                f'<div style="font-size:.7rem;color:#8A9AB5">'
                f'{total_det} pcs en stock · Seuil : {seuil_det}</div>'
                f'</div>'
                f'<div style="text-align:center;flex-shrink:0">'
                f'<div style="font-size:1.4rem;font-weight:900;color:#DC2626">+{cmd_det}</div>'
                f'<div style="font-size:.6rem;color:#8A9AB5;text-transform:uppercase;'
                f'letter-spacing:.08em">à commander</div>'
                f'</div></div>', unsafe_allow_html=True)

            # Légende
            st.markdown(
                '<div style="display:flex;gap:1rem;font-size:.62rem;color:#8A9AB5;'
                'margin-bottom:.5rem;flex-wrap:wrap">'
                '<span style="display:flex;align-items:center;gap:4px">'
                '<span style="width:12px;height:12px;border-radius:2px;'
                'background:#F0FFF4;border:1px solid #059669"></span>OK</span>'
                '<span style="display:flex;align-items:center;gap:4px">'
                '<span style="width:12px;height:12px;border-radius:2px;'
                'background:#FFF8E1;border:1px solid #D97706"></span>Faible</span>'
                '<span style="display:flex;align-items:center;gap:4px">'
                '<span style="width:12px;height:12px;border-radius:2px;'
                'background:#FFE4E4;border:1px solid #DC2626"></span>Rupture</span>'
                '<span style="margin-left:auto;font-style:italic">'
                'Stock · <span style="color:#DC2626;font-weight:700">+N</span> = à commander'
                '</span></div>', unsafe_allow_html=True)

            # Grille tailles × couleurs
            th_det = "".join(
                f'<th style="padding:8px 6px;text-align:center;font-size:.66rem;'
                f'color:#8A9AB5;text-transform:uppercase;letter-spacing:.08em;'
                f'border-bottom:2px solid #EEF1F7;border-right:1px solid #EEF1F7;'
                f'min-width:48px">{s}</th>'
                for s in sizes_det)

            rows_det = ""
            for ci_d, c_d in enumerate(colors_det):
                hx_d   = HEX_MAP.get(c_d, "#888")
                row_bg = "#FFFFFF" if ci_d % 2 == 0 else "#F8F9FC"
                cells_d = ""
                for s_d in sizes_det:
                    q_d = stk_det.get(c_d, {}).get(s_d, 0)
                    m_d = max(0, seuil_det - q_d)
                    if q_d == 0:
                        cbg = "#FFE4E4"; cq = "#DC2626"
                        plus = (f'<div style="font-size:.62rem;color:#DC2626;'
                                f'font-weight:800;line-height:1.1">+{m_d}</div>'
                                if m_d > 0 else "")
                    elif m_d > 0:
                        cbg = "#FFF8E1"; cq = "#D97706"
                        plus = (f'<div style="font-size:.62rem;color:#D97706;'
                                f'font-weight:800;line-height:1.1">+{m_d}</div>')
                    else:
                        cbg = "#F0FFF4"; cq = "#059669"; plus = ""
                    cells_d += (
                        f'<td style="padding:7px 4px;text-align:center;'
                        f'background:{cbg};border-bottom:1px solid #EEF1F7;'
                        f'border-right:1px solid #EEF1F7">'
                        f'<div style="font-size:.88rem;font-weight:800;'
                        f'color:{cq};line-height:1.2">{q_d}</div>'
                        f'{plus}</td>')
                rows_det += (
                    f'<tr>'
                    f'<td style="padding:7px 10px;border-bottom:1px solid #EEF1F7;'
                    f'border-right:2px solid #EEF1F7;background:{row_bg};white-space:nowrap">'
                    f'<span style="display:inline-flex;align-items:center;gap:7px">'
                    f'<span style="width:13px;height:13px;border-radius:50%;'
                    f'background:{hx_d};flex-shrink:0;'
                    f'border:1.5px solid rgba(0,0,0,.12)"></span>'
                    f'<span style="font-size:.8rem;font-weight:600;color:#1B2B4B">'
                    f'{c_d}</span></span></td>'
                    f'{cells_d}</tr>')

            st.markdown(
                f'<div style="background:#FFFFFF;border:2px solid #E0E5EF;'
                f'border-radius:14px;overflow:hidden;'
                f'box-shadow:0 3px 12px rgba(27,43,75,.07)">'
                f'<div style="overflow-x:auto">'
                f'<table style="width:100%;border-collapse:collapse">'
                f'<thead><tr style="background:#F4F6FA">'
                f'<th style="padding:8px 10px;text-align:left;font-size:.62rem;'
                f'color:#8A9AB5;text-transform:uppercase;letter-spacing:.1em;'
                f'border-bottom:2px solid #EEF1F7;border-right:2px solid #EEF1F7">'
                f'Couleur</th>{th_det}</tr></thead>'
                f'<tbody>{rows_det}</tbody>'
                f'</table></div></div>',
                unsafe_allow_html=True)

        # ── Exports globaux ──────────────────────────────────────────
        if orders:
            st.markdown("<div style='margin-top:.8rem'></div>",
                        unsafe_allow_html=True)
            df_o = pd.DataFrame(orders)
            e1, e2, e3 = st.columns(3)
            with e1:
                st.download_button(
                    "📥 Excel",
                    build_order_excel(orders),
                    f"commande_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            with e2:
                st.download_button(
                    "📄 CSV",
                    df_o.to_csv(index=False).encode("utf-8-sig"),
                    f"commande_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv",
                    use_container_width=True)
            with e3:
                lines = [f"🛒 *COMMANDE — {datetime.now().strftime('%d/%m/%Y')}*\n"]
                cm = ""
                for o in orders:
                    if o["Modèle"] != cm:
                        cm = o["Modèle"]
                        lines.append(f"\n*{cm}*")
                    lines.append(
                        f"  {'🔴' if o['En stock']==0 else '🟡'} "
                        f"{o['Couleur']}/{o['Taille']} → *{o['À commander']}* pcs")
                lines.append(f"\n_Total: {all_orders_tp} pièces_")
                wa_button("https://wa.me/?text=" + urllib.parse.quote("\n".join(lines)))

        st.markdown("<hr>",unsafe_allow_html=True)
        st.markdown("### 📬 Arrivage")
        st.caption("Quantités reçues → ajoutées au stock automatiquement.")
        if not stock_data:
            st.warning("Aucun stock configuré.")
        else:
            arr_m=st.selectbox("Modèle reçu",options=list(stock_data.keys()),format_func=lambda k:stock_data[k].get("model_name",k),key="arr_sel")
            if arr_m:
                md=stock_data[arr_m]; ac=md.get("colors",[]); az=md.get("sizes",[])
                df_a=pd.DataFrame([{"Couleur":c,**{s:0 for s in az}} for c in ac])
                ea=st.data_editor(df_a,use_container_width=True,hide_index=True,
                    column_config={"Couleur":st.column_config.TextColumn("Couleur",disabled=True,width="medium"),
                                   **{s:st.column_config.NumberColumn(s,min_value=0,max_value=9999,step=1,width="small") for s in az}},key=f"ae_{arr_m}")
                fournisseur_arr = st.text_input("Fournisseur (optionnel)",
                    placeholder="Ex: Fournisseur Paris, Ali Express…", key="fourn_arr")
                if st.button("✅ Valider l'arrivage",key="val_arr"):
                    total=0; cs=stock_data[arr_m].get("stock",{}); detail={}
                    for _,row in ea.iterrows():
                        col=row["Couleur"]; cs.setdefault(col,{}); detail[col]={}
                        for sz in az:
                            q=int(row[sz]); cs[col][sz]=cs[col].get(sz,0)+q
                            if q>0: detail[col][sz]=q; total+=q
                    stock_data[arr_m]["stock"]=cs
                    stock_data[arr_m]["last_update"]=datetime.now().strftime("%d/%m/%Y %H:%M")
                    save_stock(stock_data)
                    save_arrivage({"date":datetime.now().strftime("%d/%m/%Y %H:%M"),
                                   "model_key":arr_m,
                                   "model_name":md.get("model_name",arr_m),
                                   "fournisseur":fournisseur_arr or "—",
                                   "total":total, "detail":detail})
                    st.success(f"✓ {total} pièces ajoutées"); st.rerun()

        # ── Historique des arrivages ──────────────────────────────────────
        st.markdown("<hr>",unsafe_allow_html=True)
        st.markdown("### 📦 Historique des arrivages")
        arr_hist = load_arrivages()
        if not arr_hist:
            st.info("Aucun arrivage enregistré.")
        else:
            st.markdown(f"**{len(arr_hist)}** arrivage(s) enregistré(s)")
            # Filtre modèle
            arr_models = ["Tous"] + list(dict.fromkeys(a["model_name"] for a in arr_hist))
            arr_fil = st.selectbox("Filtrer par modèle",arr_models,key="arr_hist_fil")
            arr_show = arr_hist if arr_fil=="Tous" else [a for a in arr_hist if a["model_name"]==arr_fil]
            for a in arr_show:
                detail_html=""
                for col,sizes in a.get("detail",{}).items():
                    hx=HEX_MAP.get(col,"#888")
                    sizes_str=" · ".join(f"{sz}×{q}" for sz,q in sizes.items())
                    detail_html+=f'<span style="display:inline-flex;align-items:center;gap:4px;margin:2px 6px 2px 0"><span style="width:10px;height:10px;border-radius:50%;background:{hx};border:1px solid #ccc;display:inline-block"></span>{col}: {sizes_str}</span>'
                st.markdown(f"""
                <div class="history-card" style="border-left:3px solid #1565C0">
                    <div style="display:flex;justify-content:space-between">
                        <div class="history-date">📅 {a.get("date","")}</div>
                        <div style="font-size:.72rem;color:#FFF;background:#1565C0;padding:2px 10px">{a.get("total",0)} pièces</div>
                    </div>
                    <div class="history-title">{a.get("model_name","")}</div>
                    <div style="font-size:.75rem;color:#666;margin:3px 0">🏭 {a.get("fournisseur","—")}</div>
                    <div style="margin-top:6px;font-size:.75rem">{detail_html}</div>
                </div>""", unsafe_allow_html=True)
            # Export
            df_arr_h = pd.DataFrame([{"Date":a["date"],"Modèle":a["model_name"],
                "Fournisseur":a.get("fournisseur","—"),"Total pièces":a.get("total",0)} for a in arr_show])
            st.download_button("📥 Exporter arrivages (Excel)",
                _df_to_excel(df_arr_h,"Arrivages"),
                f"arrivages_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 4 : VENTES (tous les rôles)
# ══════════════════════════════════════════════════════════════════════════════

# ── Fichier file d'attente hors-ligne ─────────────────────────────────────────
OFFLINE_FILE = "ventes_offline.json"

def load_offline():  return _load(OFFLINE_FILE, [])
def save_offline(d): _save(OFFLINE_FILE, d)

def check_online() -> bool:
    """Vérifie la connexion internet (ping Google)."""
    import socket
    try:
        socket.setdefaulttimeout(2)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect(("8.8.8.8", 53))
        return True
    except Exception:
        return False

def sync_offline_queue():
    """Fusionne les ventes hors-ligne dans ventes.json si connecté."""
    queue = load_offline()
    if not queue:
        return 0
    for v in queue:
        save_vente_and_update_stock(v)
    save_offline([])
    return len(queue)

with tab_ventes:
    stock_data_v  = load_stock()
    vendeur_actuel = st.session_state.user_name

    # ════════════════════════════════════════════════════════════════════════
    # VUE ADMIN : Calendrier des ventes
    # ════════════════════════════════════════════════════════════════════════
    if is_admin:
        import calendar as _cal
        ventes_cal = load_ventes()
        now_cal    = datetime.now()

        if st.session_state.v_admin_month is None:
            st.session_state.v_admin_month = (now_cal.year, now_cal.month)
        yr_v, mo_v = st.session_state.v_admin_month

        # ── Navigation mois ─────────────────────────────────────────────────
        nav1, nav2, nav3 = st.columns([1, 3, 1])
        with nav1:
            if st.button("◀ Précédent", key="v_cal_prev", use_container_width=True):
                mo2 = mo_v - 1; yr2 = yr_v if mo2 > 0 else yr_v - 1; mo2 = mo2 if mo2 > 0 else 12
                st.session_state.v_admin_month = (yr2, mo2); st.session_state.v_admin_date = None; st.rerun()
        with nav2:
            mois_fr = ["","Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
            st.markdown(f'<div style="text-align:center;font-size:1.2rem;font-weight:700;color:#1B2B4B;padding:.4rem">{mois_fr[mo_v]} {yr_v}</div>', unsafe_allow_html=True)
        with nav3:
            if st.button("Suivant ▶", key="v_cal_next", use_container_width=True):
                mo2 = mo_v + 1; yr2 = yr_v if mo2 <= 12 else yr_v + 1; mo2 = mo2 if mo2 <= 12 else 1
                st.session_state.v_admin_month = (yr2, mo2); st.session_state.v_admin_date = None; st.rerun()

        # ── Index ventes par jour ────────────────────────────────────────────
        ventes_par_jour = {}
        for v in ventes_cal:
            ds = v.get("date","")[:10]
            try:
                dobj = datetime.strptime(ds, "%d/%m/%Y")
                if dobj.year == yr_v and dobj.month == mo_v:
                    ventes_par_jour.setdefault(ds, []).append(v)
            except: pass

        # ── En-têtes jours semaine ───────────────────────────────────────────
        jours_fr = ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]
        cols_hdr = st.columns(7)
        for ji, jn in enumerate(jours_fr):
            cols_hdr[ji].markdown(f'<div style="text-align:center;font-size:.72rem;font-weight:700;color:#8A9AB5;padding:.2rem 0">{jn}</div>', unsafe_allow_html=True)

        # ── Grille jours ─────────────────────────────────────────────────────
        cal_matrix   = _cal.monthcalendar(yr_v, mo_v)
        selected_d   = st.session_state.v_admin_date
        clicked_day_v = None

        for week in cal_matrix:
            week_cols = st.columns(7)
            for wi, day in enumerate(week):
                with week_cols[wi]:
                    if day == 0:
                        st.markdown('<div style="height:60px"></div>', unsafe_allow_html=True)
                        continue
                    day_str   = f"{day:02d}/{mo_v:02d}/{yr_v}"
                    nb_v_day  = len(ventes_par_jour.get(day_str, []))
                    is_today  = (day == now_cal.day and mo_v == now_cal.month and yr_v == now_cal.year)
                    is_sel    = (day_str == selected_d)
                    # Couleur du bouton selon état
                    if is_sel:
                        style = "background:#1B2B4B;color:#FFF;border:2px solid #1B2B4B;"
                    elif is_today:
                        style = "background:#EEF4FF;color:#1B2B4B;border:2px solid #2563EB;font-weight:700;"
                    elif nb_v_day > 0:
                        style = "background:#F0FDF4;color:#065F46;border:2px solid #059669;"
                    else:
                        style = "background:#F8F9FC;color:#8A9AB5;border:1px solid #E0E5EF;"

                    label = f"{day}" + (f"\n● {nb_v_day}" if nb_v_day > 0 else "")
                    st.markdown(
                        f'<div style="{style}border-radius:10px;padding:.4rem .2rem;'
                        f'text-align:center;font-size:.85rem;font-weight:600;'
                        f'min-height:58px;display:flex;flex-direction:column;'
                        f'align-items:center;justify-content:center;gap:2px">'
                        f'<span>{day}</span>'
                        f'{"<span style=\"font-size:.6rem;font-weight:700\">● "+str(nb_v_day)+" vente(s)</span>" if nb_v_day>0 else ""}'
                        f'</div>', unsafe_allow_html=True)
                    if st.button(f"{day}", key=f"v_day_{yr_v}_{mo_v}_{day}",
                                 use_container_width=True,
                                 help=f"{'Cliquez pour voir les ' + str(nb_v_day) + ' vente(s)' if nb_v_day>0 else 'Aucune vente'}"):
                        clicked_day_v = day_str

        if clicked_day_v:
            st.session_state.v_admin_date = clicked_day_v
            st.rerun()

        # ── Statistiques du jour sélectionné ────────────────────────────────
        if selected_d:
            ventes_jour = ventes_par_jour.get(selected_d, [])
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown(
                f'<div style="font-size:1.1rem;font-weight:700;color:#1B2B4B;margin-bottom:1rem">'
                f'📊 Statistiques — {selected_d}</div>', unsafe_allow_html=True)

            if not ventes_jour:
                st.info("Aucune vente enregistrée ce jour.")
            else:
                pieces_j  = sum(v.get("quantite",1) for v in ventes_jour)
                sm1, sm2, sm3 = st.columns(3)
                sm1.metric("💰 Transactions", len(ventes_jour))
                sm2.metric("👕 Pièces vendues", pieces_j)
                sm3.metric("🧑‍💼 Vendeurs actifs", len(set(v.get("vendeur","") for v in ventes_jour)))

                st.markdown("**Par vendeur :**")
                agg_vj = {}
                for v in ventes_jour:
                    agg_vj[v.get("vendeur","?")] = agg_vj.get(v.get("vendeur","?"),0) + v.get("quantite",1)
                for vnd, qty in sorted(agg_vj.items(), key=lambda x:-x[1]):
                    ini_vnd = "".join(p[0].upper() for p in vnd.split()[:2])
                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:.8rem;background:#F8F9FC;'
                        f'border-radius:8px;padding:.5rem .8rem;margin-bottom:.3rem">'
                        f'<div class="profil-avatar" style="width:30px;height:30px;font-size:.7rem;min-width:30px">{ini_vnd}</div>'
                        f'<strong style="color:#1B2B4B">{vnd}</strong>'
                        f'<span style="margin-left:auto;color:#059669;font-weight:700">{qty} pcs</span></div>',
                        unsafe_allow_html=True)

                st.markdown("**Détail des ventes :**")
                for v in ventes_jour:
                    hx_v2 = HEX_MAP.get(v.get("couleur",""),"#888")
                    st.markdown(
                        f'<div style="background:#FFF;border:1px solid #E0E5EF;border-radius:8px;'
                        f'padding:.6rem 1rem;margin-bottom:.3rem;display:flex;align-items:center;justify-content:space-between">'
                        f'<div><strong style="color:#1B2B4B">{v.get("model_name","?")}</strong> '
                        f'<span style="font-size:.78rem;display:inline-flex;align-items:center;gap:4px">'
                        f'<span style="width:10px;height:10px;border-radius:50%;background:{hx_v2};display:inline-block"></span>'
                        f'{v.get("couleur","?")} · {v.get("taille","?")}</span></div>'
                        f'<div style="text-align:right"><span style="color:#2563EB;font-weight:700">×{v.get("quantite",1)}</span>'
                        f'<div style="font-size:.65rem;color:#8A9AB5">{v.get("vendeur","?")}</div></div></div>',
                        unsafe_allow_html=True)

    # ── Init session state du flux vente ─────────────────────────────────────
    for k, v in [("v_step", 0), ("v_cat", None), ("v_art_key", None),
                 ("v_color", None), ("v_size", None)]:
        if k not in st.session_state:
            st.session_state[k] = v

    # ── Bandeau vendeur + statut connexion ───────────────────────────────────
    is_online = check_online()
    conn_badge = (
        '<span style="background:#D1FAE5;border:1px solid #059669;color:#065F46;'
        'font-size:.68rem;font-weight:600;padding:3px 10px;border-radius:100px">🟢 En ligne</span>'
        if is_online else
        '<span style="background:#FEF3C7;border:1px solid #D97706;color:#92400E;'
        'font-size:.68rem;font-weight:600;padding:3px 10px;border-radius:100px">🟡 Hors ligne</span>'
    )

    # ── Sync automatique des ventes offline ──────────────────────────────────
    if is_online:
        nb_sync = sync_offline_queue()
        if nb_sync:
            st.success(f"🔄 {nb_sync} vente(s) hors-ligne synchronisée(s) automatiquement !")

    st.markdown(
        f'<div style="display:flex;align-items:center;justify-content:space-between;'
        f'background:#F8F9FC;border:1px solid #E0E5EF;border-radius:10px;'
        f'padding:.8rem 1.2rem;margin-bottom:1.2rem">'
        f'<div style="display:flex;align-items:center;gap:.6rem">'
        f'<div class="profil-avatar" style="width:36px;height:36px;font-size:.85rem;'
        f'min-width:36px">{"".join(p[0].upper() for p in vendeur_actuel.split()[:2])}</div>'
        f'<div><strong style="color:#1B2B4B">{vendeur_actuel}</strong>'
        f'<div style="font-size:.7rem;color:#8A9AB5">🕐 {datetime.now().strftime("%d/%m/%Y  %H:%M")}</div></div></div>'
        f'{conn_badge}</div>',
        unsafe_allow_html=True)

    if not stock_data_v:
        st.warning("⚠️ Aucun stock configuré. L'admin doit d'abord ajouter des articles.")
    else:
        cats_v    = load_categories()
        history_v = load_history()

        # ── BARRE DE PROGRESSION ─────────────────────────────────────────────
        steps = ["Catégorie", "Article", "Couleur", "Taille & Qté", "Confirmation"]
        cur   = st.session_state.v_step
        prog_html = '<div style="display:flex;gap:4px;margin-bottom:1.4rem">'
        for si, sl in enumerate(steps):
            done   = si < cur
            active = si == cur
            bg     = "#1B2B4B" if done else ("#2563EB" if active else "#E0E5EF")
            tc     = "#FFF"    if (done or active) else "#8A9AB5"
            fw     = "700"     if active else "500"
            prog_html += (
                f'<div style="flex:1;background:{bg};color:{tc};font-size:.65rem;'
                f'font-weight:{fw};text-align:center;padding:6px 2px;border-radius:6px;'
                f'letter-spacing:.05em">{"✓ " if done else ""}{sl}</div>'
            )
        prog_html += '</div>'
        st.markdown(prog_html, unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════
        # ÉTAPE 0 — Choisir la catégorie
        # ════════════════════════════════════════════════════════════════
        if cur == 0:
            st.markdown(
                '<div style="font-size:1.3rem;font-weight:800;color:#1B2B4B;margin-bottom:1.2rem">'
                '🗂️ Choisissez une catégorie</div>', unsafe_allow_html=True)
            if not cats_v:
                st.session_state.v_cat  = "Toutes"
                st.session_state.v_step = 1
                st.rerun()

            clicked_cat = None
            nb_c = min(len(cats_v), 3)
            c_cols = st.columns(nb_c)
            for i, cat in enumerate(cats_v):
                cn = cat.get("name","") if isinstance(cat,dict) else cat
                ci = cat.get("icon","📦") if isinstance(cat,dict) else "📦"
                cc = cat.get("color","#2563EB") if isinstance(cat,dict) else "#2563EB"
                nb_art = sum(1 for e in history_v
                             if (e.get("categorie","") or "") == cn
                             and model_key(e["model_name"]) in stock_data_v)
                with c_cols[i % nb_c]:
                    st.markdown(
                        f'<div style="background:#FFF;border:2px solid {cc}44;'
                        f'border-radius:18px;padding:1.8rem 1rem;text-align:center;'
                        f'margin-bottom:.6rem;box-shadow:0 3px 12px {cc}22">'
                        f'<div style="font-size:3rem;margin-bottom:.6rem">{ci}</div>'
                        f'<div style="font-weight:800;color:#1B2B4B;font-size:1.1rem;'
                        f'margin-bottom:.4rem">{cn}</div>'
                        f'<div style="background:{cc}22;color:{cc};font-size:.75rem;'
                        f'font-weight:700;padding:3px 12px;border-radius:100px;'
                        f'display:inline-block">{nb_art} article(s)</div>'
                        f'</div>', unsafe_allow_html=True)
                    if st.button(f"Entrer →", key=f"vcat_{i}", use_container_width=True):
                        clicked_cat = cn
            if clicked_cat:
                st.session_state.v_cat  = clicked_cat
                st.session_state.v_step = 1
                st.rerun()

        # ════════════════════════════════════════════════════════════════
        # ÉTAPE 1 — Choisir l'article (grille photos)
        # ════════════════════════════════════════════════════════════════
        elif cur == 1:
            cat_obj_v = next((c for c in cats_v if
                (c.get("name") if isinstance(c,dict) else c) == st.session_state.v_cat), None)
            cat_icon_v  = cat_obj_v.get("icon","📦") if isinstance(cat_obj_v,dict) else "📦"
            cat_color_v = cat_obj_v.get("color","#2563EB") if isinstance(cat_obj_v,dict) else "#2563EB"
            st.markdown(
                f'<div style="background:{cat_color_v}18;border:1.5px solid {cat_color_v}55;'
                f'border-radius:12px;padding:.7rem 1.2rem;display:inline-flex;align-items:center;'
                f'gap:.6rem;margin-bottom:1rem">'
                f'<span style="font-size:1.5rem">{cat_icon_v}</span>'
                f'<span style="font-weight:800;color:{cat_color_v};font-size:1.1rem">'
                f'{st.session_state.v_cat}</span></div>',
                unsafe_allow_html=True)
            st.markdown(
                '<div style="font-size:1.1rem;font-weight:800;color:#1B2B4B;margin-bottom:1rem">'
                '👕 Choisissez un article</div>', unsafe_allow_html=True)

            # Index catégorie depuis l'historique pour les anciens articles
            hist_cat_idx = {}
            for e in history_v:
                mk_e2 = model_key(e["model_name"])
                if mk_e2 not in hist_cat_idx:
                    hist_cat_idx[mk_e2] = e.get("categorie","") or ""

            arts_cat = []
            sel_cat = st.session_state.v_cat
            for mk_e, sd_e in stock_data_v.items():
                # Catégorie depuis le stock (nouveaux articles) ou depuis l'historique (anciens)
                ec = sd_e.get("categorie","") or hist_cat_idx.get(mk_e, "")
                if sel_cat != "Toutes" and ec.strip().lower() != sel_cat.strip().lower():
                    continue
                # Créer une entrée synthétique compatible avec le reste du code
                entry_synth = {
                    "model_name": sd_e.get("model_name", mk_e),
                    "categorie":  ec,
                    "b64_thumb":  sd_e.get("b64_thumb",""),
                }
                arts_cat.append((mk_e, entry_synth))

            # Tri par popularité
            ventes_count_v = {}
            for vv in load_ventes():
                ventes_count_v[vv.get("model_key","")] = ventes_count_v.get(vv.get("model_key",""), 0) + vv.get("quantite",1)
            arts_cat.sort(key=lambda x: ventes_count_v.get(x[0], 0), reverse=True)

            clicked_art = None

            if not arts_cat:
                st.info("Aucun article en stock pour cette catégorie.")
            else:
                for row_start in range(0, len(arts_cat), 3):
                    row_arts = arts_cat[row_start:row_start + 3]
                    c0, c1, c2 = st.columns(3)
                    row_cols = [c0, c1, c2]
                    for ci3, (mk_e, entry) in enumerate(row_arts):
                        sd_art = stock_data_v.get(mk_e, {})
                        total_stock = sum(
                            v for c_dict in sd_art.get("stock",{}).values()
                            for v in c_dict.values())
                        sbc  = "#059669" if total_stock > 5 else "#D97706" if total_stock > 0 else "#DC2626"
                        slbl = (f"✅ {total_stock} en stock" if total_stock > 5
                                else f"⚠️ {total_stock} restants" if total_stock > 0
                                else "❌ Épuisé")
                        b64  = sd_art.get("b64_thumb","") or entry.get("b64_thumb","")
                        photo_html = (
                            f'<img src="data:image/jpeg;base64,{b64}" '
                            f'style="width:100%;height:140px;object-fit:cover;'
                            f'border-radius:10px;margin-bottom:.6rem;display:block">'
                            if b64 else
                            f'<div style="width:100%;height:140px;background:#EEF1F7;'
                            f'border-radius:10px;display:flex;align-items:center;'
                            f'justify-content:center;font-size:3rem;margin-bottom:.6rem">👕</div>'
                        )
                        with row_cols[ci3]:
                            st.markdown(
                                f'<div style="background:#FFF;border:2px solid #E0E5EF;'
                                f'border-radius:14px;padding:.8rem;margin-bottom:.8rem;'
                                f'box-shadow:0 2px 8px #0001">'
                                f'{photo_html}'
                                f'<div style="font-weight:800;color:#1B2B4B;font-size:.95rem;'
                                f'margin-bottom:.3rem">{entry["model_name"]}</div>'
                                f'<div style="font-size:.75rem;color:{sbc};font-weight:700">{slbl}</div>'
                                f'</div>', unsafe_allow_html=True)
                            if total_stock > 0:
                                if st.button("Choisir", key=f"vart_{mk_e}", use_container_width=True):
                                    clicked_art = mk_e
                            else:
                                st.button("Épuisé", key=f"vart_{mk_e}", disabled=True, use_container_width=True)

            if clicked_art:
                st.session_state.v_art_key = clicked_art
                st.session_state.v_step    = 2
                st.rerun()

            if st.button("← Retour catégories", key="vback_cat"):
                st.session_state.v_step = 0
                st.rerun()

        # ════════════════════════════════════════════════════════════════
        # ÉTAPE 2 — Choisir la couleur
        # ════════════════════════════════════════════════════════════════
        elif cur == 2:
            mk_v     = st.session_state.v_art_key
            sd_art   = stock_data_v.get(mk_v, {})
            art_name = sd_art.get("model_name", mk_v)
            colors_v = sd_art.get("colors", [])

            st.markdown(f'<div style="font-size:.82rem;color:#8A9AB5;margin-bottom:.8rem">'
                        f'📂 {st.session_state.v_cat} › <strong style="color:#1B2B4B">'
                        f'{art_name}</strong></div>', unsafe_allow_html=True)
            st.markdown('<span class="section-label">Choisissez la couleur</span>',
                        unsafe_allow_html=True)

            clicked_color = None   # ← collecte du clic HORS colonnes

            if not colors_v:
                st.warning("Aucune couleur définie pour cet article.")
            else:
                # Toujours 4 colonnes fixes
                fixed_cols = st.columns(4)
                for i, color in enumerate(colors_v):
                    hx = HEX_MAP.get(color, "#888888")
                    stock_color = sum(sd_art.get("stock",{}).get(color,{}).values()) \
                                  if sd_art.get("stock",{}).get(color) else 0
                    with fixed_cols[i % 4]:
                        st.markdown(
                            f'<div style="background:#FFF;border:2px solid #E0E5EF;'
                            f'border-radius:12px;padding:1rem .5rem;text-align:center;'
                            f'margin-bottom:.4rem">'
                            f'<div style="width:46px;height:46px;border-radius:50%;'
                            f'background:{hx};margin:0 auto .5rem;'
                            f'border:3px solid rgba(0,0,0,.12)"></div>'
                            f'<div style="font-weight:600;color:#1B2B4B;font-size:.78rem">'
                            f'{color}</div>'
                            f'<div style="font-size:.65rem;color:#8A9AB5;margin-top:.2rem">'
                            f'{stock_color} pcs</div></div>',
                            unsafe_allow_html=True)
                        if st.button(color, key=f"vcol_{i}",
                                     use_container_width=True):
                            clicked_color = color   # ← enregistre sans rerun

            # ── Rerun HORS des colonnes ──
            if clicked_color:
                st.session_state.v_color = clicked_color
                st.session_state.v_step  = 3
                st.rerun()

            if st.button("← Retour articles", key="vback_art"):
                st.session_state.v_step = 1
                st.rerun()

        # ════════════════════════════════════════════════════════════════
        # ÉTAPE 3 — Choisir la taille + quantité
        # ════════════════════════════════════════════════════════════════
        elif cur == 3:
            mk_v     = st.session_state.v_art_key
            color_v  = st.session_state.v_color
            sd_art   = stock_data_v.get(mk_v, {})
            art_name = sd_art.get("model_name", mk_v)
            sizes_v  = sd_art.get("sizes", [])
            hx_v     = HEX_MAP.get(color_v, "#888")

            st.markdown(
                f'<div style="font-size:.82rem;color:#8A9AB5;margin-bottom:.8rem">'
                f'📂 {st.session_state.v_cat} › {art_name} › '
                f'<span style="display:inline-flex;align-items:center;gap:4px">'
                f'<span style="width:12px;height:12px;border-radius:50%;background:{hx_v};'
                f'display:inline-block;border:1px solid rgba(0,0,0,.15)"></span>'
                f'<strong style="color:#1B2B4B">{color_v}</strong></span></div>',
                unsafe_allow_html=True)
            st.markdown('<span class="section-label">Choisissez la taille</span>',
                        unsafe_allow_html=True)

            clicked_size = None   # ← collecte du clic HORS colonnes

            if sizes_v:
                # Toujours 5 colonnes fixes pour éviter le bug React DOM
                sz_fixed = st.columns(5)
                for i, sz in enumerate(sizes_v):
                    dispo = sd_art.get("stock",{}).get(color_v,{}).get(sz, 0)
                    badge_col  = ("#059669" if dispo > 5
                                  else "#D97706" if dispo > 0 else "#DC2626")
                    sel_border = ("2px solid #2563EB"
                                  if st.session_state.v_size == sz
                                  else "2px solid #E0E5EF")
                    with sz_fixed[i % 5]:
                        st.markdown(
                            f'<div style="background:#FFF;border:{sel_border};'
                            f'border-radius:10px;padding:.9rem .4rem;text-align:center;'
                            f'margin-bottom:.4rem">'
                            f'<div style="font-size:1.3rem;font-weight:800;color:#1B2B4B">'
                            f'{sz}</div>'
                            f'<div style="font-size:.65rem;color:{badge_col};font-weight:600;'
                            f'margin-top:.3rem">{dispo} dispo</div></div>',
                            unsafe_allow_html=True)
                        disabled_sz = (dispo == 0)
                        btn_lbl = "✕ Épuisé" if disabled_sz else "Choisir"
                        if st.button(btn_lbl, key=f"vsz_{i}",
                                     use_container_width=True,
                                     disabled=disabled_sz):
                            clicked_size = sz   # ← enregistre sans rerun

            # ── Rerun HORS des colonnes ──
            if clicked_size:
                st.session_state.v_size = clicked_size
                st.rerun()

            if st.session_state.v_size:
                sz_sel   = st.session_state.v_size
                dispo_sel = sd_art.get("stock",{}).get(color_v,{}).get(sz_sel, 0)
                st.markdown("<hr>", unsafe_allow_html=True)
                qa, qb, qc = st.columns([2, 2, 2])
                with qa:
                    vqty = st.number_input(
                        f"Quantité (max {dispo_sel})",
                        min_value=1, max_value=max(1, dispo_sel),
                        value=1, step=1, key="v_qty_input")
                with qb:
                    vnote = st.text_input("Note (optionnel)",
                                          placeholder="Ex: Solde, échange…",
                                          key="v_note_input")
                with qc:
                    st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
                    if st.button("✅ Enregistrer la vente", key="v_go_confirm",
                                 use_container_width=True):
                        qty_v  = int(vqty)
                        note_v = vnote
                        dispo_v = sd_art.get("stock",{}).get(color_v,{}).get(sz_sel, 0)
                        now_v = datetime.now()
                        vente_rec = {
                            "date":        now_v.strftime("%d/%m/%Y %H:%M"),
                            "vendeur":     vendeur_actuel,
                            "model_key":   mk_v,
                            "model_name":  art_name,
                            "couleur":     color_v,
                            "taille":      sz_sel,
                            "quantite":    qty_v,
                            "note":        note_v,
                            "stock_avant": dispo_v,
                            "stock_apres": dispo_v - qty_v,
                        }
                        if is_online:
                            save_vente_and_update_stock(vente_rec)
                            sync_msg = "✅  Vente enregistrée !"
                        else:
                            q = load_offline()
                            q.append(vente_rec)
                            save_offline(q)
                            sd_loc = load_stock()
                            sd_loc.setdefault(mk_v, {}).setdefault("stock", {}).setdefault(color_v, {})
                            sd_loc[mk_v]["stock"][color_v][sz_sel] = max(
                                0, sd_loc[mk_v]["stock"][color_v].get(sz_sel, 0) - qty_v)
                            save_stock(sd_loc)
                            sync_msg = "🟡  Vente sauvegardée localement"
                        for k in ["v_step","v_cat","v_art_key","v_color","v_size","v_qty","v_note"]:
                            st.session_state[k] = 0 if k == "v_step" else None
                        st.success(sync_msg)
                        sd2 = load_stock()
                        ns  = sd2.get(mk_v,{}).get("stock",{}).get(color_v,{}).get(sz_sel, 0)
                        seuil_v = sd2.get(mk_v,{}).get("seuil_min", 0)
                        if ns < seuil_v:
                            st.warning(f"🔔 Stock {color_v}/{sz_sel} sous le seuil ({ns} < {seuil_v}). À commander !")
                        st.rerun()

            if st.button("← Retour couleurs", key="vback_col"):
                st.session_state.v_size = None
                st.session_state.v_step = 2
                st.rerun()

        # ════════════════════════════════════════════════════════════════
        # ÉTAPE 4 — Confirmation & enregistrement
        # ════════════════════════════════════════════════════════════════
        elif cur == 4:
            mk_v     = st.session_state.v_art_key
            color_v  = st.session_state.v_color
            sz_v     = st.session_state.v_size
            qty_v    = st.session_state.get("v_qty", 1)
            note_v   = st.session_state.get("v_note", "")
            sd_art   = stock_data_v.get(mk_v, {})
            art_name = sd_art.get("model_name", mk_v)
            hx_v     = HEX_MAP.get(color_v, "#888")
            dispo    = sd_art.get("stock",{}).get(color_v,{}).get(sz_v, 0)

            st.markdown(
                f'<div style="background:#FFF;border:2px solid #2563EB;border-radius:14px;'
                f'padding:1.6rem;max-width:480px;margin:0 auto 1.2rem">'
                f'<div style="font-size:.7rem;letter-spacing:.15em;text-transform:uppercase;'
                f'color:#8A9AB5;margin-bottom:.8rem;font-weight:600">Récapitulatif de la vente</div>'
                f'<div style="display:flex;flex-direction:column;gap:.6rem">'
                f'<div style="display:flex;justify-content:space-between">'
                f'<span style="color:#4A5568;font-size:.85rem">Vendeur</span>'
                f'<strong style="color:#1B2B4B">{vendeur_actuel}</strong></div>'
                f'<div style="display:flex;justify-content:space-between">'
                f'<span style="color:#4A5568;font-size:.85rem">Article</span>'
                f'<strong style="color:#1B2B4B">{art_name}</strong></div>'
                f'<div style="display:flex;justify-content:space-between;align-items:center">'
                f'<span style="color:#4A5568;font-size:.85rem">Couleur</span>'
                f'<span style="display:inline-flex;align-items:center;gap:6px">'
                f'<span style="width:14px;height:14px;border-radius:50%;background:{hx_v};'
                f'border:1px solid rgba(0,0,0,.15);display:inline-block"></span>'
                f'<strong style="color:#1B2B4B">{color_v}</strong></span></div>'
                f'<div style="display:flex;justify-content:space-between">'
                f'<span style="color:#4A5568;font-size:.85rem">Taille</span>'
                f'<strong style="color:#1B2B4B">{sz_v}</strong></div>'
                f'<div style="display:flex;justify-content:space-between">'
                f'<span style="color:#4A5568;font-size:.85rem">Quantité</span>'
                f'<strong style="color:#2563EB;font-size:1.1rem">{qty_v}</strong></div>'
                f'<div style="display:flex;justify-content:space-between">'
                f'<span style="color:#4A5568;font-size:.85rem">Stock après</span>'
                f'<strong style="color:{"#DC2626" if dispo-qty_v<=0 else "#059669"}">'
                f'{dispo - qty_v} pcs</strong></div>'
                + (f'<div style="display:flex;justify-content:space-between">'
                   f'<span style="color:#4A5568;font-size:.85rem">Note</span>'
                   f'<span style="color:#1B2B4B;font-size:.85rem">{note_v}</span></div>'
                   if note_v else "")
                + f'</div></div>',
                unsafe_allow_html=True)

            conf1, conf2 = st.columns(2)
            with conf1:
                if st.button("✅  Confirmer la vente", key="v_confirm_final",
                             use_container_width=True):
                    now_v = datetime.now()
                    vente_rec = {
                        "date":        now_v.strftime("%d/%m/%Y %H:%M"),
                        "vendeur":     vendeur_actuel,
                        "model_key":   mk_v,
                        "model_name":  art_name,
                        "couleur":     color_v,
                        "taille":      sz_v,
                        "quantite":    qty_v,
                        "note":        note_v,
                        "stock_avant": dispo,
                        "stock_apres": dispo - qty_v,
                    }
                    if is_online:
                        save_vente_and_update_stock(vente_rec)
                        sync_msg = "✅  Vente enregistrée et synchronisée en ligne !"
                    else:
                        # Sauvegarde hors-ligne dans la file d'attente
                        q = load_offline()
                        q.append(vente_rec)
                        save_offline(q)
                        # Mise à jour du stock locale quand même
                        sd_loc = load_stock()
                        sd_loc.setdefault(mk_v, {}).setdefault("stock", {}).setdefault(color_v, {})
                        sd_loc[mk_v]["stock"][color_v][sz_v] = max(
                            0, sd_loc[mk_v]["stock"][color_v].get(sz_v, 0) - qty_v)
                        save_stock(sd_loc)
                        sync_msg = "🟡  Vente sauvegardée localement (sera synchronisée à la reconnexion)"

                    # Reset du flux
                    for k in ["v_step","v_cat","v_art_key","v_color",
                               "v_size","v_qty","v_note"]:
                        st.session_state[k] = 0 if k == "v_step" else None
                    st.success(sync_msg)
                    sd2    = load_stock()
                    ns     = sd2.get(mk_v,{}).get("stock",{}).get(color_v,{}).get(sz_v, 0)
                    seuil_v = sd2.get(mk_v,{}).get("seuil_min", 0)
                    if ns < seuil_v:
                        st.warning(f"🔔 Stock {color_v}/{sz_v} sous le seuil ({ns} < {seuil_v}). À commander !")
                    st.rerun()

            with conf2:
                if st.button("✏️  Modifier", key="v_edit_back",
                             use_container_width=True):
                    st.session_state.v_step = 3
                    st.rerun()

    # File d'attente offline visible
    pending = load_offline()
    if pending:
        st.markdown(
            f'<div style="background:#FEF3C7;border:1.5px solid #D97706;border-radius:10px;'
            f'padding:.8rem 1.2rem;margin-top:1rem;display:flex;align-items:center;gap:.8rem">'
            f'<span style="font-size:1.2rem">🟡</span>'
            f'<span style="color:#92400E;font-size:.82rem"><strong>{len(pending)} vente(s)</strong>'
            f' en attente de synchronisation — elles seront envoyées automatiquement '
            f'dès que la connexion sera rétablie.</span></div>',
            unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # CLASSEMENT + RAPPORT — uniquement à l'étape 0 ou pour l'admin
    # ══════════════════════════════════════════════════════════════════════════
    if st.session_state.get("v_step", 0) == 0 or is_admin:
        st.markdown("<hr>", unsafe_allow_html=True)
        ventes_rank = load_ventes()
        st.markdown("### 🏆 Classement des articles les plus vendus")

        # Filtre période
        rk1,rk2=st.columns([2,2])
        with rk1:
            periode=st.selectbox("Période",["Tout","Aujourd'hui","Cette semaine","Ce mois"],key="rank_period")
        with rk2:
            top_n=st.selectbox("Afficher le top",[5,10,20,"Tout"],key="rank_n")

        # Filtrage par période
        now_r=datetime.now()
        def in_period(v):
            try:
                vd=datetime.strptime(v.get("date","").split(" ")[0],"%d/%m/%Y")
                if periode=="Aujourd'hui": return vd.date()==now_r.date()
                if periode=="Cette semaine": return (now_r.date()-vd.date()).days<7
                if periode=="Ce mois": return vd.month==now_r.month and vd.year==now_r.year
                return True
            except: return True

        vr_filt=[v for v in ventes_rank if in_period(v)]

        if not vr_filt:
            st.info(f"Aucune vente pour la période : {periode}")
        else:
            # Agrégation par (modèle, couleur, taille)
            from collections import defaultdict
            agg=defaultdict(int)
            for v in vr_filt:
                key_r=(v.get("model_name",""),v.get("couleur",""),v.get("taille",""))
                agg[key_r]+=v.get("quantite",1)

            ranked=sorted(agg.items(),key=lambda x:-x[1])
            if top_n!="Tout": ranked=ranked[:int(top_n)]

            max_qty=ranked[0][1] if ranked else 1

            # Affichage podium + barre
            medals=["🥇","🥈","🥉"]
            for rank_i,((mod,col,sz),qty) in enumerate(ranked,1):
                hx=HEX_MAP.get(col,"#888888")
                medal=medals[rank_i-1] if rank_i<=3 else f"#{rank_i}"
                pct=int(qty/max_qty*100)
                bg="#FFF9E6" if rank_i==1 else "#F8F8F8" if rank_i<=3 else "#FFFFFF"
                border_l="3px solid #CFB53B" if rank_i==1 else "3px solid #AAA" if rank_i==2 else "3px solid #CD7F32" if rank_i==3 else "1px solid #E8E8E8"
                st.markdown(f"""
                <div style="background:{bg};border-left:{border_l};padding:10px 14px;
                    margin-bottom:8px;border-radius:0 2px 2px 0">
                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px">
                        <div style="display:flex;align-items:center;gap:10px">
                            <span style="font-size:1.2rem;min-width:28px">{medal}</span>
                            <div>
                                <div style="font-weight:600;font-size:.88rem;color:#111">{mod}</div>
                                <div style="display:flex;align-items:center;gap:6px;margin-top:3px">
                                    <span style="width:11px;height:11px;border-radius:50%;background:{hx};
                                        border:1px solid #ccc;display:inline-block"></span>
                                    <span style="font-size:.75rem;color:#555">{col}</span>
                                    <span style="background:#111;color:#FFF;font-size:.65rem;
                                        padding:1px 7px;letter-spacing:.08em">{sz}</span>
                                </div>
                            </div>
                        </div>
                        <div style="text-align:right">
                            <div style="font-size:1.3rem;font-weight:700;color:#111">{qty}</div>
                            <div style="font-size:.62rem;color:#888;letter-spacing:.1em">PIÈCES</div>
                        </div>
                    </div>
                    <div style="background:#E8E8E8;height:4px;border-radius:2px">
                        <div style="background:{'#CFB53B' if rank_i==1 else '#111'};
                            width:{pct}%;height:4px;border-radius:2px;transition:width .3s"></div>
                    </div>
                </div>""", unsafe_allow_html=True)

            # Résumé par taille
            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown('<span class="section-label">Répartition par taille</span>',unsafe_allow_html=True)
            agg_sz=defaultdict(int)
            for v in vr_filt: agg_sz[v.get("taille","")]+= v.get("quantite",1)
            total_sz=sum(agg_sz.values())
            sz_sorted=sorted(agg_sz.items(),key=lambda x:-x[1])
            sz_cols=st.columns(len(sz_sorted)) if len(sz_sorted)<=6 else st.columns(6)
            for i,(sz,qty) in enumerate(sz_sorted):
                with sz_cols[i%len(sz_cols)]:
                    pct_sz=int(qty/total_sz*100) if total_sz else 0
                    st.markdown(f"""
                    <div style="text-align:center;background:#FFF;border:1px solid #E8E8E8;padding:10px 6px">
                        <div style="font-size:1rem;font-weight:700;color:#111">{qty}</div>
                        <div style="font-size:.9rem;font-weight:600;color:#333;margin:2px 0">{sz}</div>
                        <div style="font-size:.68rem;color:#AAA">{pct_sz}%</div>
                    </div>""", unsafe_allow_html=True)

            # Export classement
            st.markdown("<br>",unsafe_allow_html=True)
            df_rank=pd.DataFrame([{"Rang":i+1,"Modèle":m,"Couleur":c,"Taille":s,"Pièces vendues":q}
                                   for i,((m,c,s),q) in enumerate(ranked)])
            st.download_button("📥 Exporter le classement (Excel)",
                build_rank_excel(df_rank,periode),
                f"classement_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=False)

        # ── Rapport journalier ───────────────────────────────────────────────
        st.markdown("<hr>",unsafe_allow_html=True)
        st.markdown("### 📋 Rapport journalier")
        today_str = datetime.now().strftime("%d/%m/%Y")
        ventes_jour = [v for v in load_ventes() if v.get("date","").startswith(today_str)]
        pcs_jour = sum(v.get("quantite",1) for v in ventes_jour)

        rj1,rj2,rj3 = st.columns(3)
        rj1.metric("Transactions aujourd'hui", len(ventes_jour))
        rj2.metric("Pièces vendues aujourd'hui", pcs_jour)
        rj3.metric("Vendeurs actifs", len(set(v.get("vendeur","") for v in ventes_jour)))

        col_rj1, col_rj2 = st.columns(2)
        with col_rj1:
            rapport_txt = build_rapport_journalier()
            wa_rapport = "https://wa.me/?text=" + urllib.parse.quote(rapport_txt)
            wa_button(wa_rapport, label="📤  Envoyer le rapport sur WhatsApp")
        with col_rj2:
            st.download_button("📥  Télécharger le rapport (TXT)",
                rapport_txt.encode("utf-8"),
                f"rapport_{datetime.now().strftime('%Y%m%d')}.txt",
                "text/plain", use_container_width=True)

    # ══════════════════════════════════════════════════════════════════════════
    # ARTICLES À ROTATION LENTE
    # ══════════════════════════════════════════════════════════════════════════
    if is_admin:
        st.markdown("<hr>",unsafe_allow_html=True)
        st.markdown("### 🐢 Articles à rotation lente")
        st.caption("Articles en stock mais peu ou pas vendus récemment.")
        sl1, sl2 = st.columns([2,1])
        with sl1:
            seuil_rot = st.slider("Seuil : pas de vente depuis (jours)", 7, 60, 14, key="rot_seuil")
        rot_lente = get_rotation_lente(seuil_rot)
        if not rot_lente:
            st.success(f"✅ Tous les articles ont été vendus dans les {seuil_rot} derniers jours.")
        else:
            st.warning(f"⚠️ **{len(rot_lente)}** article(s) sans vente depuis plus de {seuil_rot} jours")
            for art in rot_lente:
                hx = HEX_MAP.get(art["Couleur"],"#888")
                jours = art["Jours sans vente"]
                color_j = "#CC0000" if jours>=30 else "#FF8C00" if jours>=14 else "#888"
                jours_txt = "Jamais vendu" if jours==999 else f"{jours} jours"
                st.markdown(f"""
                <div class="history-card" style="border-left:3px solid {color_j}">
                    <div style="display:flex;justify-content:space-between;align-items:center">
                        <div>
                            <div class="history-title" style="margin-bottom:4px">{art["Modèle"]}</div>
                            <div style="display:flex;align-items:center;gap:6px">
                                <span style="width:11px;height:11px;border-radius:50%;background:{hx};border:1px solid #ccc;display:inline-block"></span>
                                <span style="font-size:.78rem">{art["Couleur"]}</span>
                                <span class="tag tag-dark">{art["Taille"]}</span>
                                <span class="tag">En stock : {art["En stock"]}</span>
                            </div>
                        </div>
                        <div style="text-align:right">
                            <div style="font-size:1.1rem;font-weight:700;color:{color_j}">{jours_txt}</div>
                            <div style="font-size:.62rem;color:#AAA">Dernière vente : {art["Dernière vente"]}</div>
                        </div>
                    </div>
                </div>""", unsafe_allow_html=True)
            df_rot = pd.DataFrame([{k:v for k,v in a.items() if k!="model_key"} for a in rot_lente])
            st.download_button("📥  Exporter rotation lente (Excel)",
                _df_to_excel(df_rot,"Rotation lente"),
                f"rotation_lente_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("<hr>",unsafe_allow_html=True)
    st.markdown("### 📅 Historique des ventes")
    ventes_all=load_ventes()
    if not ventes_all:
        st.info("Aucune vente enregistrée.")
    else:
        tv=sum(v.get("quantite",1) for v in ventes_all)
        auj=sum(v.get("quantite",1) for v in ventes_all if v.get("date","").startswith(datetime.now().strftime("%d/%m/%Y")))
        m1,m2,m3=st.columns(3)
        m1.metric("Total pièces vendues",tv); m2.metric("Vendues aujourd'hui",auj)
        m3.metric("Nb vendeurs actifs",len(set(v.get("vendeur","") for v in ventes_all)))

        st.markdown("<br>",unsafe_allow_html=True)
        # Filtres (admin voit tout, vendeur voit seulement ses ventes)
        fv1,fv2,fv3,fv4=st.columns(4)
        with fv1:
            if is_admin:
                filt_vendeur=st.selectbox("Vendeur",["Tous"]+list(dict.fromkeys(v.get("vendeur","") for v in ventes_all)),key="fv_v")
            else:
                filt_vendeur=vendeur_actuel
                st.markdown(f'<span class="section-label">Vendeur</span><br><strong>{vendeur_actuel}</strong>',unsafe_allow_html=True)
        with fv2: filt_mod=st.selectbox("Modèle",["Tous"]+list(dict.fromkeys(v["model_name"] for v in ventes_all)),key="fv_m")
        with fv3: filt_col=st.selectbox("Couleur",["Toutes"]+list(dict.fromkeys(v["couleur"] for v in ventes_all)),key="fv_c")
        with fv4: filt_sz=st.selectbox("Taille",["Toutes"]+list(dict.fromkeys(v["taille"] for v in ventes_all)),key="fv_s")

        fv=ventes_all
        if filt_vendeur not in ["Tous",None]: fv=[v for v in fv if v.get("vendeur","")==filt_vendeur]
        if filt_mod!="Tous":   fv=[v for v in fv if v["model_name"]==filt_mod]
        if filt_col!="Toutes": fv=[v for v in fv if v["couleur"]==filt_col]
        if filt_sz!="Toutes":  fv=[v for v in fv if v["taille"]==filt_sz]

        st.markdown(f"**{len(fv)}** vente(s) affichée(s) — **{sum(v.get('quantite',1) for v in fv)}** pièces")
        st.markdown("<br>",unsafe_allow_html=True)

        sd_v=load_stock()
        for idx,v in enumerate(fv):
            hx=HEX_MAP.get(v.get("couleur",""),"#888")
            sa=v.get("stock_avant","—"); sap=v.get("stock_apres","—")
            seuil_v2=sd_v.get(v.get("model_key",""),{}).get("seuil_min",0)
            alerte=f'<span style="color:#CC0000;font-size:.7rem;margin-left:6px">⚠ sous le seuil</span>' if isinstance(sap,int) and sap<seuil_v2 else ""
            vndr_badge=f'<span class="tag" style="background:#E8F5E9;color:#1A7A3C">{v.get("vendeur","")}</span>' if v.get("vendeur") else ""
            note_html=f'<div style="font-size:.72rem;color:#888;font-style:italic;margin-top:3px">{v["note"]}</div>' if v.get("note") else ""
            st.markdown(f"""
            <div class="history-card" style="border-left:3px solid #25a244">
                <div style="display:flex;justify-content:space-between;align-items:center">
                    <div class="history-date">📅 {v.get("date","")}</div>
                    <div style="display:flex;gap:6px;align-items:center">
                        {vndr_badge}
                        <div style="font-size:.72rem;color:#FFF;background:#25a244;padding:2px 10px;letter-spacing:.05em">
                            {v.get("quantite",1)} pièce{'s' if v.get("quantite",1)>1 else ''}
                        </div>
                    </div>
                </div>
                <div class="history-title">{v.get("model_name","")}</div>
                <div style="display:flex;align-items:center;gap:8px;margin-top:4px">
                    <span style="width:12px;height:12px;border-radius:50%;background:{hx};border:1px solid #ccc;display:inline-block"></span>
                    <span style="font-size:.8rem">{v.get("couleur","")}</span>
                    <span class="tag tag-dark">{v.get("taille","")}</span>
                </div>
                <div style="font-size:.72rem;color:#888;margin-top:4px">Stock: {sa} → {sap}{alerte}</div>
                {note_html}
            </div>""", unsafe_allow_html=True)
            if is_admin and st.button("Supprimer",key=f"dv_{idx}"):
                va=load_ventes()
                if v in va: va.remove(v); save_ventes(va); st.rerun()

        st.markdown("<hr>",unsafe_allow_html=True)
        ev1,ev2=st.columns(2)
        with ev1: st.download_button("📥 Excel",build_ventes_excel(fv),f"ventes_{datetime.now().strftime('%Y%m%d')}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
        with ev2: st.download_button("📄 CSV",pd.DataFrame([{"Date":v.get("date"),"Vendeur":v.get("vendeur",""),"Modèle":v.get("model_name"),"Couleur":v.get("couleur"),"Taille":v.get("taille"),"Qté":v.get("quantite",1),"Note":v.get("note","")} for v in fv]).to_csv(index=False).encode("utf-8-sig"),f"ventes_{datetime.now().strftime('%Y%m%d')}.csv","text/csv",use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET ALIMENTATION ADMIN
# ══════════════════════════════════════════════════════════════════════════════
if tab_ali_admin:
    with tab_ali_admin:
        pending_list = load_ali_pending()
        hist_list    = load_ali_hist()
        stock_ap     = load_stock()

        sub_pending, sub_hist = st.tabs(["📋 Demandes en attente", "📜 Historique"])

        # ── Sous-onglet 1 : Demandes en attente ─────────────────────────────
        with sub_pending:
            nb_p = len(pending_list)
            if nb_p == 0:
                st.info("✅ Aucune demande d'alimentation en attente.")
            else:
                st.markdown(f'<span class="section-label">{nb_p} demande(s) en attente</span>', unsafe_allow_html=True)
                for pi, req in enumerate(pending_list):
                    with st.expander(f"📦 {req.get('model_name','?')} — {req.get('livreur','?')} — {req.get('date','?')}", expanded=(pi==0)):
                        # Affichage recap
                        rc1, rc2 = st.columns([2,1])
                        with rc1:
                            st.markdown(f"**Article :** {req.get('model_name','?')}")
                            st.markdown(f"**Catégorie :** {req.get('categorie','?')}")
                            st.markdown(f"**Livreur :** {req.get('livreur','?')}")
                            st.markdown(f"**Date :** {req.get('date','?')}")
                            st.markdown("**Quantités proposées :**")
                            for k, v in req.get("qtys",{}).items():
                                color_q, size_q = k.split("__") if "__" in k else (k, "")
                                hx_q = HEX_MAP.get(color_q,"#888")
                                st.markdown(
                                    f'<div style="display:inline-flex;align-items:center;gap:.4rem;'
                                    f'background:#F4F6FA;border-radius:8px;padding:.3rem .7rem;margin:.2rem">'
                                    f'<span style="width:10px;height:10px;border-radius:50%;background:{hx_q};display:inline-block"></span>'
                                    f'<span>{color_q}</span><strong>{size_q}</strong><span style="color:#059669;font-weight:700">+{v}</span></div>',
                                    unsafe_allow_html=True)

                        with rc2:
                            mk_req = req.get("model_key","")
                            b64_req = stock_ap.get(mk_req,{}).get("b64_thumb","")
                            if b64_req:
                                st.markdown(f'<img src="data:image/png;base64,{b64_req}" style="width:100%;border-radius:10px">', unsafe_allow_html=True)

                        # Modifier les quantités avant approbation
                        if st.session_state.ali_edit_idx == pi:
                            st.markdown("**✏️ Modifier les quantités :**")
                            new_qtys = {}
                            for k, v in req.get("qtys",{}).items():
                                color_q, size_q = k.split("__") if "__" in k else (k, "")
                                new_v = st.number_input(f"{color_q} / {size_q}", min_value=0, max_value=500,
                                                        value=v, key=f"edit_ali_{pi}_{k}")
                                new_qtys[k] = new_v
                            sav1, sav2 = st.columns(2)
                            with sav1:
                                if st.button("💾 Sauvegarder modifications", key=f"save_ali_edit_{pi}", use_container_width=True):
                                    pending_list[pi]["qtys"] = {k:v for k,v in new_qtys.items() if v>0}
                                    save_ali_pending(pending_list)
                                    st.session_state.ali_edit_idx = None
                                    st.rerun()
                            with sav2:
                                if st.button("✕ Annuler", key=f"cancel_ali_edit_{pi}", use_container_width=True):
                                    st.session_state.ali_edit_idx = None
                                    st.rerun()
                        else:
                            b1, b2, b3 = st.columns(3)
                            with b1:
                                if st.button("✏️ Modifier", key=f"ali_edit_{pi}", use_container_width=True):
                                    st.session_state.ali_edit_idx = pi
                                    st.rerun()
                            with b2:
                                if st.button("✅ Approuver", key=f"ali_approve_{pi}", use_container_width=True):
                                    # Mettre à jour le stock
                                    mk_ap = req.get("model_key","")
                                    for k, qty_ap in req.get("qtys",{}).items():
                                        color_ap, size_ap = k.split("__") if "__" in k else (k,"")
                                        stock_ap.setdefault(mk_ap,{}).setdefault("stock",{}).setdefault(color_ap,{})
                                        stock_ap[mk_ap]["stock"][color_ap][size_ap] = \
                                            stock_ap[mk_ap]["stock"][color_ap].get(size_ap,0) + qty_ap
                                    save_stock(stock_ap)
                                    # Historique
                                    req["statut"] = "approuve"
                                    req["date_approbation"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                                    hist_list.insert(0, req)
                                    save_ali_hist(hist_list)
                                    pending_list.pop(pi)
                                    save_ali_pending(pending_list)
                                    st.success(f"✅ Stock mis à jour pour {req.get('model_name','?')} !")
                                    st.rerun()
                            with b3:
                                if st.button("❌ Rejeter", key=f"ali_reject_{pi}", use_container_width=True):
                                    req["statut"] = "rejete"
                                    req["date_approbation"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                                    hist_list.insert(0, req)
                                    save_ali_hist(hist_list)
                                    pending_list.pop(pi)
                                    save_ali_pending(pending_list)
                                    st.warning("❌ Demande rejetée.")
                                    st.rerun()

        # ── Sous-onglet 2 : Historique ───────────────────────────────────────
        with sub_hist:
            import calendar as _cal2
            now_ali_cal = datetime.now()
            if st.session_state.ali_admin_month is None:
                st.session_state.ali_admin_month = (now_ali_cal.year, now_ali_cal.month)
            yr_a, mo_a = st.session_state.ali_admin_month

            # Navigation mois
            an1, an2, an3 = st.columns([1, 3, 1])
            with an1:
                if st.button("◀ Mois précédent", key="ali_cal_prev", use_container_width=True):
                    mo2a = mo_a - 1
                    yr2a = yr_a if mo2a > 0 else yr_a - 1
                    mo2a = mo2a if mo2a > 0 else 12
                    st.session_state.ali_admin_month = (yr2a, mo2a)
                    st.session_state.ali_admin_date = None
                    st.rerun()
            with an2:
                st.markdown(
                    f'<div style="text-align:center;font-family:Space Grotesk,sans-serif;'
                    f'font-size:1.2rem;font-weight:700;color:#1B2B4B;padding:.5rem">'
                    f'{_cal2.month_name[mo_a].capitalize()} {yr_a}</div>',
                    unsafe_allow_html=True)
            with an3:
                if st.button("Mois suivant ▶", key="ali_cal_next", use_container_width=True):
                    mo2a = mo_a + 1
                    yr2a = yr_a if mo2a <= 12 else yr_a + 1
                    mo2a = mo2a if mo2a <= 12 else 1
                    st.session_state.ali_admin_month = (yr2a, mo2a)
                    st.session_state.ali_admin_date = None
                    st.rerun()

            # Index alimentation par jour
            ali_par_jour = {}
            for h in hist_list:
                ds_a = h.get("date","")[:10]
                try:
                    dobj_a = datetime.strptime(ds_a, "%d/%m/%Y")
                    if dobj_a.year == yr_a and dobj_a.month == mo_a:
                        ali_par_jour.setdefault(ds_a, []).append(h)
                except: pass

            # Calendrier
            jours_s = ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]
            hdr_a = '<div style="display:grid;grid-template-columns:repeat(7,1fr);gap:4px;margin-bottom:4px">'
            for j in jours_s:
                hdr_a += f'<div style="text-align:center;font-size:.7rem;font-weight:700;color:#8A9AB5;padding:.3rem">{j}</div>'
            hdr_a += '</div>'
            st.markdown(hdr_a, unsafe_allow_html=True)

            cal_mat_a = _cal2.monthcalendar(yr_a, mo_a)
            sel_a = st.session_state.ali_admin_date
            clicked_day_a = None

            for week_a in cal_mat_a:
                wk_cols_a = st.columns(7)
                for wi_a, day_a in enumerate(week_a):
                    with wk_cols_a[wi_a]:
                        if day_a == 0:
                            st.markdown('<div style="height:52px"></div>', unsafe_allow_html=True)
                            continue
                        day_str_a = f"{day_a:02d}/{mo_a:02d}/{yr_a}"
                        nb_a = len(ali_par_jour.get(day_str_a, []))
                        is_today_a = (day_a == now_ali_cal.day and mo_a == now_ali_cal.month and yr_a == now_ali_cal.year)
                        is_sel_a   = (day_str_a == sel_a)
                        border_a = "3px solid #059669" if is_sel_a else ("2px solid #1B2B4B" if is_today_a else "1px solid #E0E5EF")
                        bg_a     = "#F0FDF4" if is_sel_a else ("#F4F6FA" if is_today_a else "#FFF")
                        badge_a  = f'<div style="font-size:.6rem;color:#059669;font-weight:700">{nb_a}ali</div>' if nb_a > 0 else ""
                        dot_a    = '<div style="width:8px;height:8px;border-radius:50%;background:#059669;margin:2px auto 0"></div>' if nb_a > 0 else ""
                        st.markdown(
                            f'<div style="background:{bg_a};border:{border_a};border-radius:10px;'
                            f'padding:.4rem .2rem;text-align:center;font-size:.85rem;font-weight:600;'
                            f'min-height:58px;display:flex;flex-direction:column;'
                            f'align-items:center;justify-content:center;gap:2px">'
                            f'<span style="font-weight:{"700" if is_today_a else "500"};color:#1B2B4B">{day_a}</span>'
                            f'{"<span style=\"font-size:.6rem;font-weight:700;color:#059669\">● "+str(nb_a)+" ali</span>" if nb_a>0 else ""}'
                            f'</div>', unsafe_allow_html=True)
                        if st.button(f"{day_a}", key=f"ali_day_{yr_a}_{mo_a}_{day_a}",
                                     use_container_width=True,
                                     help=f"{'Cliquez — '+str(nb_a)+' alimentation(s)' if nb_a>0 else 'Aucune alimentation'}"):
                            clicked_day_a = day_str_a

            if clicked_day_a:
                st.session_state.ali_admin_date = clicked_day_a
                st.rerun()

            # ── Détail du jour sélectionné ───────────────────────────────────
            if sel_a:
                ali_jour = ali_par_jour.get(sel_a, [])
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown(
                    f'<div style="font-family:Space Grotesk,sans-serif;font-size:1.1rem;'
                    f'font-weight:700;color:#1B2B4B;margin-bottom:1rem">📦 Alimentations — {sel_a}</div>',
                    unsafe_allow_html=True)
                if not ali_jour:
                    st.info("Aucune alimentation ce jour.")
                else:
                    for h in ali_jour:
                        statut_h = h.get("statut","?")
                        color_h  = "#059669" if statut_h=="approuve" else "#DC2626"
                        badge_h  = "✅ Approuvé" if statut_h=="approuve" else "❌ Rejeté"
                        total_h  = sum(h.get("qtys",{}).values())
                        st.markdown(
                            f'<div style="background:#FFF;border:1px solid #E0E5EF;border-radius:12px;'
                            f'padding:1rem 1.2rem;margin-bottom:.6rem;display:flex;justify-content:space-between;align-items:center">'
                            f'<div><strong style="color:#1B2B4B">{h.get("model_name","?")}</strong> '
                            f'<span style="font-size:.75rem;color:#8A9AB5">par {h.get("livreur","?")} — {h.get("date","?")}</span><br>'
                            f'<span style="font-size:.78rem;color:#4A5568">+{total_h} pièce(s) · {h.get("categorie","?")}</span></div>'
                            f'<span style="background:{color_h}22;color:{color_h};border:1px solid {color_h}44;'
                            f'font-size:.65rem;font-weight:700;padding:3px 10px;border-radius:100px">{badge_h}</span></div>',
                            unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# ONGLET 5 : GESTION DES PROFILS (admin seulement)
# ══════════════════════════════════════════════════════════════════════════════
if tab_admin:
    with tab_admin:
        profils_data=load_profils()
        vendeurs=profils_data.get("vendeurs",[])

        # ════════════════════════════════════════════════════════
        # GESTION DES CATÉGORIES
        # ════════════════════════════════════════════════════════
        st.markdown("### 🗂️ Gestion des catégories")
        st.caption("Créez vos catégories ici. Elles seront disponibles lors de l'ajout d'articles.")
        st.markdown("<hr>", unsafe_allow_html=True)

        cats_admin = load_categories()

        # ── Formulaire ajout ─────────────────────────────────────────────
        st.markdown('<span class="section-label">Nouvelle catégorie</span>', unsafe_allow_html=True)
        ca1, ca2, ca3, ca4 = st.columns([3, 1, 1, 1])
        with ca1:
            new_cat_name = st.text_input("Nom de la catégorie",
                placeholder="Ex: T-shirts Homme, Polos Femme, Accessoires…",
                label_visibility="visible", key="new_cat_name")
        with ca2:
            new_cat_icon = st.text_input("Icône (emoji)",
                value="👕", max_chars=4,
                key="new_cat_icon", label_visibility="visible")
        with ca3:
            CAT_COLORS = ["#D4A843","#4F8EF7","#2DD4A0","#F76B6B","#C084FC","#FB923C","#34D399"]
            new_cat_color = st.selectbox("Couleur badge",
                options=CAT_COLORS,
                format_func=lambda x: {"#D4A843":"🟡 Or","#4F8EF7":"🔵 Bleu",
                    "#2DD4A0":"🟢 Vert","#F76B6B":"🔴 Rouge","#C084FC":"🟣 Violet",
                    "#FB923C":"🟠 Orange","#34D399":"💚 Émeraude"}.get(x,x),
                key="new_cat_color")
        with ca4:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button("➕  Ajouter", key="btn_add_cat", use_container_width=True):
                n = new_cat_name.strip()
                existing_names = [(c.get("name",c) if isinstance(c,dict) else c) for c in cats_admin]
                if not n:
                    st.warning("Saisissez un nom de catégorie.")
                elif n in existing_names:
                    st.warning(f"La catégorie **{n}** existe déjà.")
                else:
                    cats_admin.append({"name": n, "icon": new_cat_icon or "📦",
                                       "color": new_cat_color})
                    save_categories(cats_admin)
                    st.success(f"✓ Catégorie **{n}** créée !")
                    st.rerun()

        # ── Liste des catégories ──────────────────────────────────────────
        if not cats_admin:
            st.info("Aucune catégorie. Commencez par en créer une ci-dessus.")
        else:
            st.markdown(
                f'<span class="section-label" style="margin-top:1rem">'
                f'{len(cats_admin)} catégorie(s)</span>', unsafe_allow_html=True)

            for ci, cat in enumerate(cats_admin):
                cat_n = cat.get("name", cat) if isinstance(cat, dict) else cat
                cat_i = cat.get("icon", "📦") if isinstance(cat, dict) else "📦"
                cat_c = cat.get("color", "#D4A843") if isinstance(cat, dict) else "#D4A843"
                # Compte les articles de cette catégorie
                nb_art = sum(1 for e in load_history()
                             if (e.get("categorie","") or "Non classé") == cat_n)
                cc1, cc2, cc3 = st.columns([.08, .7, .22])
                with cc1:
                    st.markdown(f'<div style="font-size:1.8rem;line-height:1;'
                                f'padding-top:.3rem">{cat_i}</div>', unsafe_allow_html=True)
                with cc2:
                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:.8rem;padding:.5rem 0">'
                        f'<strong style="color:var(--text);font-size:.95rem">{cat_n}</strong>'
                        f'<span style="background:{cat_c}22;border:1px solid {cat_c};'
                        f'color:{cat_c};font-size:.62rem;padding:2px 8px;border-radius:100px;'
                        f'font-weight:600">{nb_art} article(s)</span></div>',
                        unsafe_allow_html=True)
                with cc3:
                    ec1, ec2 = st.columns(2)
                    with ec1:
                        if st.button("✏️", key=f"edit_cat_{ci}", use_container_width=True,
                                     help="Modifier cette catégorie"):
                            st.session_state.edit_cat_idx = ci
                            st.rerun()
                    with ec2:
                        if st.button("🗑", key=f"del_cat_{ci}", use_container_width=True,
                                     help="Supprimer"):
                            cats_admin.pop(ci)
                            save_categories(cats_admin)
                            st.rerun()

                # Formulaire d'édition inline
                if st.session_state.edit_cat_idx == ci:
                    with st.form(key=f"form_edit_cat_{ci}"):
                        st.markdown(f"**✏️ Modifier : {cat_n}**")
                        ef1, ef2, ef3 = st.columns([3,1,1])
                        with ef1: new_cn = st.text_input("Nom", value=cat_n, key=f"ecn_{ci}")
                        with ef2: new_ci = st.text_input("Icône", value=cat_i, max_chars=4, key=f"eci_{ci}")
                        with ef3:
                            CAT_COLORS2 = ["#D4A843","#4F8EF7","#2DD4A0","#F76B6B","#C084FC","#FB923C","#34D399"]
                            new_cc = st.selectbox("Couleur", CAT_COLORS2,
                                index=CAT_COLORS2.index(cat_c) if cat_c in CAT_COLORS2 else 0,
                                format_func=lambda x: {"#D4A843":"🟡","#4F8EF7":"🔵","#2DD4A0":"🟢",
                                    "#F76B6B":"🔴","#C084FC":"🟣","#FB923C":"🟠","#34D399":"💚"}.get(x,x),
                                key=f"ecc_{ci}")
                        sb1, sb2 = st.columns(2)
                        with sb1: save_ok = st.form_submit_button("💾 Sauvegarder", use_container_width=True)
                        with sb2: cancel_ok = st.form_submit_button("✕ Annuler", use_container_width=True)
                    if save_ok:
                        cats_admin[ci] = {"name": new_cn.strip() or cat_n, "icon": new_ci or cat_i, "color": new_cc}
                        save_categories(cats_admin)
                        st.session_state.edit_cat_idx = None
                        st.rerun()
                    if cancel_ok:
                        st.session_state.edit_cat_idx = None
                        st.rerun()

        st.markdown("<hr>", unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════
        st.markdown("### 👥 Gestion des vendeurs")
        st.markdown("<hr>",unsafe_allow_html=True)

        # ── Ajouter un vendeur ──
        st.markdown('<span class="section-label">Ajouter un vendeur</span>',unsafe_allow_html=True)
        na1,na2=st.columns([3,1])
        with na1: new_name=st.text_input("",placeholder="Ex: Ahmed, Sara Dupont…",label_visibility="collapsed",key="new_vend")
        with na2:
            if st.button("➕  Ajouter",key="add_vend"):
                n=new_name.strip()
                if n and n not in vendeurs:
                    vendeurs.append(n); profils_data["vendeurs"]=vendeurs
                    save_profils(profils_data); st.success(f"✓ Vendeur **{n}** ajouté"); st.rerun()
                elif n in vendeurs: st.warning(f"**{n}** existe déjà.")
                else: st.warning("Saisissez un nom.")

        # ── Liste des vendeurs ──
        if not vendeurs:
            st.info("Aucun vendeur configuré.")
        else:
            st.markdown(f'<span class="section-label" style="margin-top:1rem">{len(vendeurs)} vendeur(s) actif(s)</span>',unsafe_allow_html=True)
            for i,nom in enumerate(vendeurs):
                vc1,vc2,vc3=st.columns([.6,.1,.3])
                ini="".join(p[0].upper() for p in nom.split()[:2])
                with vc1:
                    nb_v=sum(1 for v in load_ventes() if v.get("vendeur","")==nom)
                    st.markdown(f'<div style="display:flex;align-items:center;gap:10px;padding:8px 0">'
                                f'<div class="profil-avatar" style="width:36px;height:36px;font-size:.85rem;min-width:36px">{ini}</div>'
                                f'<div><strong>{nom}</strong><br><span style="font-size:.68rem;color:#888">{nb_v} vente(s) enregistrée(s)</span></div></div>',unsafe_allow_html=True)
                with vc2:
                    if st.button("✏️", key=f"edit_v_{i}", help="Modifier le nom"):
                        st.session_state.edit_vend_idx = i
                        st.rerun()
                with vc3:
                    if st.button("🗑", key=f"del_v_{i}", help="Supprimer"):
                        vendeurs.remove(nom); profils_data["vendeurs"]=vendeurs
                        save_profils(profils_data); st.rerun()

                if st.session_state.edit_vend_idx == i:
                    with st.form(key=f"form_edit_v_{i}"):
                        new_vnom = st.text_input("Nouveau nom", value=nom, key=f"evn_{i}")
                        vs1, vs2 = st.columns(2)
                        with vs1: vsave = st.form_submit_button("💾 Sauvegarder", use_container_width=True)
                        with vs2: vcancel = st.form_submit_button("✕ Annuler", use_container_width=True)
                    if vsave and new_vnom.strip():
                        vendeurs[i] = new_vnom.strip()
                        profils_data["vendeurs"] = vendeurs
                        save_profils(profils_data)
                        st.session_state.edit_vend_idx = None
                        st.rerun()
                    if vcancel:
                        st.session_state.edit_vend_idx = None
                        st.rerun()

        st.markdown("<hr>",unsafe_allow_html=True)

        # ── Changer le PIN admin ──
        st.markdown("### 🔐 Changer le code PIN admin")
        pp1,pp2,pp3=st.columns(3)
        with pp1: old_pin=st.text_input("PIN actuel",type="password",key="old_pin")
        with pp2: new_pin=st.text_input("Nouveau PIN",type="password",key="new_pin")
        with pp3: new_pin2=st.text_input("Confirmer",type="password",key="new_pin2")
        if st.button("💾 Changer le PIN",key="chg_pin"):
            if old_pin!=profils_data.get("admin_pin","1234"): st.error("PIN actuel incorrect.")
            elif len(new_pin)<4: st.error("Le PIN doit faire au moins 4 caractères.")
            elif new_pin!=new_pin2: st.error("Les deux PINs ne correspondent pas.")
            else:
                profils_data["admin_pin"]=new_pin; save_profils(profils_data)
                st.success("✓ PIN modifié avec succès !")

        st.markdown("<hr>",unsafe_allow_html=True)

        # ── PIN Alimentation ──────────────────────────────────────────────────
        st.markdown("### 📦 Code PIN Alimentation")
        st.caption("Ce code est donné aux livreurs pour accéder à l'espace Alimentation Stock.")
        current_ali_pin = profils_data.get("alimentation_pin","")
        ap1, ap2, ap3 = st.columns(3)
        with ap1: new_ali_pin = st.text_input("Nouveau PIN alimentation", type="password", placeholder="• • • •", key="new_ali_pin")
        with ap2: new_ali_pin2 = st.text_input("Confirmer", type="password", placeholder="• • • •", key="new_ali_pin2")
        with ap3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button("💾 Définir le PIN", key="set_ali_pin", use_container_width=True):
                if len(new_ali_pin) < 4: st.error("Le PIN doit faire au moins 4 caractères.")
                elif new_ali_pin != new_ali_pin2: st.error("Les deux PINs ne correspondent pas.")
                else:
                    profils_data["alimentation_pin"] = new_ali_pin
                    save_profils(profils_data)
                    st.success("✓ PIN alimentation défini !")
                    st.rerun()

        if current_ali_pin:
            st.markdown(f'<div style="background:#F0FDF4;border:1px solid #059669;border-radius:10px;padding:.8rem 1rem;margin:.8rem 0;font-size:.85rem;color:#065F46">PIN alimentation actuel : <strong>{"•"*len(current_ali_pin)}</strong> ({len(current_ali_pin)} caractères)</div>', unsafe_allow_html=True)
            wa_pin_msg = f"Bonjour,\n\nVoici votre code d'accès pour l'alimentation du stock Latif Shop :\n\n🔐 Code PIN : *{current_ali_pin}*\n\n📱 Accédez via : https://latifshop-teeshirt.streamlit.app\n\nChoisissez \"📦 Alimentation Stock\" sur la page d'accueil."
            wa_pin_url = f"https://wa.me/?text={urllib.parse.quote(wa_pin_msg)}"
            st.markdown(
                f'<a href="{wa_pin_url}" target="_blank">'
                f'<button style="background:#25D366;color:#FFF;border:none;border-radius:10px;'
                f'padding:.65rem 1.5rem;font-size:.85rem;font-weight:700;cursor:pointer">'
                f'📱 Partager le PIN par WhatsApp</button></a>',
                unsafe_allow_html=True)

        st.markdown("<hr>",unsafe_allow_html=True)

        # ── Mode hors-ligne : sauvegarde & restauration ──────────────────
        st.markdown("### 💾 Sauvegarde & Mode hors-ligne")
        st.caption("Exportez toutes vos données en un fichier ZIP. Restaurez-les sur n'importe quel appareil.")

        nb1,nb2=st.columns(2)
        with nb1:
            st.markdown("**📤 Exporter toutes les données**")
            nb_ventes=len(load_ventes()); nb_models=len(load_history()); nb_arr=len(load_arrivages())
            st.markdown(f'<div style="background:#F5F5F5;border:1px solid #E0E0E0;padding:10px 14px;font-size:.8rem;margin-bottom:8px">'
                        f'📊 {nb_models} modèle(s) · 💰 {nb_ventes} vente(s) · 📦 {nb_arr} arrivage(s)</div>',
                        unsafe_allow_html=True)
            st.download_button(
                "📥  Télécharger la sauvegarde (.zip)",
                export_backup_zip(),
                f"sauvegarde_inventaire_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                "application/zip", use_container_width=True)

        with nb2:
            st.markdown("**📥 Restaurer depuis une sauvegarde**")
            st.caption("⚠️ Attention : écrase les données actuelles.")
            restore_file=st.file_uploader("",type=["zip"],label_visibility="collapsed",key="restore_zip")
            if restore_file:
                if st.button("🔄  Restaurer maintenant",key="do_restore"):
                    try:
                        import_backup_zip(restore_file.read())
                        st.success("✓ Données restaurées avec succès !"); st.rerun()
                    except Exception as ex:
                        st.error(f"Erreur lors de la restauration : {ex}")

        st.markdown("""
        <div style="background:#EEF4FF;border:1px solid #C3D8FF;padding:12px 16px;font-size:.78rem;color:#1a3a6e;margin-top:8px">
            💡 <strong>Mode hors-ligne :</strong> L'application fonctionne sans internet lorsqu'elle est lancée
            localement (<code>streamlit run app.py</code>). Sur Streamlit Cloud, téléchargez régulièrement
            la sauvegarde ZIP pour ne pas perdre vos données.
        </div>
        """, unsafe_allow_html=True)
